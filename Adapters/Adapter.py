# -*- coding: utf-8 -*-
from __future__ import annotations

import os
import sys
import json
import xml.etree.ElementTree as ET
from dataclasses import dataclass, field
from typing import Any, Dict, List, Optional, Tuple, Set

from PySide6 import QtCore, QtGui, QtWidgets

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from theme_toggle import (
    ThemeToggle, theme, is_dark_theme, create_back_button, go_to_main_menu,
    resolve_icon_path, apply_dark_titlebar,
    load_saved_theme, enable_theme_sync,
    RowHoverDelegate, install_viewport_row_highlighter, setup_hover_tracking,
)

# ----------------- Тема и логотип -----------------
BG = "#FFFFFF"
FG = "#222222"
ACCENT_ORANGE = "#F7921E"
ACCENT_ORANGE_HOVER = "#FFA74B"
BTN_GRAY = "#D9D9D9"
BTN_GRAY_HOVER = "#C9C9C9"

def _plugins_dir() -> str:
    # ...\Plugins\Manager_Adapter\Adapters\Adapter.py -> ...\Plugins
    return os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))


def _resolve_logo_path() -> str:
    candidates = [
        # legacy absolute path (might exist on some machines)
        r"C:\Users\dviktorov\Desktop\Project\Manager\icon\Manager-scaled.png",
        # from Nexus Desktop-manager (design reference mentioned by user)
        os.path.join(_plugins_dir(), "Nexus", "Dekstop-manager", "icon", "Manager-scaled.png"),
        # common local fallbacks (if icon gets copied into this plugin later)
        os.path.join(os.path.dirname(__file__), "icon", "Manager-scaled.png"),
        os.path.join(os.path.dirname(__file__), "assets", "Manager-scaled.png"),
    ]
    for p in candidates:
        try:
            if p and os.path.exists(p):
                return p
        except Exception:
            continue
    return ""


LOGO_PATH = _resolve_logo_path()
ICON_DIR = os.path.join(os.path.dirname(__file__), "icon")
DEFAULT_BASE_URL = "http://localhost:5000"

_ICON_FALLBACK_FILES = {
    "ok": "ok.png",
    "circle": "circle.png",
    "check": "check.png",
    "select": "select.png",
    "poloska": "poloska.png",
}


def _resolve_icon_path(filename: str) -> str:
    if not filename:
        return ""
    candidates = [
        os.path.join(_plugins_dir(), "Nexus", "Dekstop-manager", "icon", filename),
        os.path.join(_plugins_dir(), "Nexus", "Dekstop-manager", "icon", "white", filename),
        os.path.join(os.path.dirname(__file__), "icon", filename),
        os.path.join(os.path.dirname(__file__), "assets", filename),
    ]
    for p in candidates:
        try:
            if p and os.path.exists(p):
                return p
        except Exception:
            continue
    return ""


def _pm(size: int) -> QtGui.QPixmap:
    pm = QtGui.QPixmap(size, size)
    pm.fill(QtCore.Qt.transparent)
    return pm


def _icon_ok(size: int = 16, color: str = ACCENT_ORANGE) -> QtGui.QIcon:
    pm = _pm(size)
    p = QtGui.QPainter(pm)
    try:
        p.setRenderHint(QtGui.QPainter.Antialiasing, True)
        pen = QtGui.QPen(QtGui.QColor(color), max(2, size // 8))
        pen.setCapStyle(QtCore.Qt.RoundCap)
        pen.setJoinStyle(QtCore.Qt.RoundJoin)
        p.setPen(pen)
        # simple check mark
        p.drawLine(int(size * 0.20), int(size * 0.55), int(size * 0.42), int(size * 0.75))
        p.drawLine(int(size * 0.42), int(size * 0.75), int(size * 0.80), int(size * 0.28))
    finally:
        p.end()
    return QtGui.QIcon(pm)


def _icon_circle(size: int = 16, border_color: str = "#B0B0B0") -> QtGui.QIcon:
    pm = _pm(size)
    p = QtGui.QPainter(pm)
    try:
        p.setRenderHint(QtGui.QPainter.Antialiasing, True)
        pen = QtGui.QPen(QtGui.QColor(border_color), max(2, size // 9))
        pen.setCapStyle(QtCore.Qt.RoundCap)
        p.setPen(pen)
        p.setBrush(QtCore.Qt.NoBrush)
        pad = max(2, size // 6)
        p.drawEllipse(pad, pad, size - 2 * pad, size - 2 * pad)
    finally:
        p.end()
    return QtGui.QIcon(pm)


def _icon_arrow(direction: str, size: int = 18, color: str = "#333333") -> QtGui.QIcon:
    pm = _pm(size)
    p = QtGui.QPainter(pm)
    try:
        p.setRenderHint(QtGui.QPainter.Antialiasing, True)
        pen = QtGui.QPen(QtGui.QColor(color), max(2, size // 9))
        pen.setCapStyle(QtCore.Qt.RoundCap)
        pen.setJoinStyle(QtCore.Qt.RoundJoin)
        p.setPen(pen)
        cx, cy = size / 2.0, size / 2.0
        r = size * 0.28
        if direction == "up":
            a = QtCore.QPointF(cx, cy - r)
            b = QtCore.QPointF(cx - r, cy + r)
            c = QtCore.QPointF(cx + r, cy + r)
        else:
            a = QtCore.QPointF(cx, cy + r)
            b = QtCore.QPointF(cx - r, cy - r)
            c = QtCore.QPointF(cx + r, cy - r)
        p.setBrush(QtGui.QBrush(QtGui.QColor(color)))
        p.drawPolygon(QtGui.QPolygonF([a, b, c]))
    finally:
        p.end()
    return QtGui.QIcon(pm)

def _icon_checkbox(checked: bool, size: int = 18, color: str = "#333333") -> QtGui.QIcon:
    pm = _pm(size)
    p = QtGui.QPainter(pm)
    try:
        p.setRenderHint(QtGui.QPainter.Antialiasing, True)
        pen = QtGui.QPen(QtGui.QColor(color), max(2, size // 10))
        pen.setCapStyle(QtCore.Qt.RoundCap)
        pen.setJoinStyle(QtCore.Qt.RoundJoin)
        p.setPen(pen)
        pad = max(2, size // 6)
        rect = QtCore.QRectF(pad, pad, size - 2 * pad, size - 2 * pad)
        p.setBrush(QtCore.Qt.NoBrush)
        p.drawRoundedRect(rect, 3, 3)
        if checked:
            p.drawLine(QtCore.QPointF(size * 0.25, size * 0.52), QtCore.QPointF(size * 0.44, size * 0.70))
            p.drawLine(QtCore.QPointF(size * 0.44, size * 0.70), QtCore.QPointF(size * 0.77, size * 0.30))
    finally:
        p.end()
    return QtGui.QIcon(pm)

# ----------------- HTTP helpers -------------------
try:
    import requests
except Exception:
    requests = None

def _need_requests():
    if requests is None:
        raise RuntimeError("Не установлен requests. Установи: pip install requests")

def _http_get(url: str, params=None, timeout=60):
    _need_requests()
    r = requests.get(url, headers={"accept": "application/json"}, params=params, timeout=timeout)
    if r.status_code >= 400:
        raise RuntimeError(f"GET {url} -> {r.status_code}\n{r.text}")
    try:
        return r.json()
    except Exception:
        return json.loads(r.text or "null")

def api_get_projects(base_url: str) -> List[Dict[str, Any]]:
    data = _http_get(f"{base_url.rstrip('/')}/api/project/projects") or []
    return [{"id": p.get("id"), "title": p.get("title") or p.get("name") or f"ID {p.get('id')}"} for p in data]

def api_get_containers(base_url: str, project_id: int) -> List[Dict[str, Any]]:
    data = _http_get(f"{base_url.rstrip('/')}/api/imcContainer/getProjectImcContainers/{project_id}") or []
    return [{"id": c.get("id"), "title": c.get("title") or f"ID {c.get('id')}"} for c in data]

def api_get_params_for_container(base_url: str, container_id: int) -> List[Dict[str, Any]]:
    data = _http_get(
        f"{base_url.rstrip('/')}/api/imcParameterDefinition/imcParameterDefinitions",
        params=[("containerIds", int(container_id))]
    ) or []
    rows = []
    for r in data:
        rows.append({
            "code": (r.get("code") or "").strip(),
            "isNumeric": bool(r.get("isNumeric")),
        })
    return rows

def api_get_global_component(base_url: str, comp_type: int = 1) -> dict:
    url = f"{base_url.rstrip('/')}/api/globalComponent/globalComponent/{int(comp_type)}"
    return _http_get(url) or {}

def flatten_global_attributes_with_types(gc: dict) -> List[Tuple[str, Optional[bool]]]:
    out: List[Tuple[str, Optional[bool]]] = []

    def walk_section(sec: dict, prefix: str = ""):
        name = (sec.get("Name") or "").strip()
        path = f"{prefix}.{name}" if prefix and name else (name or prefix)
        for a in (sec.get("Attributes") or []):
            nm = (a.get("Name") or "").strip()
            if not nm:
                continue
            isnum = a.get("IsNumeric")
            full = f"{path}.{nm}" if path else nm
            out.append((full, bool(isnum) if isnum is not None else None))
        for s in (sec.get("Sections") or []):
            walk_section(s, path)

    content = (gc.get("content") or {})
    for a in (content.get("Attributes") or []):
        nm = (a.get("Name") or "").strip()
        if nm:
            isnum = a.get("IsNumeric")
            out.append((nm, bool(isnum) if isnum is not None else None))
    for sec in (content.get("Sections") or []):
        walk_section(sec, "")
    # уникализация
    seen = {}
    for name, t in out:
        if name not in seen:
            seen[name] = t
    return sorted(seen.items(), key=lambda x: x[0].lower())

# ----------------- Модель адаптера -----------------
def _b(v: bool) -> str:
    return "true" if v else "false"

@dataclass
class TransformSettings:
    trim: str = ""               # "", "Left", "Right", "Both"
    case: str = ""               # "", "Upper", "Lower", "Title"
    prepare_numeric: bool = False
    replaces: List[Tuple[str, str]] = field(default_factory=list)

def _next_uid():
    _next_uid._c = getattr(_next_uid, '_c', 0) + 1
    return _next_uid._c

@dataclass
class Binding:
    parameter_code: str
    uid: int = field(default_factory=_next_uid)
    src_is_numeric: Optional[bool] = None
    is_enabled: bool = True
    src_model_title: str = ""
    transform: TransformSettings = field(default_factory=TransformSettings)

@dataclass
class BindingQueue:
    attribute_full_name: str
    default_is_enabled: bool = True
    bindings: List[Binding] = field(default_factory=list)

class AdapterDoc:
    def __init__(self):
        self.queues: Dict[str, BindingQueue] = {}

    def ensure_attr(self, name: str) -> BindingQueue:
        k = (name or "").strip()
        if not k:
            raise ValueError("Имя атрибута пустое")
        if k not in self.queues:
            self.queues[k] = BindingQueue(attribute_full_name=k)
        return self.queues[k]

    def list_attrs(self) -> List[BindingQueue]:
        return [self.queues[k] for k in sorted(self.queues.keys(), key=str.lower)]

    def to_xml(self) -> ET.ElementTree:
        root = ET.Element("EC.Entities.Adapter", {
            "xmlns:xsd": "http://www.w3.org/2001/XMLSchema",
            "xmlns:xsi": "http://www.w3.org/2001/XMLSchema-instance",
        })
        bqs = ET.SubElement(root, "BindingQueues")
        for q in self.list_attrs():
            q_el = ET.SubElement(bqs, "BindingQueue", {
                "AttributeFullName": q.attribute_full_name,
                "DefaultIsEnabled": _b(q.default_is_enabled),
            })
            binds_el = ET.SubElement(q_el, "Bindings")
            for idx, b in enumerate(q.bindings):
                t = getattr(b, "transform", None)
                attrs = {
                    "Index": str(idx),
                    "IsEnabled": _b(b.is_enabled),
                    "ParameterCode": b.parameter_code,
                    "TransformPrepareTextForNumeric": _b(getattr(t, "prepare_numeric", False)),
                }
                if t and getattr(t, "trim", ""):
                    attrs["TransformTrim"] = t.trim
                if t and getattr(t, "case", ""):
                    attrs["TransformCase"] = t.case
                bind_el = ET.SubElement(binds_el, "Binding", attrs)

                repls_el = ET.SubElement(bind_el, "TransformReplaces")
                if t:
                    for src, dst in getattr(t, "replaces", []) or []:
                        ET.SubElement(repls_el, "TransformReplace", {
                            "ReplaceableText": src,
                            "ReplacingText": dst
                        })
        return ET.ElementTree(root)

    @classmethod
    def from_xml(cls, path: str) -> "AdapterDoc":
        tree = ET.parse(path)
        root = tree.getroot()
        doc = cls()
        bqs = root.find("BindingQueues")
        if bqs is None:
            return doc
        for q_el in bqs.findall("BindingQueue"):
            attr = q_el.get("AttributeFullName", "").strip()
            q = doc.ensure_attr(attr)
            q.default_is_enabled = (q_el.get("DefaultIsEnabled", "false").lower() == "true")
            binds_el = q_el.find("Bindings")
            if not binds_el:
                continue
            pairs: List[Tuple[int, Binding]] = []
            for b_el in binds_el.findall("Binding"):
                try:
                    idx = int(b_el.get("Index", "0"))
                except Exception:
                    idx = 0
                t = TransformSettings()
                t.prepare_numeric = (b_el.get("TransformPrepareTextForNumeric", "false").lower() == "true")
                t.trim = (b_el.get("TransformTrim") or "").strip()
                t.case = (b_el.get("TransformCase") or "").strip()

                repls: List[Tuple[str, str]] = []
                repls_el = b_el.find("TransformReplaces")
                if repls_el is not None:
                    for r in repls_el.findall("TransformReplace"):
                        src = r.get("ReplaceableText", "") or ""
                        dst = r.get("ReplacingText", "") or ""
                        repls.append((src, dst))
                t.replaces = repls

                b = Binding(
                    parameter_code=b_el.get("ParameterCode", ""),
                    is_enabled=(b_el.get("IsEnabled", "true").lower() == "true"),
                    src_is_numeric=None,
                    src_model_title="",
                    transform=t
                )
                pairs.append((idx, b))
            pairs.sort(key=lambda t: t[0])
            q.bindings = [p[1] for p in pairs]
        return doc

# ----------------- Диалоги -----------------
class TransformDialog(QtWidgets.QDialog):
    def __init__(self, parent, binding: Binding):
        super().__init__(parent)
        self.setWindowTitle("Настройка преобразований")
        self.binding = binding
        self.resize(700, 480)

        v = QtWidgets.QVBoxLayout(self)

        row1 = QtWidgets.QHBoxLayout()
        v.addLayout(row1)
        row1.addWidget(QtWidgets.QLabel("Пробелы:"))
        self.cmbTrim = QtWidgets.QComboBox()
        self.cmbTrim.addItems(["Без изменений", "Удалить слева", "Удалить справа", "Удалить с обеих сторон"])
        row1.addWidget(self.cmbTrim)

        row2 = QtWidgets.QHBoxLayout()
        v.addLayout(row2)
        row2.addWidget(QtWidgets.QLabel("Регистр:"))
        self.cmbCase = QtWidgets.QComboBox()
        self.cmbCase.addItems(["Без изменений", "Привести к верхнему", "Привести к нижнему", "Заголовок"])
        row2.addWidget(self.cmbCase)

        v.addWidget(QtWidgets.QLabel("Сопоставление строк"))
        self.table = QtWidgets.QTableWidget(0, 2)
        self.table.setHorizontalHeaderLabels(["Замена с", "Замена на"])
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.verticalHeader().setVisible(False)
        self.table.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)

        v.addWidget(self.table)

        tb = QtWidgets.QHBoxLayout()
        v.addLayout(tb)
        self.btnAdd = QtWidgets.QPushButton("+")
        self.btnDel = QtWidgets.QPushButton("-")
        tb.addWidget(self.btnAdd)
        tb.addWidget(self.btnDel)
        tb.addStretch(1)

        bb = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel)
        v.addWidget(bb)

        self.btnAdd.clicked.connect(self.add_row)
        self.btnDel.clicked.connect(self.del_rows)
        bb.accepted.connect(self.accept_and_save)
        bb.rejected.connect(self.reject)

        trim_map_rev = {"": "Без изменений", "Left":"Удалить слева", "Right":"Удалить справа", "Both":"Удалить с обеих сторон"}
        case_map_rev = {"": "Без изменений", "Upper":"Привести к верхнему", "Lower":"Привести к нижнему", "Title":"Заголовок"}
        self.cmbTrim.setCurrentText(trim_map_rev.get(binding.transform.trim or "", "Без изменений"))
        self.cmbCase.setCurrentText(case_map_rev.get(binding.transform.case or "", "Без изменений"))
        for src, dst in (binding.transform.replaces or []):
            self.add_row(src, dst)

        # Подсветка латиницы/кириллицы
        self.table.itemChanged.connect(self._recolor_rows)
        self._recolor_rows()

    def _recolor_rows(self):
        import re as _re
        lat = _re.compile(r"[A-Za-z]")
        cyr = _re.compile(r"[А-Яа-яЁё]")
        for r in range(self.table.rowCount()):
            for c in (0,1):
                it = self.table.item(r,c)
                if not it:
                    continue
                txt = it.text() or ""
                has_lat = bool(lat.search(txt))
                has_cyr = bool(cyr.search(txt))
                if has_lat and not has_cyr:
                    it.setForeground(QtGui.QBrush(QtGui.QColor('#1f6feb')))  # латиница — синяя
                elif has_cyr and not has_lat:
                    it.setForeground(QtGui.QBrush(QtGui.QColor('#d73a49')))  # кириллица — красная
                elif has_lat and has_cyr:
                    it.setForeground(QtGui.QBrush(QtGui.QColor('#8a2be2')))  # смешанный — фиолетовый
                else:
                    it.setForeground(QtGui.QBrush())

    def add_row(self, src: str = "", dst: str = ""):
        r = self.table.rowCount()
        self.table.insertRow(r)
        self.table.setItem(r, 0, QtWidgets.QTableWidgetItem(src))
        self.table.setItem(r, 1, QtWidgets.QTableWidgetItem(dst))
        self._recolor_rows()

    def del_rows(self):
        sel = self.table.selectionModel().selectedRows()
        for m in reversed(sel):
            self.table.removeRow(m.row())
        self._recolor_rows()

    def accept_and_save(self):
        trim_map = {"Без изменений":"", "Удалить слева":"Left", "Удалить справа":"Right", "Удалить с обеих сторон":"Both"}
        case_map = {"Без изменений":"", "Привести к верхнему":"Upper", "Привести к нижнему":"Lower", "Заголовок":"Title"}
        self.binding.transform.trim = trim_map.get(self.cmbTrim.currentText(), "")
        self.binding.transform.case = case_map.get(self.cmbCase.currentText(), "")
        repl = []
        for r in range(self.table.rowCount()):
            s = (self.table.item(r,0).text() if self.table.item(r,0) else "").strip()
            d = (self.table.item(r,1).text() if self.table.item(r,1) else "").strip()
            if s or d:
                repl.append((s,d))
        self.binding.transform.replaces = repl
        self.accept()

class ModelFilterDialog(QtWidgets.QDialog):
    def __init__(self, parent, models: List[str], selected: Optional[Set[str]]):
        super().__init__(parent)
        self.setWindowTitle("Фильтр по моделям")
        self.resize(520, 600)
        v = QtWidgets.QVBoxLayout(self)

        h = QtWidgets.QHBoxLayout()
        btn_all = QtWidgets.QPushButton("Выделить все")
        btn_none = QtWidgets.QPushButton("Снять все")
        h.addWidget(btn_all); h.addWidget(btn_none); h.addStretch(1)
        v.addLayout(h)

        self.list = QtWidgets.QListWidget()
        self.list.setSelectionMode(QtWidgets.QAbstractItemView.NoSelection)
        v.addWidget(self.list)

        sel = set(models) if selected is None else set(selected)
        for m in models:
            it = QtWidgets.QListWidgetItem(m)
            it.setFlags(it.flags() | QtCore.Qt.ItemIsUserCheckable)
            it.setCheckState(QtCore.Qt.Checked if m in sel else QtCore.Qt.Unchecked)
            self.list.addItem(it)

        btn_all.clicked.connect(lambda: self._set_all(QtCore.Qt.Checked))
        btn_none.clicked.connect(lambda: self._set_all(QtCore.Qt.Unchecked))

        bb = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel)
        v.addWidget(bb)
        bb.accepted.connect(self.accept)
        bb.rejected.connect(self.reject)

    def _set_all(self, state):
        for i in range(self.list.count()):
            it = self.list.item(i); it.setCheckState(state)

    def chosen(self) -> List[str]:
        out = []
        for i in range(self.list.count()):
            it = self.list.item(i)
            if it.checkState() == QtCore.Qt.Checked:
                out.append(it.text())
        return out

# ----------------- Главное окно -----------------

class RowDragTable(QtWidgets.QTableWidget):
    """Таблица с перетаскиванием СТРОК (только по вертикали), без «ломания» ячеек."""
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        self.setSelectionMode(QtWidgets.QAbstractItemView.ExtendedSelection)
        self.setDragEnabled(True)
        self.setAcceptDrops(True)
        self.setDropIndicatorShown(True)
        self.setDragDropMode(QtWidgets.QAbstractItemView.DragDrop)  # без InternalMove
        self.setDefaultDropAction(QtCore.Qt.MoveAction)
        self.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        self._drag_rows = []   # внутреннее хранилище
        self.on_rows_moved = None  # callback(List[int] src_rows_sorted, int dest_row)

    def startDrag(self, supportedActions):
        # фиксируем исходные строки
        rows = sorted({i.row() for i in self.selectedIndexes()})
        if not rows:
            return
        self._drag_rows = rows
        drag = QtGui.QDrag(self)
        mime = QtCore.QMimeData()
        mime.setData("application/x-rows", b",".join(str(r).encode("utf-8") for r in rows))
        drag.setMimeData(mime)
        # простая картинка для dnd
        pm = QtGui.QPixmap(self.viewport().visibleRegion().boundingRect().size())
        pm.fill(QtCore.Qt.transparent)
        painter = QtGui.QPainter(pm)
        # рисуем первую переносимую строку как превью
        if rows:
            r = rows[0]
            rect = self.visualRect(self.model().index(r, 0))
            rect = rect.adjusted(0, 0, self.viewport().width()-1, rect.height())
            painter.fillRect(rect, QtGui.QBrush(QtGui.QColor(0, 0, 0, 30)))
        painter.end()
        drag.setPixmap(pm)
        drag.exec(QtCore.Qt.MoveAction)

    def dragEnterEvent(self, e: QtGui.QDragEnterEvent):
        if e.mimeData().hasFormat("application/x-rows"):
            e.acceptProposedAction()
        else:
            e.ignore()

    def dragMoveEvent(self, e: QtGui.QDragMoveEvent):
        if e.mimeData().hasFormat("application/x-rows"):
            e.acceptProposedAction()
        else:
            e.ignore()

    def dropEvent(self, e: QtGui.QDropEvent):
        if not e.mimeData().hasFormat("application/x-rows"):
            e.ignore(); return
        # вычисляем строку вставки
        pos = e.position().toPoint() if hasattr(e, "position") else e.pos()
        dest_row = self.rowAt(pos.y())
        if dest_row < 0:
            dest_row = self.rowCount()
        rows = self._drag_rows or []
        if not rows:
            e.ignore(); return
        # вызовем обратный вызов у владельца
        if callable(self.on_rows_moved):
            self.on_rows_moved(rows, dest_row)
        e.acceptProposedAction()


class MainWin(QtWidgets.QMainWindow):
    def _build_ui(self):
        self.ui = QtWidgets.QWidget(self)
        self.ui.setObjectName("centralwidget")
        self.setCentralWidget(self.ui)

        self.verticalLayout_central = QtWidgets.QVBoxLayout(self.ui)
        self.verticalLayout_central.setObjectName("verticalLayout_central")
        self.verticalLayout_central.setSpacing(8)
        self.verticalLayout_central.setContentsMargins(10, 10, 10, 10)

        # Header with back button and theme toggle
        header_h_layout = QtWidgets.QHBoxLayout()
        header_h_layout.setContentsMargins(0, 0, 0, 0)
        self._btn_back = create_back_button(self, icon_dir=ICON_DIR)
        self._btn_back.clicked.connect(lambda: go_to_main_menu(self))
        header_h_layout.addWidget(self._btn_back)
        header_h_layout.addStretch()
        self._theme_toggle = ThemeToggle(self)
        self._theme_toggle.setChecked(is_dark_theme(QtWidgets.QApplication.instance()))
        self._theme_toggle.toggled.connect(self._on_theme_toggled)
        header_h_layout.addWidget(self._theme_toggle)
        self.verticalLayout_central.addLayout(header_h_layout)

        # Top bar
        layout_top = QtWidgets.QHBoxLayout()
        layout_top.setObjectName("layoutTopBar")
        self.labelBase = QtWidgets.QLabel("Base URL:", self.ui)
        self.labelBase.setObjectName("labelBase")
        self.editBaseUrl = QtWidgets.QLineEdit(self.ui)
        self.editBaseUrl.setObjectName("editBaseUrl")
        self.editBaseUrl.setMinimumWidth(360)
        self.editBaseUrl.setPlaceholderText("http://localhost:5000")
        self.btnLoadProjects = QtWidgets.QPushButton("Загрузить проекты", self.ui)
        self.btnLoadProjects.setObjectName("btnLoadProjects")
        self.btnImportXml = QtWidgets.QPushButton("Импорт Adapter.xml", self.ui)
        self.btnImportXml.setObjectName("btnImportXml")
        self.btnImportXml.setToolTip("Импортировать существующий Adapter.xml")
        self.btnImportExcel = QtWidgets.QPushButton("Импорт Excel", self.ui)
        self.btnImportExcel.setObjectName("btnImportExcel")
        self.btnLoadGlobal = QtWidgets.QPushButton("Загрузить общие атрибуты", self.ui)
        self.btnLoadGlobal.setObjectName("btnLoadGlobal")

        layout_top.addWidget(self.labelBase)
        layout_top.addWidget(self.editBaseUrl)
        layout_top.addWidget(self.btnLoadProjects)
        layout_top.addWidget(self.btnImportXml)
        layout_top.addWidget(self.btnImportExcel)
        layout_top.addWidget(self.btnLoadGlobal)
        layout_top.addStretch(1)
        self.verticalLayout_central.addLayout(layout_top)

        # Project bar
        layout_proj = QtWidgets.QHBoxLayout()
        layout_proj.setObjectName("layoutProjectBar")
        self.labelProject = QtWidgets.QLabel("Проект:", self.ui)
        self.labelProject.setObjectName("labelProject")
        self.comboProject = QtWidgets.QComboBox(self.ui)
        self.comboProject.setObjectName("comboProject")
        self.comboProject.setSizeAdjustPolicy(QtWidgets.QComboBox.AdjustToContentsOnFirstShow)
        self.btnModelsWindow = QtWidgets.QPushButton("Окно моделей", self.ui)
        self.btnModelsWindow.setObjectName("btnModelsWindow")

        layout_proj.addWidget(self.labelProject)
        layout_proj.addWidget(self.comboProject)
        layout_proj.addWidget(self.btnModelsWindow)
        layout_proj.addStretch(1)
        self.verticalLayout_central.addLayout(layout_proj)

        # Main split
        self.layoutMainSplit = QtWidgets.QHBoxLayout()
        self.layoutMainSplit.setObjectName("layoutMainSplit")

        # Attributes
        self.groupAttributes = QtWidgets.QGroupBox("Атрибуты", self.ui)
        self.groupAttributes.setObjectName("groupAttributes")
        layout_attrs = QtWidgets.QVBoxLayout(self.groupAttributes)
        layout_attrs.setObjectName("layoutAttributes")

        layout_attr_search = QtWidgets.QHBoxLayout()
        layout_attr_search.setObjectName("layoutAttrSearch")
        self.labelAttrSearch = QtWidgets.QLabel("Поиск:", self.groupAttributes)
        self.labelAttrSearch.setObjectName("labelAttrSearch")
        self.editAttrSearch = QtWidgets.QLineEdit(self.groupAttributes)
        self.editAttrSearch.setObjectName("editAttrSearch")
        layout_attr_search.addWidget(self.labelAttrSearch)
        layout_attr_search.addWidget(self.editAttrSearch)
        layout_attrs.addLayout(layout_attr_search)

        self.treeAttributes = QtWidgets.QTreeView(self.groupAttributes)
        self.treeAttributes.setObjectName("treeAttributes")
        self.treeAttributes.setHeaderHidden(True)
        self.treeAttributes.setSelectionMode(QtWidgets.QAbstractItemView.ExtendedSelection)
        try:
            self.treeAttributes.setUniformRowHeights(True)
        except Exception:
            pass
        layout_attrs.addWidget(self.treeAttributes)

        layout_attr_btns = QtWidgets.QHBoxLayout()
        layout_attr_btns.setObjectName("layoutAttrButtons")
        self.btnAttrToggleDefault = QtWidgets.QPushButton("Вкл/Выкл по умолчанию", self.groupAttributes)
        self.btnAttrToggleDefault.setObjectName("btnAttrToggleDefault")
        self.btnAttrMarkWithBindings = QtWidgets.QPushButton("Отметить с привязками", self.groupAttributes)
        self.btnAttrMarkWithBindings.setObjectName("btnAttrMarkWithBindings")
        layout_attr_btns.addWidget(self.btnAttrToggleDefault)
        layout_attr_btns.addWidget(self.btnAttrMarkWithBindings)
        layout_attrs.addLayout(layout_attr_btns)

        self.labelAttrSource = QtWidgets.QLabel("Параметры не загружены", self.groupAttributes)
        self.labelAttrSource.setObjectName("labelAttrSource")
        layout_attrs.addWidget(self.labelAttrSource)

        # Bindings
        self.groupBindings = QtWidgets.QGroupBox("Привязки атрибута", self.ui)
        self.groupBindings.setObjectName("groupBindings")
        layout_bind = QtWidgets.QVBoxLayout(self.groupBindings)
        layout_bind.setObjectName("layoutBindings")

        self.tableBindings = RowDragTable(self.groupBindings)
        self.tableBindings.setObjectName("tableBindings")
        layout_bind.addWidget(self.tableBindings)

        layout_bind_btns = QtWidgets.QHBoxLayout()
        layout_bind_btns.setObjectName("layoutBindButtons")
        self.btnBindDelete = QtWidgets.QPushButton("Удалить выбранные", self.groupBindings)
        self.btnBindDelete.setObjectName("btnBindDelete")
        self.btnBindToggle = None
        self.btnTransform = QtWidgets.QPushButton("Преобразования", self.groupBindings)
        self.btnTransform.setObjectName("btnTransform")
        self.btnUp = QtWidgets.QPushButton("▲", self.groupBindings)
        self.btnUp.setObjectName("btnUp")
        self.btnUp.setMinimumWidth(36)
        self.btnDown = QtWidgets.QPushButton("▼", self.groupBindings)
        self.btnDown.setObjectName("btnDown")
        self.btnDown.setMinimumWidth(36)

        layout_bind_btns.addWidget(self.btnBindDelete)
        layout_bind_btns.addWidget(self.btnTransform)
        layout_bind_btns.addStretch(1)
        layout_bind_btns.addWidget(self.btnUp)
        layout_bind_btns.addWidget(self.btnDown)
        layout_bind.addLayout(layout_bind_btns)

        # Params
        self.groupParams = QtWidgets.QGroupBox("Параметры моделей", self.ui)
        self.groupParams.setObjectName("groupParams")
        layout_params = QtWidgets.QVBoxLayout(self.groupParams)
        layout_params.setObjectName("layoutParams")

        layout_param_filter = QtWidgets.QHBoxLayout()
        layout_param_filter.setObjectName("layoutParamFilter")
        self.labelFilter = QtWidgets.QLabel("Фильтр по шифру:", self.groupParams)
        self.labelFilter.setObjectName("labelFilter")
        self.editParamFilter = QtWidgets.QLineEdit(self.groupParams)
        self.editParamFilter.setObjectName("editParamFilter")
        self.btnModelFilter = QtWidgets.QPushButton("Фильтр по моделям", self.groupParams)
        self.btnModelFilter.setObjectName("btnModelFilter")
        layout_param_filter.addWidget(self.labelFilter)
        layout_param_filter.addWidget(self.editParamFilter)
        layout_param_filter.addWidget(self.btnModelFilter)
        layout_params.addLayout(layout_param_filter)

        self.tableParams = QtWidgets.QTableWidget(self.groupParams)
        self.tableParams.setObjectName("tableParams")
        layout_params.addWidget(self.tableParams)

        self.btnAddSelected = QtWidgets.QPushButton("Добавить выделенные → привязки", self.groupParams)
        self.btnAddSelected.setObjectName("btnAddSelected")
        layout_params.addWidget(self.btnAddSelected)

        self.layoutMainSplit.addWidget(self.groupAttributes, 1)
        self.layoutMainSplit.addWidget(self.groupBindings, 2)
        self.layoutMainSplit.addWidget(self.groupParams, 3)
        self.verticalLayout_central.addLayout(self.layoutMainSplit)

        # Bottom
        layout_bottom = QtWidgets.QHBoxLayout()
        layout_bottom.setObjectName("layoutBottom")
        layout_bottom.addStretch(1)
        self.btnSave = QtWidgets.QPushButton("Сохранить адаптер", self.ui)
        self.btnSave.setObjectName("btnSave")
        self.btnSave.setMinimumWidth(320)
        self.btnSaveGlobal = QtWidgets.QPushButton("Сохранить общие атрибуты", self.ui)
        self.btnSaveGlobal.setObjectName("btnSaveGlobal")
        self.btnSaveGlobal.setMinimumWidth(320)
        layout_bottom.addWidget(self.btnSave)
        layout_bottom.addWidget(self.btnSaveGlobal)
        layout_bottom.addStretch(1)
        self.verticalLayout_central.addLayout(layout_bottom)

        self.lblStatus = QtWidgets.QLabel("", self.ui)
        self.lblStatus.setObjectName("lblStatus")
        self.verticalLayout_central.addWidget(self.lblStatus)

    def __init__(self):
        super().__init__()
        self._pending_theme_apply = False
        self._build_ui()

        self.setWindowTitle("Редактор адаптера")
        self.resize(1420, 760)
        self.setMinimumSize(1000, 680)

        self.doc = AdapterDoc()
        self.projects: List[Dict[str, Any]] = []
        self.containers: List[Dict[str, Any]] = []
        self.per_model_params: Dict[int, List[Dict[str, Any]]] = {}
        self.selected_attr: Optional[str] = None
        self.attr_types: Dict[str, Optional[bool]] = {}
        self.param_models_map: Dict[int, List[str]] = {}
        self.bind_models_map: Dict[int, List[str]] = {}
        self._is_filling_bind_table: bool = False
        self._icon_bind_checked: Optional[QtGui.QIcon] = None
        self._icon_bind_unchecked: Optional[QtGui.QIcon] = None
        self._icon_bind_indeterminate: Optional[QtGui.QIcon] = None

        # Фильтры "Параметры моделей"
        self.type_filter: Optional[bool] = None     # None=оба, True=число, False=текст
        self.model_filter_set: Optional[Set[str]] = None  # None=все модели, иначе только выбранные

        # (виджет удалён из интерфейса, но логика поддерживает значение по умолчанию)
        self.chkAutoEnable = None

        # Пропорции колонок: Атрибуты, Привязки, Параметры (равные)
        try:
            self.layoutMainSplit.setStretch(0, 1)
            self.layoutMainSplit.setStretch(1, 1)
            self.layoutMainSplit.setStretch(2, 1)
        except Exception:
            pass


        # Модель для дерева атрибутов
        self.attrModel: QtGui.QStandardItemModel = QtGui.QStandardItemModel(0, 2, self.treeAttributes)
        self.treeAttributes.setModel(self.attrModel)
        self.attrModel.setHorizontalHeaderLabels(["Атрибуты", ""])
        self.treeAttributes.header().hide()
        try:
            self.treeAttributes.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
            self.treeAttributes.setAllColumnsShowFocus(True)
        except Exception:
            pass
        try:
            hdr = self.treeAttributes.header()
            hdr.setStretchLastSection(False)
            hdr.setSectionResizeMode(0, QtWidgets.QHeaderView.Stretch)
            hdr.setSectionResizeMode(1, QtWidgets.QHeaderView.Fixed)
            self.treeAttributes.setColumnWidth(1, 28)
        except Exception:
            pass

        # иконки статуса атрибутов (справа)
        try:
            app = QtWidgets.QApplication.instance()
            ok_path = resolve_icon_path("ok", ICON_DIR, app=app) or _resolve_icon_path(_ICON_FALLBACK_FILES["ok"])
            circle_path = resolve_icon_path("circle", ICON_DIR, app=app) or _resolve_icon_path(_ICON_FALLBACK_FILES["circle"])
            self._icon_attr_assigned = QtGui.QIcon(ok_path) if ok_path else _icon_ok()
            self._icon_attr_unassigned = QtGui.QIcon(circle_path) if circle_path else _icon_circle()
        except Exception:
            self._icon_attr_assigned = _icon_ok()
            self._icon_attr_unassigned = _icon_circle()

        # Иконки чекбоксов (как в других окнах: unchecked=check, checked=select)
        self._reload_bind_checkbox_icons(is_dark_theme(QtWidgets.QApplication.instance()))

        # Таблицы
        self.tableBindings.setColumnCount(4)
        self.tableBindings.setHorizontalHeaderLabels(["", "Имя параметра", "Преобразование", "Источник"])
        self.tableBindings.horizontalHeader().setStretchLastSection(True)
        self.tableBindings.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        self.tableBindings.setSelectionMode(QtWidgets.QAbstractItemView.ExtendedSelection)
        self.tableBindings.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        try:
            self.tableBindings.setShowGrid(False)
        except Exception:
            pass
        # Размеры столбцов привязок
        self.tableBindings.horizontalHeader().setSectionResizeMode(0, QtWidgets.QHeaderView.ResizeToContents)
        self.tableBindings.horizontalHeader().setSectionResizeMode(1, QtWidgets.QHeaderView.Stretch)
        self.tableBindings.horizontalHeader().setSectionResizeMode(2, QtWidgets.QHeaderView.ResizeToContents)
        self.tableBindings.horizontalHeader().setSectionResizeMode(3, QtWidgets.QHeaderView.Stretch)
        self.tableBindings.setColumnWidth(0, 40)
        try:
            self.tableBindings.setIconSize(QtCore.QSize(18, 18))
        except Exception:
            pass
        try:
            h = self.tableBindings.horizontalHeader()
            h.setSectionsClickable(True)
            h.sectionClicked.connect(self._on_bind_header_clicked)
        except Exception:
            pass

        # callback на перестановку строк для обновления порядка в данных
        def _rows_moved(src_rows: list[int], dest_row: int):
            q = self.current_queue()
            if not q or not src_rows:
                return
            n = len(q.bindings)
            src = sorted(set([r for r in src_rows if 0 <= r < n]))
            moving = [q.bindings[i] for i in src]
            remaining = [q.bindings[i] for i in range(n) if i not in src]
            # позиция вставки в массиве remaining
            # если вставка ниже исходного блока - сдвигаем индекс на количество удалённых строк до неё
            shift = sum(1 for i in src if i < dest_row)
            ins = max(0, min(len(remaining), dest_row - shift))
            new_order = remaining[:ins] + moving + remaining[ins:]
            q.bindings = new_order
            self.fill_bind_table(q)
            # выделим перемещённые строки
            for i in range(ins, ins + len(moving)):
                self.tableBindings.selectRow(i)
        self.tableBindings.on_rows_moved = _rows_moved

        # drag&drop
        self.tableBindings.setDragEnabled(True)
        self.tableBindings.setAcceptDrops(True)
        self.tableBindings.setDropIndicatorShown(True)
        self.tableBindings.setDragDropMode(QtWidgets.QAbstractItemView.InternalMove)
        self.tableBindings.setDragDropOverwriteMode(False)
        self.tableBindings.setDefaultDropAction(QtCore.Qt.MoveAction)
        self.tableBindings.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        self.tableBindings.viewport().installEventFilter(self)
        # context menu
        self.tableBindings.setContextMenuPolicy(QtCore.Qt.CustomContextMenu)
        self.tableBindings.customContextMenuRequested.connect(self._on_bind_context_menu)

        self.tableParams.setColumnCount(3)
        self.tableParams.setHorizontalHeaderLabels(["Шифр", "Тип", "Модель"])
        self.tableParams.horizontalHeader().setStretchLastSection(False)
        self.tableParams.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        self.tableParams.setSelectionMode(QtWidgets.QAbstractItemView.ExtendedSelection)
        self.tableParams.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        try:
            self.tableParams.setShowGrid(False)
        except Exception:
            pass
        # размеры колонок: шифр тянется, тип и модель по содержимому
        self.tableParams.horizontalHeader().setSectionResizeMode(0, QtWidgets.QHeaderView.ResizeToContents)  # Шифр
        self.tableParams.horizontalHeader().setSectionResizeMode(1, QtWidgets.QHeaderView.ResizeToContents)  # Тип
        self.tableParams.horizontalHeader().setSectionResizeMode(2, QtWidgets.QHeaderView.Stretch)  # Модель
        # сделаем минимальную ширину для колонки "Шифр" покрупнее
        self.tableParams.horizontalHeader().setMinimumSectionSize(40)

        # кликабельный заголовок для фильтра по типу
        header = self.tableParams.horizontalHeader()
        header.setSectionsClickable(True)
        header.sectionClicked.connect(self.on_param_header_clicked)

        # Внешний вид
        self.apply_style()
        self.setup_icon_buttons()

        # События
        self.editBaseUrl.setText(DEFAULT_BASE_URL)
        self.btnLoadProjects.clicked.connect(self.load_projects)
        self.btnImportXml.clicked.connect(self.import_xml)
        self.btnLoadGlobal.clicked.connect(self.load_global_attrs)
        self.btnImportExcel.clicked.connect(self.import_excel)
        btnModelsWindow = getattr(self, 'btnModelsWindow', None)
        if btnModelsWindow is not None:
            btnModelsWindow.clicked.connect(self.open_models_window)
        self.comboProject.currentIndexChanged.connect(self.on_project_changed)
        self.editAttrSearch.textChanged.connect(lambda *_: self.refresh_attr_tree(True))
        self.treeAttributes.selectionModel().selectionChanged.connect(self.on_attr_pick)
        self.btnAttrToggleDefault.clicked.connect(self.attrs_toggle_default)
        self.btnAttrMarkWithBindings.clicked.connect(self.attrs_mark_with_bindings)
        if self.btnBindDelete is not None:
            self.btnBindDelete.clicked.connect(self.bind_delete)
        if self.btnBindToggle is not None:
            self.btnBindToggle.clicked.connect(self.bind_toggle_enabled)
        if self.btnTransform is not None:
            self.btnTransform.clicked.connect(self.open_transform_dialog)
        if self.btnUp is not None:
            self.btnUp.clicked.connect(self.bind_move_up)
        if self.btnDown is not None:
            self.btnDown.clicked.connect(self.bind_move_down)
        self.btnAddSelected.clicked.connect(self.add_selected_params_as_bindings)
        self.btnSave.clicked.connect(self.export_xml)
        if self.btnSaveGlobal is not None:
            self.btnSaveGlobal.clicked.connect(self.export_global_attrs_xml)
        self.tableParams.itemDoubleClicked.connect(lambda *_: self.add_selected_params_as_bindings())
        self.tableBindings.itemDoubleClicked.connect(lambda *_: self.bind_delete())
        self.tableBindings.cellClicked.connect(self._on_bind_cell_clicked)
        self.editParamFilter.textChanged.connect(lambda *_: self.fill_param_table())
        self.btnModelFilter.clicked.connect(self.open_model_filter_dialog)

        # Старт
        self.load_projects()
        self.refresh_attr_tree(True)

    def changeEvent(self, event: QtCore.QEvent) -> None:
        try:
            t = event.type()
            if t in (QtCore.QEvent.StyleChange, QtCore.QEvent.PaletteChange, QtCore.QEvent.ApplicationPaletteChange):
                if not self._pending_theme_apply:
                    self._pending_theme_apply = True
                    QtCore.QTimer.singleShot(0, self._apply_theme_extras_from_app)
        except Exception:
            pass
        super().changeEvent(event)

    def _reload_attr_status_icons(self) -> None:
        try:
            app = QtWidgets.QApplication.instance()
            ok_path = resolve_icon_path("ok", ICON_DIR, app=app) or _resolve_icon_path(_ICON_FALLBACK_FILES["ok"])
            circle_path = resolve_icon_path("circle", ICON_DIR, app=app) or _resolve_icon_path(_ICON_FALLBACK_FILES["circle"])
            self._icon_attr_assigned = QtGui.QIcon(ok_path) if ok_path else _icon_ok()
            self._icon_attr_unassigned = QtGui.QIcon(circle_path) if circle_path else _icon_circle()
        except Exception:
            self._icon_attr_assigned = _icon_ok()
            self._icon_attr_unassigned = _icon_circle()

    def _apply_theme_extras(self, dark: bool) -> None:
        self._reload_bind_checkbox_icons(bool(dark))
        self._reload_attr_status_icons()
        try:
            self.setup_icon_buttons()
        except Exception:
            pass
        try:
            self.refresh_attr_tree(True)
        except Exception:
            pass
        q = self.current_queue()
        if q:
            self.fill_bind_table(q)

    def _apply_theme_extras_from_app(self) -> None:
        self._pending_theme_apply = False
        try:
            app = QtWidgets.QApplication.instance()
            self._apply_theme_extras(is_dark_theme(app))
        except Exception:
            pass

    def _on_theme_toggled(self, dark: bool):
        app = QtWidgets.QApplication.instance()
        theme(app, dark, icon_dir=ICON_DIR)
        self._apply_theme_extras(bool(dark))

 # ----------------- Внешний вид -----------------
    def apply_style(self):
        app = QtWidgets.QApplication.instance()
        theme(app, is_dark_theme(app), icon_dir=ICON_DIR)
        
        # Установка делегата для скругленной подсветки строк
        for table in [self.tableBindings, self.tableParams]:
            delegate = RowHoverDelegate(table)
            table.setItemDelegate(delegate)
            # Установка row highlighter для скругленных фонов
            install_viewport_row_highlighter(table)
            # Включение отслеживания наведения мыши
            setup_hover_tracking(table)
        self._apply_theme_extras(is_dark_theme(app))

    def _reload_bind_checkbox_icons(self, dark: bool):
        try:
            app = QtWidgets.QApplication.instance()
            off_path = resolve_icon_path("check", ICON_DIR, app=app) or _resolve_icon_path(_ICON_FALLBACK_FILES["check"])
            on_path = resolve_icon_path("select", ICON_DIR, app=app) or _resolve_icon_path(_ICON_FALLBACK_FILES["select"])
            mid_path = resolve_icon_path("poloska", ICON_DIR, app=app) or _resolve_icon_path(_ICON_FALLBACK_FILES["poloska"])
            self._icon_bind_unchecked = QtGui.QIcon(off_path) if off_path else None
            self._icon_bind_checked = QtGui.QIcon(on_path) if on_path else None
            self._icon_bind_indeterminate = QtGui.QIcon(mid_path) if mid_path else None
        except Exception:
            self._icon_bind_unchecked = None
            self._icon_bind_checked = None
            self._icon_bind_indeterminate = None

    def setup_icon_buttons(self):
        icon_key_by_filename = {
            "arrow-up.png": "arrow_up",
            "arrow-down.png": "arrow_down",
            "delete.png": "delete",
            "free-icon-setting-3288004.png": "gear",
        }
        def _norm(name: str) -> str:
            return icon_key_by_filename.get(str(name or ""), str(name or ""))

        def _apply(btn: Optional[QtWidgets.QPushButton], icon_name: str, tooltip: str):
            if btn is None:
                return
            try:
                btn.setText("")
            except Exception:
                pass
            try:
                btn.setToolTip(tooltip)
            except Exception:
                pass
            icon = None
            icon_key = _norm(icon_name)
            app = QtWidgets.QApplication.instance()
            icon_path = resolve_icon_path(icon_key, ICON_DIR, app=app) or _resolve_icon_path(icon_name)
            if icon_path:
                try:
                    icon = QtGui.QIcon(icon_path)
                except Exception:
                    icon = None
            else:
                try:
                    if icon_key in ("arrow_up", "arrow-up.png"):
                        icon = _icon_arrow("up")
                    elif icon_key in ("arrow_down", "arrow-down.png"):
                        icon = _icon_arrow("down")
                    elif icon_key in ("delete", "delete.png"):
                        icon = self.style().standardIcon(QtWidgets.QStyle.SP_TrashIcon)
                    else:
                        icon = self.style().standardIcon(QtWidgets.QStyle.SP_FileDialogDetailedView)
                except Exception:
                    icon = None
            if icon is not None:
                try:
                    btn.setIcon(icon)
                except Exception:
                    pass
            try:
                btn.setIconSize(QtCore.QSize(18, 18))
                btn.setMinimumSize(QtCore.QSize(34, 34))
                btn.setMaximumSize(QtCore.QSize(34, 34))
            except Exception:
                pass

        _apply(self.btnBindDelete, "delete", "Удалить выбранные")
        _apply(self.btnTransform, "gear", "Преобразования")
        _apply(self.btnUp, "arrow_up", "Вверх")
        _apply(self.btnDown, "arrow_down", "Вниз")

    # ---------- API helpers ----------
    def base(self) -> str:
        s = self.editBaseUrl.text().strip() or DEFAULT_BASE_URL
        return s.rstrip("/")

    # ---------- Проекты/модели/параметры ----------
    def on_project_changed(self, idx: int):
        """При смене проекта автоматически подгружаем модели и сбрасываем параметры/фильтры."""
        if idx is None or idx < 0 or idx >= len(self.projects):
            return
        # загрузка контейнеров для выбранного проекта
        pid = int(self.projects[idx]["id"])
        try:
            self.containers = sorted(api_get_containers(self.base(), pid), key=lambda c: (c["title"] or "").lower())
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "API", f"Не удалось получить модели:\n{e}")
            return
        # сбрасываем параметры и фильтры по моделям (т.к. модели другие)
        self.per_model_params.clear()
        self.model_filter_set = None
        self.fill_param_table()
        self.lblStatus.setText(f"Загружено моделей: {len(self.containers)}")
    
    def load_projects(self):
        try:
            self.projects = sorted(api_get_projects(self.base()), key=lambda p: (p["title"] or "").lower())
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "API", f"Не удалось получить проекты:\n{e}")
            return
        self.comboProject.clear()
        for p in self.projects:
            self.comboProject.addItem(p["title"])
        # автоматически подгрузим модели для выбранного проекта
        self.on_project_changed(self.comboProject.currentIndex())
        self.containers.clear()
        self.per_model_params.clear()
        self.tableParams.setRowCount(0)
        self.lblStatus.setText("")

    def ensure_containers_loaded(self) -> bool:
        i = self.comboProject.currentIndex()
        if i < 0:
            QtWidgets.QMessageBox.warning(self, "API", "Выбери проект.")
            return False
        if self.containers:
            return True
        pid = int(self.projects[i]["id"])
        try:
            self.containers = sorted(api_get_containers(self.base(), pid), key=lambda c: (c["title"] or "").lower())
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "API", f"Не удалось получить модели:\n{e}")
            return False
        return True

    def open_models_window(self):
        if not self.ensure_containers_loaded():
            return
        dlg = QtWidgets.QDialog(self)
        dlg.setWindowTitle("Выбор моделей")
        dlg.resize(700, 420)
        v = QtWidgets.QVBoxLayout(dlg)

        lst = QtWidgets.QListWidget()
        lst.setSelectionMode(QtWidgets.QAbstractItemView.ExtendedSelection)
        for c in self.containers:
            lst.addItem(c["title"])
        v.addWidget(lst)

        h = QtWidgets.QHBoxLayout()
        v.addLayout(h)
        btnAll = QtWidgets.QPushButton("Выделить все")
        btnNone = QtWidgets.QPushButton("Снять выделение")
        btnLoad = QtWidgets.QPushButton("Загрузить параметры")
        h.addWidget(btnAll); h.addWidget(btnNone); h.addStretch(1); h.addWidget(btnLoad)

        btnAll.clicked.connect(lambda: lst.selectAll())
        btnNone.clicked.connect(lambda: lst.clearSelection())

        def do_load():
            sels = lst.selectedIndexes()
            if not sels:
                QtWidgets.QMessageBox.warning(dlg, "API", "Выдели минимум одну модель.")
                return
            self.per_model_params.clear()
            self.model_filter_set = None  # сбрасываем фильтр моделей при новой загрузке
            errs = 0
            for i in sels:
                idx = i.row()
                cid = int(self.containers[idx]["id"])
                try:
                    self.per_model_params[cid] = api_get_params_for_container(self.base(), cid)
                except Exception as e:
                    errs += 1
                    QtWidgets.QMessageBox.critical(dlg, "API", f"Не получил параметры для «{self.containers[idx]['title']}»:\n{e}")
            self.fill_param_table()
            self.lblStatus.setText(f"Загружено моделей: {len(sels)}; ошибок: {errs}")
            dlg.accept()

        btnLoad.clicked.connect(do_load)
        dlg.exec()

    def current_models(self) -> List[str]:
        title_by_cid = {c["id"]: c["title"] for c in self.containers}
        return sorted([title_by_cid.get(cid, f"ID {cid}") for cid in self.per_model_params.keys()], key=str.lower)

    def fill_param_table(self):
        title_by_cid = {c["id"]: c["title"] for c in self.containers}
        grouped: Dict[str, Dict[str, Any]] = {}

        code_q = (self.editParamFilter.text() or "").lower().strip()

        for cid, items in self.per_model_params.items():
            model_title = title_by_cid.get(cid, f"ID {cid}")
            for r in items:
                code = r["code"]
                # 1) по шифру
                if code_q and code_q not in code.lower():
                    continue
                # 2) по типу
                if self.type_filter is not None and bool(r["isNumeric"]) != bool(self.type_filter):
                    continue
                # 3) по моделям
                if self.model_filter_set is not None and model_title not in self.model_filter_set:
                    continue
                g = grouped.setdefault(code, {"isnum": r["isNumeric"], "models": set()})
                g["isnum"] = g["isnum"] or r["isNumeric"]
                g["models"].add(model_title)

        rows = []
        for code, data in grouped.items():
            models = sorted(list(data["models"]), key=str.lower)
            isnum = "Число" if data["isnum"] else "Текст"
            rows.append((code, isnum, models))

        rows.sort(key=lambda x: (x[0].lower(), x[2][0].lower() if x[2] else ""))

        self.tableParams.setRowCount(0)
        self.param_models_map.clear()
        for i, (code, isnum, models) in enumerate(rows):
            self.tableParams.insertRow(i)
            self.tableParams.setItem(i, 0, QtWidgets.QTableWidgetItem(code))
            self.tableParams.setItem(i, 1, QtWidgets.QTableWidgetItem(isnum))
            it2 = QtWidgets.QTableWidgetItem(", ".join(models))
            it2.setToolTip("\n".join(models))
            self.tableParams.setItem(i, 2, it2)
            self.param_models_map[i] = models
        # Ужимаем колонку 'Тип' под содержимое
        self.tableParams.resizeColumnToContents(1)

    # ---------- Фильтры заголовков ----------
    
    def on_param_header_clicked(self, section: int):
        if section == 1:
            self._show_type_filter_menu()
        elif section == 2:
            # по клику на "Модель" открываем окно выбора моделей
            self.open_model_filter_dialog()

    def _show_type_filter_menu(self):
        menu = QtWidgets.QMenu(self)
        act_all = QtGui.QAction("Оба", menu)
        act_text = QtGui.QAction("Текст", menu)
        act_num = QtGui.QAction("Число", menu)
        ag = QtGui.QActionGroup(menu)
        for a in (act_all, act_text, act_num):
            a.setCheckable(True)
            ag.addAction(a); menu.addAction(a)
        if self.type_filter is None:
            act_all.setChecked(True)
        elif self.type_filter is False:
            act_text.setChecked(True)
        else:
            act_num.setChecked(True)

        def apply_choice(a: QtGui.QAction):
            if a is act_all:
                self.type_filter = None
            elif a is act_text:
                self.type_filter = False
            else:
                self.type_filter = True
            self.fill_param_table()
        ag.triggered.connect(apply_choice)
        menu.exec(QtGui.QCursor.pos())

    def open_model_filter_dialog(self):
        models = self.current_models()
        if not models:
            QtWidgets.QMessageBox.information(self, "Фильтр по моделям", "Сначала загрузите параметры через «Окно моделей».")
            return
        dlg = ModelFilterDialog(self, models, self.model_filter_set)
        if dlg.exec() == QtWidgets.QDialog.Accepted:
            chosen = dlg.chosen()
            self.model_filter_set = None if len(chosen) == len(models) else set(chosen)
            self.fill_param_table()

    # ---------- Атрибуты слева ----------
    def _attr_tree_state(self):
        expanded = set()
        for i in range(self.attrModel.rowCount()):
            item = self.attrModel.item(i, 0)
            if item and item.hasChildren() and self.treeAttributes.isExpanded(item.index()):
                expanded.add(item.text())
        sel = None
        indexes = self.treeAttributes.selectionModel().selectedIndexes()
        if indexes:
            it = self.attrModel.itemFromIndex(indexes[0])
            data = it.data()
            if data and data[0] == "leaf":
                sel = data[1]
        return expanded, sel

    def _restore_attr_tree_state(self, expanded, selected):
        for i in range(self.attrModel.rowCount()):
            item = self.attrModel.item(i, 0)
            if item and item.text() in expanded:
                self.treeAttributes.setExpanded(item.index(), True)
        if selected:
            def walk(parent=None):
                rng = range(self.attrModel.rowCount()) if parent is None else range(parent.rowCount())
                for i in rng:
                    it = self.attrModel.item(i,0) if parent is None else parent.child(i)
                    if not it: continue
                    data = it.data()
                    if data and data[0]=="leaf" and data[1]==selected:
                        self.treeAttributes.setCurrentIndex(it.index()); return True
                    if it.hasChildren() and walk(it): return True
                return False
            walk()

    def refresh_attr_tree(self, preserve_state: bool = True):
        expanded, selected = self._attr_tree_state() if preserve_state else (set(), None)
        q = (self.editAttrSearch.text() or "").lower().strip()
        self.attrModel.removeRows(0, self.attrModel.rowCount())
        buckets: Dict[str, List[str]] = {}
        for a in self.doc.list_attrs():
            full = a.attribute_full_name.strip()
            if q and q not in full.lower(): continue
            if "." in full:
                grp, leaf = full.split(".", 1)
            else:
                grp, leaf = "", full
            buckets.setdefault(grp, []).append(leaf)

        for grp in sorted(buckets.keys(), key=str.lower):
            if grp:
                parent = QtGui.QStandardItem(grp)
                parent.setEditable(False)
                parent_status = QtGui.QStandardItem("")
                parent_status.setEditable(False)
                try:
                    parent_status.setTextAlignment(QtCore.Qt.AlignCenter)
                except Exception:
                    pass
                self.attrModel.appendRow([parent, parent_status])
                for leaf in sorted(set(buckets[grp]), key=str.lower):
                    full = f"{grp}.{leaf}"
                    text = self.decorate_leaf(full)
                    child = QtGui.QStandardItem(text)
                    child.setData(("leaf", full))
                    child.setEditable(False)
                    status = QtGui.QStandardItem("")
                    status.setEditable(False)
                    status.setData(("leaf", full))
                    try:
                        status.setTextAlignment(QtCore.Qt.AlignCenter)
                    except Exception:
                        pass
                    try:
                        qobj = self.doc.queues.get(full)
                        has_bindings = bool(qobj and qobj.bindings)
                        if has_bindings and getattr(self, "_icon_attr_assigned", None):
                            status.setIcon(self._icon_attr_assigned)  # type: ignore[arg-type]
                        elif (not has_bindings) and getattr(self, "_icon_attr_unassigned", None):
                            status.setIcon(self._icon_attr_unassigned)  # type: ignore[arg-type]
                    except Exception:
                        pass
                    parent.appendRow([child, status])
            else:
                for leaf in sorted(set(buckets.get("", [])), key=str.lower):
                    full = leaf
                    text = self.decorate_leaf(full)
                    item = QtGui.QStandardItem(text)
                    item.setData(("leaf", full))
                    item.setEditable(False)
                    status = QtGui.QStandardItem("")
                    status.setEditable(False)
                    status.setData(("leaf", full))
                    try:
                        status.setTextAlignment(QtCore.Qt.AlignCenter)
                    except Exception:
                        pass
                    try:
                        qobj = self.doc.queues.get(full)
                        has_bindings = bool(qobj and qobj.bindings)
                        if has_bindings and getattr(self, "_icon_attr_assigned", None):
                            status.setIcon(self._icon_attr_assigned)  # type: ignore[arg-type]
                        elif (not has_bindings) and getattr(self, "_icon_attr_unassigned", None):
                            status.setIcon(self._icon_attr_unassigned)  # type: ignore[arg-type]
                    except Exception:
                        pass
                    self.attrModel.appendRow([item, status])
        self._restore_attr_tree_state(expanded, selected)

    def decorate_leaf(self, full: str) -> str:
        return full

    def on_attr_pick(self):
        indexes = self.treeAttributes.selectionModel().selectedIndexes()
        if not indexes:
            self.selected_attr = None; self.tableBindings.setRowCount(0); return
        item: QtGui.QStandardItem = self.attrModel.itemFromIndex(indexes[0])
        data = item.data()
        if not data or data[0] != "leaf":
            self.selected_attr = None; self.tableBindings.setRowCount(0); return
        name = data[1]
        if "." not in name:
            self.selected_attr = None; self.tableBindings.setRowCount(0); return
        self.selected_attr = name
        q = self.doc.ensure_attr(name)
        self.fill_bind_table(q)

    def attrs_toggle_default(self):
        changed = False
        for idx in self.treeAttributes.selectionModel().selectedIndexes():
            item = self.attrModel.itemFromIndex(idx); data = item.data()
            if not data or data[0] != "leaf": continue
            name = data[1]; q = self.doc.ensure_attr(name)
            q.default_is_enabled = not q.default_is_enabled; changed = True
        if changed:
            self.refresh_attr_tree(True)

    def attrs_mark_with_bindings(self):
        changed = False
        for name, q in self.doc.queues.items():
            if q.bindings and not q.default_is_enabled:
                q.default_is_enabled = True; changed = True
        if changed:
            self.refresh_attr_tree(True)
            self.lblStatus.setText("Отмечены атрибуты с привязками.")
        else:
            self.lblStatus.setText("Нет изменений.")

    # ---------- Таблица привязок ----------
    def fill_bind_table(self, q: BindingQueue):
        self._is_filling_bind_table = True
        try:
            self.tableBindings.setRowCount(0)
            dst_isnum = self.attr_types.get(q.attribute_full_name, None)
            dst_s = "Число" if dst_isnum else "Текст" if dst_isnum is not None else "?"
            self.bind_models_map.clear()
            for i, b in enumerate(q.bindings):
                src_s = "Число" if b.src_is_numeric else "Текст" if b.src_is_numeric is not None else "?"
                conv = f"{src_s} -> {dst_s}"
                self.tableBindings.insertRow(i)

                it0 = QtWidgets.QTableWidgetItem("")
                it0.setData(QtCore.Qt.UserRole, b.uid)
                it0.setTextAlignment(QtCore.Qt.AlignCenter)
                it0.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                if b.is_enabled:
                    if self._icon_bind_checked is not None:
                        it0.setIcon(self._icon_bind_checked)
                    else:
                        it0.setIcon(_icon_checkbox(True))
                else:
                    if self._icon_bind_unchecked is not None:
                        it0.setIcon(self._icon_bind_unchecked)
                    else:
                        it0.setIcon(_icon_checkbox(False))
                self.tableBindings.setItem(i, 0, it0)

                self.tableBindings.setItem(i, 1, QtWidgets.QTableWidgetItem(b.parameter_code))
                self.tableBindings.setItem(i, 2, QtWidgets.QTableWidgetItem(conv))
                it3 = QtWidgets.QTableWidgetItem(b.src_model_title or "")
                if b.src_model_title:
                    it3.setToolTip("\n".join([m.strip() for m in b.src_model_title.split(",") if m.strip()]))
                self.tableBindings.setItem(i, 3, it3)
                self.bind_models_map[i] = [m.strip() for m in (b.src_model_title or "").split(",") if m.strip()]
            self._update_bind_header_checkbox(q)
        finally:
            self._is_filling_bind_table = False

    def _update_bind_header_checkbox(self, q: BindingQueue):
        try:
            header_item = self.tableBindings.horizontalHeaderItem(0)
            if header_item is None:
                header_item = QtWidgets.QTableWidgetItem("")
                self.tableBindings.setHorizontalHeaderItem(0, header_item)
            header_item.setText("")
            if not q.bindings:
                icon = self._icon_bind_unchecked or _icon_checkbox(False)
            else:
                enabled = [bool(b.is_enabled) for b in q.bindings]
                if all(enabled):
                    icon = self._icon_bind_checked or _icon_checkbox(True)
                elif any(enabled):
                    icon = self._icon_bind_indeterminate or (self._icon_bind_unchecked or _icon_checkbox(False))
                else:
                    icon = self._icon_bind_unchecked or _icon_checkbox(False)
            header_item.setIcon(icon)
        except Exception:
            pass

    def _on_bind_header_clicked(self, section: int):
        if section != 0:
            return
        if self._is_filling_bind_table:
            return
        q = self.current_queue()
        if not q or not q.bindings:
            return
        enable = not all(bool(b.is_enabled) for b in q.bindings)
        for b in q.bindings:
            b.is_enabled = enable
        self.fill_bind_table(q)
        self.refresh_attr_tree(True)

    def _on_bind_cell_clicked(self, row: int, col: int):
        if self._is_filling_bind_table:
            return
        if col != 0:
            return
        q = self.current_queue()
        if not q:
            return
        if not (0 <= row < len(q.bindings)):
            return
        q.bindings[row].is_enabled = not q.bindings[row].is_enabled
        self.fill_bind_table(q)
        self._reselect_rows([row])

    def current_queue(self) -> Optional[BindingQueue]:
        if not self.selected_attr: return None
        return self.doc.ensure_attr(self.selected_attr)

    def add_selected_params_as_bindings(self):
        q = self.current_queue()
        if not q:
            QtWidgets.QMessageBox.warning(self, "Атрибут", "Выбери атрибут-лист слева."); return
        sels = self.tableParams.selectionModel().selectedRows()
        if not sels:
            QtWidgets.QMessageBox.warning(self, "Параметры", "Отметь строки справа."); return
        for m in sels:
            r = m.row()
            code = self.tableParams.item(r, 0).text()
            isnum = self.tableParams.item(r, 1).text() == "Число"
            models = self.param_models_map.get(r, [])
            enabled = self.chkAutoEnable.isChecked() if self.chkAutoEnable is not None else True
            b = Binding(
                parameter_code=code,
                src_is_numeric=isnum,
                is_enabled=enabled,
                src_model_title=", ".join(models)
            )
            q.bindings.append(b)
        self.fill_bind_table(q)
        self.refresh_attr_tree(True)

    def bind_delete(self):
        q = self.current_queue()
        if not q: return
        sels = sorted([m.row() for m in self.tableBindings.selectionModel().selectedRows()], reverse=True)
        for r in sels:
            if 0 <= r < len(q.bindings): q.bindings.pop(r)
        self.fill_bind_table(q); self.refresh_attr_tree(True)

    def bind_toggle_enabled(self):
        q = self.current_queue()
        if not q: return
        sels = [m.row() for m in self.tableBindings.selectionModel().selectedRows()]
        for r in sels:
            if 0 <= r < len(q.bindings):
                q.bindings[r].is_enabled = not q.bindings[r].is_enabled
        self.fill_bind_table(q)

    def bind_move_up(self):
        q = self.current_queue()
        if not q: return
        sels = sorted([m.row() for m in self.tableBindings.selectionModel().selectedRows()])
        if not sels or sels[0] == 0: return
        for r in sels: q.bindings[r-1], q.bindings[r] = q.bindings[r], q.bindings[r-1]
        self.fill_bind_table(q); self._reselect_rows([r-1 for r in sels])

    def bind_move_down(self):
        q = self.current_queue()
        if not q: return
        sels = sorted([m.row() for m in self.tableBindings.selectionModel().selectedRows()], reverse=True)
        if not sels or sels[0] >= len(q.bindings)-1: return
        for r in sels: q.bindings[r], q.bindings[r+1] = q.bindings[r+1], q.bindings[r]
        self.fill_bind_table(q); self._reselect_rows([r+1 for r in reversed(sels)])

    def _reselect_rows(self, rows: List[int]):
        sel = self.tableBindings.selectionModel(); sel.clearSelection()
        for r in rows:
            if 0 <= r < self.tableBindings.rowCount():
                self.tableBindings.selectRow(r)
        if rows:
            self.tableBindings.scrollTo(self.tableBindings.model().index(rows[0], 0))

    def _on_bind_context_menu(self, pos: QtCore.QPoint):
        idx = self.tableBindings.indexAt(pos)
        if not idx.isValid(): return
        self.tableBindings.selectRow(idx.row())
        q = self.current_queue()
        if not q: return
        r = idx.row()
        if not (0 <= r < len(q.bindings)): return
        b = q.bindings[r]
        dlg = TransformDialog(self, b)
        if dlg.exec() == QtWidgets.QDialog.Accepted:
            self.fill_bind_table(q)

    # drag&drop sync

    def open_transform_dialog(self):
        q = self.current_queue()
        if not q:
            return
        sels = [m.row() for m in self.tableBindings.selectionModel().selectedRows()]
        if not sels:
            QtWidgets.QMessageBox.warning(self, "Преобразования", "Выбери строку привязки.")
            return
        r = sels[0]
        if not (0 <= r < len(q.bindings)):
            return
        b = q.bindings[r]
        dlg = TransformDialog(self, b)
        if dlg.exec() == QtWidgets.QDialog.Accepted:
            self.fill_bind_table(q)

    # ---------- Импорт/Экспорт/Глобальные атрибуты ----------
    def load_global_attrs(self):
        try:
            gc = api_get_global_component(self.base(), comp_type=1)
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "API", f"Не удалось получить общие атрибуты:\n{e}")
            return
        pairs = flatten_global_attributes_with_types(gc)
        self.doc = AdapterDoc()
        self.attr_types.clear()
        self.selected_attr = None
        for full, isnum in pairs:
            if not full: continue
            self.doc.ensure_attr(full)
            self.attr_types[full] = isnum
        self.tableBindings.setRowCount(0)
        self.labelAttrSource.setText("Параметры загружены из общих атрибутов")
        self.refresh_attr_tree(True)
        QtWidgets.QMessageBox.information(self, "Общие атрибуты", f"Получено: {len(pairs)}")

    def import_xml(self):
        path, _ = QtWidgets.QFileDialog.getOpenFileName(self, "Открыть Adapter.xml", "", "Adapter XML (*.xml);;All files (*.*)")
        if not path: return
        try:
            self.doc = AdapterDoc.from_xml(path)
            self.attr_types.clear()
            self.selected_attr = None
            self.tableBindings.setRowCount(0)
            self.labelAttrSource.setText("Параметры загружены из импортированного файла")
            self.refresh_attr_tree(True)
            QtWidgets.QMessageBox.information(self, "Импорт", f"Загружено атрибутов: {len(self.doc.queues)}")
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "Импорт", f"Не удалось импортировать:\n{e}")

    def export_xml(self):
        if not self.doc.queues:
            QtWidgets.QMessageBox.warning(self, "Экспорт", "Нет данных для сохранения."); return
        path, _ = QtWidgets.QFileDialog.getSaveFileName(self, "Сохранить Adapter.xml", "Adapter.xml", "Adapter XML (*.xml);;All files (*.*)")
        if not path: return
        try:
            tree = self.doc.to_xml(); tree.write(path, encoding="utf-8", xml_declaration=True)
            QtWidgets.QMessageBox.information(self, "Экспорт", f"Сохранено:\n{path}")
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "Экспорт", f"Ошибка сохранения:\n{e}")

    def export_global_attrs_xml(self):
        names = sorted((self.doc.queues or {}).keys(), key=lambda s: (s or "").lower())
        if not names:
            QtWidgets.QMessageBox.warning(self, "Общие атрибуты", "Нет общих атрибутов для сохранения.")
            return

        path, _ = QtWidgets.QFileDialog.getSaveFileName(
            self,
            "Сохранить общие атрибуты.xml",
            "Общие атрибуты.xml",
            "XML (*.xml);;All files (*.*)",
        )
        if not path:
            return

        try:
            root = ET.Element(
                "EC.Entities.AttributeTree",
                {
                    "xmlns:xsd": "http://www.w3.org/2001/XMLSchema",
                    "xmlns:xsi": "http://www.w3.org/2001/XMLSchema-instance",
                },
            )
            sec = ET.SubElement(root, "Section")
            sections = ET.SubElement(sec, "Sections", {"Name": "IBIM", "IsComitted": "false"})
            ET.SubElement(sections, "Section")
            attrs = ET.SubElement(sections, "Attributes")
            for i, nm in enumerate(names, start=1):
                is_num = bool(self.attr_types.get(nm)) if self.attr_types else False
                ET.SubElement(
                    attrs,
                    "Attribute",
                    {
                        "Name": nm,
                        "IsComitted": "false",
                        "Id": str(i),
                        "Title": "",
                        "Description": "",
                        "Uom": "",
                        "IsNumeric": _b(is_num),
                        "ReportColumnType": "Text",
                    },
                )
            ET.SubElement(sec, "Attributes")

            tree = ET.ElementTree(root)
            tree.write(path, encoding="utf-8", xml_declaration=True)
            QtWidgets.QMessageBox.information(self, "Общие атрибуты", f"Сохранено:\n{path}")
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "Общие атрибуты", f"Ошибка сохранения:\n{e}")

def main():
    app = QtWidgets.QApplication(sys.argv)
    try:
        theme(app, load_saved_theme(False), icon_dir=ICON_DIR, persist=False)
        enable_theme_sync(app, ICON_DIR)
    except Exception:
        pass
    win = MainWin()
    win.show()
    apply_dark_titlebar(win)
    sys.exit(app.exec())
# === HOTFIX: импорт параметров из Excel (вставить перед if __name__ == "__main__":) ===

def _deamon_import_excel(self):
    from PySide6 import QtWidgets
    try:
        import openpyxl
    except Exception:
        QtWidgets.QMessageBox.critical(self, "Импорт Excel", "Не установлен openpyxl.\nУстанови: pip install openpyxl")
        return

    path, _ = QtWidgets.QFileDialog.getOpenFileName(self, "Открыть Excel", "", "Excel (*.xlsx *.xlsm)")
    if not path:
        return

    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active

        def _val(r, c):
            v = ws.cell(row=r, column=c).value
            return (str(v).strip() if v is not None else "")

        # Поиск строки заголовков
        header_row = None
        cols = {"name": None, "type": None, "list": None, "rep_from": None, "rep_to": None, "group": None}
        WANT = {
            "name": ["наименование параметра"],
            "type": ["тип параметра"],
            "list": ["список параметров"],
            "rep_from": ["замена с", "замена c"],
            "rep_to": ["замена на"],
            "group": ["группа параметров"],
        }
        for r in range(1, min(ws.max_row, 30) + 1):
            row_vals = [(_val(r, c)).lower() for c in range(1, ws.max_column + 1)]
            found = {}
            for key, variants in WANT.items():
                for c, txt in enumerate(row_vals, start=1):
                    if any(v in txt for v in variants):
                        found[key] = c
                        break
            if {"name", "type", "list"}.issubset(found.keys()):
                header_row = r
                for k, v in found.items():
                    cols[k] = v
                break
        if not header_row:
            raise RuntimeError("Не нашёл строку заголовков - ожидаю: Наименование параметра / Тип параметра / Список параметров.")

        # Группа из верхних ячеек (пример: A1='Укажите группу параметров:', B1='IBIM')
        default_group = ""
        for rr in range(1, min(ws.max_row, 8) + 1):
            for cc in range(1, min(ws.max_column, 8) + 1):
                raw = ws.cell(row=rr, column=cc).value
                if not raw:
                    continue
                s = str(raw).strip()
                low = s.lower()
                if "групп" in low:
                    val = None
                    if ":" in s:
                        after = s.split(":", 1)[1].strip()
                        val = after if after else None
                    if not val:
                        right = ws.cell(row=rr, column=cc+1).value
                        if right:
                            val = str(right).strip()
                    if val:
                        default_group = val
                        break
            if default_group:
                break
        
        
        
        def _split_list(s: str):
            # Строгий CSV: значения в двойных кавычках, запятая как разделитель.
            # Обратный слэш НЕ является escape-символом; чтобы вставить кавычку внутри значения, используйте удвоенную кавычку "".
            if not s:
                return []
            import csv, io
            reader = csv.reader(io.StringIO(s), delimiter=',', quotechar='"', doublequote=True, escapechar=None, skipinitialspace=True)
            try:
                row = next(reader)
                return [t for t in row if t is not None and t != '']
            except Exception:
                s = s.replace(';', ',')
                return [t.strip() for t in s.split(',') if t.strip()]
# Подготовка документа
        if not hasattr(self, "attr_types"):
            self.attr_types = {}
        if not hasattr(self, "attr_groups"):
            self.attr_groups = {}

        new_doc = AdapterDoc()
        self.attr_types.clear()
        self.attr_groups.clear()
        self.selected_attr = None

        r = header_row + 1
        while r <= ws.max_row:
            name = _val(r, cols["name"]) if cols["name"] else ""
            ptype = (_val(r, cols["type"]) if cols["type"] else "").lower()
            plist = _val(r, cols["list"]) if cols["list"] else ""
            rep_from = _val(r, cols["rep_from"]) if cols["rep_from"] else ""
            rep_to = _val(r, cols["rep_to"]) if cols["rep_to"] else ""
            grp = _val(r, cols["group"]) if cols["group"] else ""

            if not (name or plist or ptype or grp):
                r += 1
                continue

            isnum = True if "чис" in ptype else False
            full_name = f"{(grp or default_group)}.{name}" if (grp or default_group) else name
            q = new_doc.ensure_attr(full_name)
            self.attr_types[full_name] = isnum
            if (grp or default_group):
                self.attr_groups[full_name] = (grp or default_group)

            lefts = _split_list(rep_from)
            rights = _split_list(rep_to)
            pairs = [(l, rights[i]) for i, l in enumerate(lefts) if i < len(rights)]

            params = _split_list(plist)
            for code in params:
                b = Binding(parameter_code=code, src_is_numeric=None, is_enabled=True, src_model_title="Excel")
                if pairs:
                    t = TransformSettings()
                    t.replaces = pairs
                    b.transform = t
                q.bindings.append(b)
            r += 1

        self.doc = new_doc
        self.tableBindings.setRowCount(0)
        self.labelAttrSource.setText("Параметры загружены из Excel")

        # Обновляем дерево, выделяем первый атрибут и принудительно рисуем привязки
        self.refresh_attr_tree(False)

        selected_index = None
        for i in range(self.attrModel.rowCount()):
            root_item = self.attrModel.item(i, 0)
            if not root_item:
                continue
            data = root_item.data()
            if data and data[0] == "leaf":
                selected_index = root_item.index()
                break
            if root_item.hasChildren():
                child = root_item.child(0)
                if child:
                    cdata = child.data()
                    if cdata and cdata[0] == "leaf":
                        selected_index = child.index()
                        break
        if selected_index:
            self.treeAttributes.setCurrentIndex(selected_index)
            self.on_attr_pick()
            try:
                name_item = self.attrModel.itemFromIndex(selected_index)
                data = name_item.data() if name_item else None
                if data and data[0] == 'leaf':
                    q = self.doc.ensure_attr(data[1])
                    self.fill_bind_table(q)
            except Exception:
                pass
            if self.tableBindings.rowCount() > 0:
                self.tableBindings.selectRow(0)
                self.tableBindings.setFocus()

        QtWidgets.QMessageBox.information(self, "Импорт Excel", f"Загружено атрибутов: {len(self.doc.queues)}")
    except Exception as e:
        QtWidgets.QMessageBox.critical(self, "Импорт Excel", f"Не удалось импортировать Excel:\n{e}")
# Подмешиваем метод в класс, если его не было
MainWin.import_excel = _deamon_import_excel
# === /HOTFIX ===

if __name__ == "__main__":
    main()
