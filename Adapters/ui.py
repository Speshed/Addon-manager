# -*- coding: utf-8 -*-
from __future__ import annotations

import os
import sys
import json
import xml.etree.ElementTree as ET
from dataclasses import dataclass, field
from typing import Any, Dict, List, Optional, Tuple, Set

from PySide6 import QtCore, QtGui, QtWidgets

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from shared.theme_toggle import (
    ThemeToggle, theme, is_dark_theme, create_back_button, go_to_main_menu,
    resolve_icon_path, apply_dark_titlebar,
    load_saved_theme, enable_theme_sync,
    RowHoverDelegate, install_viewport_row_highlighter, setup_hover_tracking, PALETTE,
    _tint_pixmap,
)
from shared.dialogs import wire_dialog_button_box

# ----------------- Тема и логотип -----------------
BG = "#FFFFFF"
FG = "#222222"
ACCENT_ORANGE = "#F7921E"
ACCENT_ORANGE_HOVER = "#FFA74B"
BTN_GRAY = "#D9D9D9"
BTN_GRAY_HOVER = "#C9C9C9"
BIND_CHECKBOX_ICON_SIZE = 16

def _plugins_dir() -> str:
    # ...\Plugins\Manager_Adapter\Adapters\ui.py -> ...\Plugins
    return os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))


def _app_root_dir() -> str:
    return os.path.dirname(os.path.abspath(os.path.dirname(__file__)))


def _resolve_logo_path() -> str:
    candidates = [
        os.path.join(_app_root_dir(), "icon", "Manager-scaled.png"),
        os.path.join(_app_root_dir(), "icon", "logo.png"),
        os.path.join(os.path.dirname(__file__), "assets", "Manager-scaled.png"),
    ]
    for p in candidates:
        try:
            if p and os.path.exists(p):
                return p
        except Exception:
            continue
    return ""


LOGO_PATH = _resolve_logo_path()
ICON_DIR = os.path.join(_app_root_dir(), "icon")
TITLEBAR_ICON_PATH = os.path.join(_app_root_dir(), "icon", "logo.ico")
DEFAULT_BASE_URL = "http://localhost:5000"


def _runtime_base_url() -> str:
    return (os.environ.get("LARIX_API_BASE_URL") or DEFAULT_BASE_URL).rstrip("/")

_ICON_FALLBACK_FILES = {
    "ok": "ok.png",
    "circle": "circle.png",
    "check": "check.png",
    "select": "select.png",
    "poloska": "poloska.png",
}


def _resolve_icon_path(filename: str) -> str:
    if not filename:
        return ""
    candidates = [
        os.path.join(_app_root_dir(), "icon", filename),
        os.path.join(os.path.dirname(__file__), "assets", filename),
    ]
    for p in candidates:
        try:
            if p and os.path.exists(p):
                return p
        except Exception:
            continue
    return ""


class ErrorDialog(QtWidgets.QDialog):
    def __init__(self, parent, title: str, message: str):
        super().__init__(parent)
        self.setWindowTitle(title)
        self.setWindowFlags(self.windowFlags() & ~QtCore.Qt.WindowType.WindowContextHelpButtonHint)
        self.setMinimumWidth(320)
        vlayout = QtWidgets.QVBoxLayout(self)
        vlayout.setSpacing(16)
        vlayout.setContentsMargins(20, 20, 20, 20)
        hlayout = QtWidgets.QHBoxLayout()
        icon_label = QtWidgets.QLabel()
        app = QtWidgets.QApplication.instance()
        icon_path = resolve_icon_path("error", ICON_DIR, app=app, tint_in_dark=True)
        if os.path.exists(icon_path):
            pm = QtGui.QPixmap(icon_path)
            if not pm.isNull():
                icon_label.setPixmap(pm.scaled(48, 48, QtCore.Qt.AspectRatioMode.KeepAspectRatio, QtCore.Qt.TransformationMode.SmoothTransformation))
        if icon_label.pixmap() is None or icon_label.pixmap().isNull():
            style = QtWidgets.QApplication.style()
            icon_label.setPixmap(style.standardIcon(QtWidgets.QStyle.StandardPixmap.SP_MessageBoxWarning).pixmap(48, 48))
        hlayout.addWidget(icon_label, 0, QtCore.Qt.AlignmentFlag.AlignTop)
        msg_label = QtWidgets.QLabel(message)
        msg_label.setWordWrap(True)
        msg_label.setAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
        msg_label.setTextInteractionFlags(QtCore.Qt.TextInteractionFlag.TextSelectableByMouse)
        hlayout.addWidget(msg_label, 1)
        vlayout.addLayout(hlayout)
        btn_box = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.StandardButton.Ok)
        wire_dialog_button_box(btn_box, self.accept)
        vlayout.addWidget(btn_box, 0, QtCore.Qt.AlignmentFlag.AlignRight)


def show_error_dialog(parent, title: str, message: str):
    dlg = ErrorDialog(parent, title, message)
    dlg.exec()


class InfoDialog(QtWidgets.QDialog):
    def __init__(self, parent, title: str, message: str):
        super().__init__(parent)
        self.setWindowTitle(title)
        self.setWindowFlags(self.windowFlags() & ~QtCore.Qt.WindowType.WindowContextHelpButtonHint)
        self.setMinimumWidth(320)
        vlayout = QtWidgets.QVBoxLayout(self)
        vlayout.setSpacing(16)
        vlayout.setContentsMargins(20, 20, 20, 20)
        hlayout = QtWidgets.QHBoxLayout()
        icon_label = QtWidgets.QLabel()
        app = QtWidgets.QApplication.instance()
        icon_path = resolve_icon_path("alert", ICON_DIR, app=app, tint_in_dark=True)
        if icon_path and os.path.exists(icon_path):
            pm = QtGui.QPixmap(icon_path)
            if not pm.isNull():
                icon_label.setPixmap(pm.scaled(48, 48, QtCore.Qt.AspectRatioMode.KeepAspectRatio, QtCore.Qt.TransformationMode.SmoothTransformation))
        if icon_label.pixmap() is None or icon_label.pixmap().isNull():
            style = QtWidgets.QApplication.style()
            icon_label.setPixmap(style.standardIcon(QtWidgets.QStyle.StandardPixmap.SP_MessageBoxInformation).pixmap(48, 48))
        hlayout.addWidget(icon_label, 0, QtCore.Qt.AlignmentFlag.AlignTop)
        msg_label = QtWidgets.QLabel(message)
        msg_label.setWordWrap(True)
        msg_label.setAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
        msg_label.setTextInteractionFlags(QtCore.Qt.TextInteractionFlag.TextSelectableByMouse)
        hlayout.addWidget(msg_label, 1)
        vlayout.addLayout(hlayout)
        btn_box = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.StandardButton.Ok)
        wire_dialog_button_box(btn_box, self.accept)
        vlayout.addWidget(btn_box, 0, QtCore.Qt.AlignmentFlag.AlignRight)


def show_info_dialog(parent, title: str, message: str):
    dlg = InfoDialog(parent, title, message)
    dlg.exec()


def _pm(size: int) -> QtGui.QPixmap:
    pm = QtGui.QPixmap(size, size)
    pm.fill(QtCore.Qt.transparent)
    return pm


def _icon_ok(size: int = 16, color: str = ACCENT_ORANGE) -> QtGui.QIcon:
    pm = _pm(size)
    p = QtGui.QPainter(pm)
    try:
        p.setRenderHint(QtGui.QPainter.Antialiasing, True)
        pen = QtGui.QPen(QtGui.QColor(color), max(2, size // 8))
        pen.setCapStyle(QtCore.Qt.RoundCap)
        pen.setJoinStyle(QtCore.Qt.RoundJoin)
        p.setPen(pen)
        # simple check mark
        p.drawLine(int(size * 0.20), int(size * 0.55), int(size * 0.42), int(size * 0.75))
        p.drawLine(int(size * 0.42), int(size * 0.75), int(size * 0.80), int(size * 0.28))
    finally:
        p.end()
    return QtGui.QIcon(pm)


def _icon_circle(size: int = 16, border_color: str = "#B0B0B0") -> QtGui.QIcon:
    pm = _pm(size)
    p = QtGui.QPainter(pm)
    try:
        p.setRenderHint(QtGui.QPainter.Antialiasing, True)
        pen = QtGui.QPen(QtGui.QColor(border_color), max(2, size // 9))
        pen.setCapStyle(QtCore.Qt.RoundCap)
        p.setPen(pen)
        p.setBrush(QtCore.Qt.NoBrush)
        pad = max(2, size // 6)
        p.drawEllipse(pad, pad, size - 2 * pad, size - 2 * pad)
    finally:
        p.end()
    return QtGui.QIcon(pm)


def _icon_arrow(direction: str, size: int = 18, color: str = "#333333") -> QtGui.QIcon:
    pm = _pm(size)
    p = QtGui.QPainter(pm)
    try:
        p.setRenderHint(QtGui.QPainter.Antialiasing, True)
        pen = QtGui.QPen(QtGui.QColor(color), max(2, size // 9))
        pen.setCapStyle(QtCore.Qt.RoundCap)
        pen.setJoinStyle(QtCore.Qt.RoundJoin)
        p.setPen(pen)
        cx, cy = size / 2.0, size / 2.0
        r = size * 0.28
        if direction == "up":
            a = QtCore.QPointF(cx, cy - r)
            b = QtCore.QPointF(cx - r, cy + r)
            c = QtCore.QPointF(cx + r, cy + r)
        else:
            a = QtCore.QPointF(cx, cy + r)
            b = QtCore.QPointF(cx - r, cy - r)
            c = QtCore.QPointF(cx + r, cy - r)
        p.setBrush(QtGui.QBrush(QtGui.QColor(color)))
        p.drawPolygon(QtGui.QPolygonF([a, b, c]))
    finally:
        p.end()
    return QtGui.QIcon(pm)

def _icon_checkbox(checked: bool, size: int = 18, color: str = "#333333") -> QtGui.QIcon:
    pm = _pm(size)
    p = QtGui.QPainter(pm)
    try:
        p.setRenderHint(QtGui.QPainter.Antialiasing, True)
        pen = QtGui.QPen(QtGui.QColor(color), max(2, size // 10))
        pen.setCapStyle(QtCore.Qt.RoundCap)
        pen.setJoinStyle(QtCore.Qt.RoundJoin)
        p.setPen(pen)
        pad = max(2, size // 6)
        rect = QtCore.QRectF(pad, pad, size - 2 * pad, size - 2 * pad)
        p.setBrush(QtCore.Qt.NoBrush)
        p.drawRoundedRect(rect, 3, 3)
        if checked:
            p.drawLine(QtCore.QPointF(size * 0.25, size * 0.52), QtCore.QPointF(size * 0.44, size * 0.70))
            p.drawLine(QtCore.QPointF(size * 0.44, size * 0.70), QtCore.QPointF(size * 0.77, size * 0.30))
    finally:
        p.end()
    return QtGui.QIcon(pm)


def _icon_checkbox_indeterminate(size: int = 18, color: str = "#333333") -> QtGui.QIcon:
    pm = _pm(size)
    p = QtGui.QPainter(pm)
    try:
        p.setRenderHint(QtGui.QPainter.Antialiasing, True)
        pen = QtGui.QPen(QtGui.QColor(color), max(2, size // 10))
        pen.setCapStyle(QtCore.Qt.RoundCap)
        pen.setJoinStyle(QtCore.Qt.RoundJoin)
        p.setPen(pen)
        pad = max(2, size // 6)
        rect = QtCore.QRectF(pad, pad, size - 2 * pad, size - 2 * pad)
        p.setBrush(QtCore.Qt.NoBrush)
        p.drawRoundedRect(rect, 3, 3)
        y = size * 0.50
        p.drawLine(QtCore.QPointF(size * 0.28, y), QtCore.QPointF(size * 0.72, y))
    finally:
        p.end()
    return QtGui.QIcon(pm)


class FullRowTreeSelectionDelegate(QtWidgets.QStyledItemDelegate):
    @staticmethod
    def _row_rect(view: QtWidgets.QTreeView, option: QtWidgets.QStyleOptionViewItem) -> QtCore.QRectF:
        # Keep tree branch/expander area untouched so arrows stay visible on hover/selection.
        branch_pad = max(12, int(view.indentation()) - 4)
        x = max(2.0, float(branch_pad))
        return QtCore.QRectF(
            x,
            float(option.rect.top()) + 1.0,
            max(0.0, float(view.viewport().width()) - x - 2.0),
            max(0.0, float(option.rect.height()) - 2.0),
        )

    def paint(self, painter, option, index):
        opt = QtWidgets.QStyleOptionViewItem(option)
        self.initStyleOption(opt, index)
        is_selected = bool(opt.state & QtWidgets.QStyle.State_Selected)
        is_hovered = bool(opt.state & QtWidgets.QStyle.State_MouseOver)
        view = option.widget if isinstance(option.widget, QtWidgets.QTreeView) else None

        # Disable built-in per-cell states; we draw row backgrounds ourselves.
        opt.state &= ~QtWidgets.QStyle.State_MouseOver
        opt.state &= ~QtWidgets.QStyle.State_Selected
        opt.state &= ~QtWidgets.QStyle.State_HasFocus

        if is_hovered:
            if view is not None and index.column() == 0 and not is_selected:
                painter.save()
                painter.setRenderHint(QtGui.QPainter.Antialiasing, True)
                row_rect = self._row_rect(view, option)
                painter.setPen(QtCore.Qt.NoPen)
                painter.setBrush(QtGui.QColor(PALETTE.SOFT_HOVER))
                painter.drawRoundedRect(row_rect, 8.0, 8.0)
                painter.restore()

        if is_selected:
            if view is not None and index.column() == 0:
                painter.save()
                painter.setRenderHint(QtGui.QPainter.Antialiasing, True)
                row_rect = self._row_rect(view, option)
                painter.setPen(QtCore.Qt.NoPen)
                painter.setBrush(QtGui.QColor(PALETTE.SELECTED))
                painter.drawRoundedRect(row_rect, 8.0, 8.0)
                painter.restore()

        if is_selected or is_hovered:
            # Orange hover/selection background needs dark readable text in dark theme.
            pal = QtGui.QPalette(opt.palette)
            dark_text = QtGui.QColor("#000000")
            for role in (QtGui.QPalette.Text, QtGui.QPalette.WindowText, QtGui.QPalette.HighlightedText):
                pal.setColor(QtGui.QPalette.Active, role, dark_text)
                pal.setColor(QtGui.QPalette.Inactive, role, dark_text)
            opt.palette = pal

        style = opt.widget.style() if opt.widget is not None else QtWidgets.QApplication.style()
        style.drawControl(QtWidgets.QStyle.CE_ItemViewItem, opt, painter, opt.widget)


class BindingsCheckDelegate(QtWidgets.QStyledItemDelegate):
    def __init__(self, parent=None, icon_checked=None, icon_unchecked=None, icon_indeterminate=None):
        super().__init__(parent)
        self._icon_checked = icon_checked
        self._icon_unchecked = icon_unchecked
        self._icon_indeterminate = icon_indeterminate
        self._tint_cache: dict[tuple, QtGui.QPixmap] = {}

    def set_icons(self, icon_checked=None, icon_unchecked=None, icon_indeterminate=None):
        self._icon_checked = icon_checked
        self._icon_unchecked = icon_unchecked
        self._icon_indeterminate = icon_indeterminate
        self._tint_cache.clear()

    def _get_tinted_pixmap(self, icon: QtGui.QIcon, size: int, color: QtGui.QColor) -> QtGui.QPixmap:
        if icon is None or icon.isNull():
            return QtGui.QPixmap()
        cache_key = (id(icon), size, color.name())
        if cache_key in self._tint_cache:
            return self._tint_cache[cache_key]
        pm = icon.pixmap(size, size)
        if pm.isNull():
            return QtGui.QPixmap()
        tinted = _tint_pixmap(pm, color)
        self._tint_cache[cache_key] = tinted
        return tinted

    @staticmethod
    def _row_rect(view: QtWidgets.QTableView, option: QtWidgets.QStyleOptionViewItem) -> QtCore.QRectF:
        margin_h = 4.0
        margin_v = 1.0
        return QtCore.QRectF(
            margin_h,
            float(option.rect.top()) + margin_v,
            float(view.viewport().width()) - margin_h * 2.0,
            float(option.rect.height()) - margin_v * 2.0,
        )

    def paint(self, painter, option, index):
        opt = QtWidgets.QStyleOptionViewItem(option)
        self.initStyleOption(opt, index)
        
        view = option.widget if isinstance(option.widget, QtWidgets.QAbstractItemView) else None
        hover_row = -1
        if view is not None:
            try:
                val = view.property("_hover_row")
                hover_row = int(val) if val is not None else -1
            except Exception:
                hover_row = -1
        
        is_selected = bool(opt.state & QtWidgets.QStyle.State_Selected)
        is_hovered = bool(hover_row >= 0 and index.row() == hover_row)
        
        opt.state &= ~QtWidgets.QStyle.State_Selected
        opt.state &= ~QtWidgets.QStyle.State_MouseOver
        opt.state &= ~QtWidgets.QStyle.State_HasFocus
        opt.backgroundBrush = QtCore.Qt.NoBrush

        if view is not None and index.column() == 0:
            painter.save()
            painter.setRenderHint(QtGui.QPainter.Antialiasing, True)
            row_rect = self._row_rect(view, option)
            painter.setPen(QtCore.Qt.NoPen)
            
            if is_selected:
                painter.setBrush(QtGui.QColor(PALETTE.SELECTED))
                painter.drawRoundedRect(row_rect, 8.0, 8.0)
            elif is_hovered:
                painter.setBrush(QtGui.QColor(PALETTE.SOFT_HOVER))
                painter.drawRoundedRect(row_rect, 8.0, 8.0)
            
            painter.restore()

        if is_selected or is_hovered:
            pal = QtGui.QPalette(opt.palette)
            dark_text = QtGui.QColor("#000000")
            for role in (QtGui.QPalette.Text, QtGui.QPalette.WindowText, QtGui.QPalette.HighlightedText):
                pal.setColor(QtGui.QPalette.Active, role, dark_text)
                pal.setColor(QtGui.QPalette.Inactive, role, dark_text)
            opt.palette = pal

        style = opt.widget.style() if opt.widget is not None else QtWidgets.QApplication.style()

        if index.column() != 0:
            style.drawControl(QtWidgets.QStyle.CE_ItemViewItem, opt, painter, opt.widget)
            return

        is_checked = bool(index.data(QtCore.Qt.UserRole + 1))
        icon = self._icon_checked if is_checked else self._icon_unchecked
        
        opt.icon = QtGui.QIcon()
        opt.text = ""
        style.drawControl(QtWidgets.QStyle.CE_ItemViewItem, opt, painter, opt.widget)

        s = BIND_CHECKBOX_ICON_SIZE
        cell_rect = option.rect
        target = QtCore.QRect(
            cell_rect.x() + (cell_rect.width() - s) // 2,
            cell_rect.y() + (cell_rect.height() - s) // 2,
            s,
            s,
        )
        
        if icon is not None and not icon.isNull():
            if is_selected or is_hovered:
                pm = self._get_tinted_pixmap(icon, s, QtGui.QColor("#222222"))
            else:
                pm = icon.pixmap(s, s)
            if not pm.isNull():
                painter.drawPixmap(target, pm)
        else:
            fallback_color = "#222222" if (is_selected or is_hovered) else "#333333"
            fallback_icon = _icon_checkbox(is_checked, size=s, color=fallback_color)
            pm = fallback_icon.pixmap(s, s)
            if not pm.isNull():
                painter.drawPixmap(target, pm)


class BindingsHeaderView(QtWidgets.QHeaderView):
    def __init__(self, orientation: QtCore.Qt.Orientation, parent=None):
        super().__init__(orientation, parent)
        self._checkbox_icon = QtGui.QIcon()

    def set_checkbox_icon(self, icon: QtGui.QIcon) -> None:
        self._checkbox_icon = icon if isinstance(icon, QtGui.QIcon) else QtGui.QIcon()
        self.viewport().update()

    def paintSection(self, painter: QtGui.QPainter, rect: QtCore.QRect, logicalIndex: int) -> None:
        super().paintSection(painter, rect, logicalIndex)
        if logicalIndex != 0 or self._checkbox_icon.isNull():
            return
        s = QtCore.QSize(BIND_CHECKBOX_ICON_SIZE, BIND_CHECKBOX_ICON_SIZE)
        pm = self._checkbox_icon.pixmap(s)
        if pm.isNull():
            return
        target = QtCore.QRect(
            rect.x() + (rect.width() - s.width()) // 2,
            rect.y() + (rect.height() - s.height()) // 2,
            s.width(),
            s.height(),
        )
        painter.drawPixmap(target, pm)

# ----------------- HTTP helpers -------------------
try:
    import requests
except Exception:
    requests = None

def _need_requests():
    if requests is None:
        raise RuntimeError("Не установлен requests. Установи: pip install requests")

def _http_get(url: str, params=None, timeout=60):
    _need_requests()
    r = requests.get(url, headers={"accept": "application/json"}, params=params, timeout=timeout)
    if r.status_code >= 400:
        raise RuntimeError(f"GET {url} -> {r.status_code}\n{r.text}")
    try:
        return r.json()
    except Exception:
        return json.loads(r.text or "null")

def api_get_projects(base_url: str) -> List[Dict[str, Any]]:
    data = _http_get(f"{base_url.rstrip('/')}/api/project/projects") or []
    return [{"id": p.get("id"), "title": p.get("title") or p.get("name") or f"ID {p.get('id')}"} for p in data]

def api_get_containers(base_url: str, project_id: int) -> List[Dict[str, Any]]:
    data = _http_get(f"{base_url.rstrip('/')}/api/imcContainer/getProjectImcContainers/{project_id}") or []
    return [{"id": c.get("id"), "title": c.get("title") or f"ID {c.get('id')}"} for c in data]

def api_get_params_for_container(base_url: str, container_id: int) -> List[Dict[str, Any]]:
    data = _http_get(
        f"{base_url.rstrip('/')}/api/imcParameterDefinition/imcParameterDefinitions",
        params=[("containerIds", int(container_id))]
    ) or []
    rows = []
    for r in data:
        rows.append({
            "code": (r.get("code") or "").strip(),
            "isNumeric": bool(r.get("isNumeric")),
        })
    return rows

def api_get_global_component(base_url: str, comp_type: int = 1) -> dict:
    url = f"{base_url.rstrip('/')}/api/globalComponent/globalComponent/{int(comp_type)}"
    return _http_get(url) or {}

def flatten_global_attributes_with_types(gc: dict) -> List[Tuple[str, Optional[bool]]]:
    out: List[Tuple[str, Optional[bool]]] = []

    def walk_section(sec: dict, prefix: str = ""):
        name = (sec.get("Name") or "").strip()
        path = f"{prefix}.{name}" if prefix and name else (name or prefix)
        for a in (sec.get("Attributes") or []):
            nm = (a.get("Name") or "").strip()
            if not nm:
                continue
            isnum = a.get("IsNumeric")
            full = f"{path}.{nm}" if path else nm
            out.append((full, bool(isnum) if isnum is not None else None))
        for s in (sec.get("Sections") or []):
            walk_section(s, path)

    content = (gc.get("content") or {})
    for a in (content.get("Attributes") or []):
        nm = (a.get("Name") or "").strip()
        if nm:
            isnum = a.get("IsNumeric")
            out.append((nm, bool(isnum) if isnum is not None else None))
    for sec in (content.get("Sections") or []):
        walk_section(sec, "")
    # уникализация
    seen = {}
    for name, t in out:
        if name not in seen:
            seen[name] = t
    return sorted(seen.items(), key=lambda x: x[0].lower())

# ----------------- Модель адаптера -----------------
def _b(v: bool) -> str:
    return "true" if v else "false"

@dataclass
class TransformSettings:
    trim: str = ""               # "", "Left", "Right", "Both"
    case: str = ""               # "", "Upper", "Lower", "Title"
    prepare_numeric: bool = False
    replaces: List[Tuple[str, str]] = field(default_factory=list)

def _next_uid():
    _next_uid._c = getattr(_next_uid, '_c', 0) + 1
    return _next_uid._c

@dataclass
class Binding:
    parameter_code: str
    uid: int = field(default_factory=_next_uid)
    src_is_numeric: Optional[bool] = None
    is_enabled: bool = True
    src_model_title: str = ""
    transform: TransformSettings = field(default_factory=TransformSettings)

@dataclass
class BindingQueue:
    attribute_full_name: str
    default_is_enabled: bool = True
    bindings: List[Binding] = field(default_factory=list)

class AdapterDoc:
    def __init__(self):
        self.queues: Dict[str, BindingQueue] = {}

    def ensure_attr(self, name: str) -> BindingQueue:
        k = (name or "").strip()
        if not k:
            raise ValueError("Имя атрибута пустое")
        if k not in self.queues:
            self.queues[k] = BindingQueue(attribute_full_name=k)
        return self.queues[k]

    def list_attrs(self) -> List[BindingQueue]:
        return [self.queues[k] for k in sorted(self.queues.keys(), key=str.lower)]

    def to_xml(self) -> ET.ElementTree:
        root = ET.Element("EC.Entities.Adapter", {
            "xmlns:xsd": "http://www.w3.org/2001/XMLSchema",
            "xmlns:xsi": "http://www.w3.org/2001/XMLSchema-instance",
        })
        bqs = ET.SubElement(root, "BindingQueues")
        for q in self.list_attrs():
            q_el = ET.SubElement(bqs, "BindingQueue", {
                "AttributeFullName": q.attribute_full_name,
                "DefaultIsEnabled": _b(q.default_is_enabled),
            })
            binds_el = ET.SubElement(q_el, "Bindings")
            for idx, b in enumerate(q.bindings):
                t = getattr(b, "transform", None)
                attrs = {
                    "Index": str(idx),
                    "IsEnabled": _b(b.is_enabled),
                    "ParameterCode": b.parameter_code,
                    "TransformPrepareTextForNumeric": _b(getattr(t, "prepare_numeric", False)),
                }
                if t and getattr(t, "trim", ""):
                    attrs["TransformTrim"] = t.trim
                if t and getattr(t, "case", ""):
                    attrs["TransformCase"] = t.case
                bind_el = ET.SubElement(binds_el, "Binding", attrs)

                repls_el = ET.SubElement(bind_el, "TransformReplaces")
                if t:
                    for src, dst in getattr(t, "replaces", []) or []:
                        ET.SubElement(repls_el, "TransformReplace", {
                            "ReplaceableText": src,
                            "ReplacingText": dst
                        })
        return ET.ElementTree(root)

    @classmethod
    def from_xml(cls, path: str) -> "AdapterDoc":
        tree = ET.parse(path)
        root = tree.getroot()
        doc = cls()
        bqs = root.find("BindingQueues")
        if bqs is None:
            return doc
        for q_el in bqs.findall("BindingQueue"):
            attr = q_el.get("AttributeFullName", "").strip()
            q = doc.ensure_attr(attr)
            q.default_is_enabled = (q_el.get("DefaultIsEnabled", "false").lower() == "true")
            binds_el = q_el.find("Bindings")
            if not binds_el:
                continue
            pairs: List[Tuple[int, Binding]] = []
            for b_el in binds_el.findall("Binding"):
                try:
                    idx = int(b_el.get("Index", "0"))
                except Exception:
                    idx = 0
                t = TransformSettings()
                t.prepare_numeric = (b_el.get("TransformPrepareTextForNumeric", "false").lower() == "true")
                t.trim = (b_el.get("TransformTrim") or "").strip()
                t.case = (b_el.get("TransformCase") or "").strip()

                repls: List[Tuple[str, str]] = []
                repls_el = b_el.find("TransformReplaces")
                if repls_el is not None:
                    for r in repls_el.findall("TransformReplace"):
                        src = r.get("ReplaceableText", "") or ""
                        dst = r.get("ReplacingText", "") or ""
                        repls.append((src, dst))
                t.replaces = repls

                b = Binding(
                    parameter_code=b_el.get("ParameterCode", ""),
                    is_enabled=(b_el.get("IsEnabled", "true").lower() == "true"),
                    src_is_numeric=None,
                    src_model_title="",
                    transform=t
                )
                pairs.append((idx, b))
            pairs.sort(key=lambda t: t[0])
            q.bindings = [p[1] for p in pairs]
        return doc

# ----------------- Диалоги -----------------
class TransformDialog(QtWidgets.QDialog):
    def __init__(self, parent, binding: Binding):
        super().__init__(parent)
        self.setWindowTitle("Настройка преобразований")
        self.binding = binding
        self.resize(700, 480)

        v = QtWidgets.QVBoxLayout(self)

        row1 = QtWidgets.QHBoxLayout()
        v.addLayout(row1)
        row1.addWidget(QtWidgets.QLabel("Пробелы:"))
        self.cmbTrim = QtWidgets.QComboBox()
        self.cmbTrim.addItems(["Без изменений", "Удалить слева", "Удалить справа", "Удалить с обеих сторон"])
        row1.addWidget(self.cmbTrim)

        row2 = QtWidgets.QHBoxLayout()
        v.addLayout(row2)
        row2.addWidget(QtWidgets.QLabel("Регистр:"))
        self.cmbCase = QtWidgets.QComboBox()
        self.cmbCase.addItems(["Без изменений", "Привести к верхнему", "Привести к нижнему", "Заголовок"])
        row2.addWidget(self.cmbCase)

        v.addWidget(QtWidgets.QLabel("Сопоставление строк"))
        self.table = QtWidgets.QTableWidget(0, 2)
        self.table.setHorizontalHeaderLabels(["Замена с", "Замена на"])
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.verticalHeader().setVisible(False)
        self.table.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)

        v.addWidget(self.table)

        tb = QtWidgets.QHBoxLayout()
        v.addLayout(tb)
        self.btnAdd = QtWidgets.QPushButton("+")
        self.btnDel = QtWidgets.QPushButton("-")
        tb.addWidget(self.btnAdd)
        tb.addWidget(self.btnDel)
        tb.addStretch(1)

        bb = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel)
        v.addWidget(bb)

        self.btnAdd.clicked.connect(self.add_row)
        self.btnDel.clicked.connect(self.del_rows)
        wire_dialog_button_box(bb, self.accept_and_save, self.reject)

        trim_map_rev = {"": "Без изменений", "Left":"Удалить слева", "Right":"Удалить справа", "Both":"Удалить с обеих сторон"}
        case_map_rev = {"": "Без изменений", "Upper":"Привести к верхнему", "Lower":"Привести к нижнему", "Title":"Заголовок"}
        self.cmbTrim.setCurrentText(trim_map_rev.get(binding.transform.trim or "", "Без изменений"))
        self.cmbCase.setCurrentText(case_map_rev.get(binding.transform.case or "", "Без изменений"))
        for src, dst in (binding.transform.replaces or []):
            self.add_row(src, dst)

        # Подсветка латиницы/кириллицы
        self.table.itemChanged.connect(self._recolor_rows)
        self._recolor_rows()

    def _recolor_rows(self):
        import re as _re
        lat = _re.compile(r"[A-Za-z]")
        cyr = _re.compile(r"[А-Яа-яЁё]")
        for r in range(self.table.rowCount()):
            for c in (0,1):
                it = self.table.item(r,c)
                if not it:
                    continue
                txt = it.text() or ""
                has_lat = bool(lat.search(txt))
                has_cyr = bool(cyr.search(txt))
                if has_lat and not has_cyr:
                    it.setForeground(QtGui.QBrush(QtGui.QColor('#1f6feb')))  # латиница — синяя
                elif has_cyr and not has_lat:
                    it.setForeground(QtGui.QBrush(QtGui.QColor('#d73a49')))  # кириллица — красная
                elif has_lat and has_cyr:
                    it.setForeground(QtGui.QBrush(QtGui.QColor('#8a2be2')))  # смешанный — фиолетовый
                else:
                    it.setForeground(QtGui.QBrush())

    def add_row(self, src: str = "", dst: str = ""):
        r = self.table.rowCount()
        self.table.insertRow(r)
        self.table.setItem(r, 0, QtWidgets.QTableWidgetItem(src))
        self.table.setItem(r, 1, QtWidgets.QTableWidgetItem(dst))
        self._recolor_rows()

    def del_rows(self):
        sel = self.table.selectionModel().selectedRows()
        for m in reversed(sel):
            self.table.removeRow(m.row())
        self._recolor_rows()

    def accept_and_save(self):
        trim_map = {"Без изменений":"", "Удалить слева":"Left", "Удалить справа":"Right", "Удалить с обеих сторон":"Both"}
        case_map = {"Без изменений":"", "Привести к верхнему":"Upper", "Привести к нижнему":"Lower", "Заголовок":"Title"}
        self.binding.transform.trim = trim_map.get(self.cmbTrim.currentText(), "")
        self.binding.transform.case = case_map.get(self.cmbCase.currentText(), "")
        repl = []
        for r in range(self.table.rowCount()):
            s = (self.table.item(r,0).text() if self.table.item(r,0) else "").strip()
            d = (self.table.item(r,1).text() if self.table.item(r,1) else "").strip()
            if s or d:
                repl.append((s,d))
        self.binding.transform.replaces = repl
        self.accept()

class ModelFilterDialog(QtWidgets.QDialog):
    def __init__(self, parent, models: List[str], selected: Optional[Set[str]]):
        super().__init__(parent)
        self.setWindowTitle("Фильтр по моделям")
        self.resize(520, 600)
        v = QtWidgets.QVBoxLayout(self)

        h = QtWidgets.QHBoxLayout()
        btn_all = QtWidgets.QPushButton("Выделить все")
        btn_none = QtWidgets.QPushButton("Снять все")
        h.addWidget(btn_all); h.addWidget(btn_none); h.addStretch(1)
        v.addLayout(h)

        self.list = QtWidgets.QListWidget()
        self.list.setSelectionMode(QtWidgets.QAbstractItemView.NoSelection)
        v.addWidget(self.list)

        sel = set(models) if selected is None else set(selected)
        for m in models:
            it = QtWidgets.QListWidgetItem(m)
            it.setFlags(it.flags() | QtCore.Qt.ItemIsUserCheckable)
            it.setCheckState(QtCore.Qt.Checked if m in sel else QtCore.Qt.Unchecked)
            self.list.addItem(it)

        btn_all.clicked.connect(lambda: self._set_all(QtCore.Qt.Checked))
        btn_none.clicked.connect(lambda: self._set_all(QtCore.Qt.Unchecked))

        bb = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel)
        v.addWidget(bb)
        wire_dialog_button_box(bb, self.accept, self.reject)

    def _set_all(self, state):
        for i in range(self.list.count()):
            it = self.list.item(i); it.setCheckState(state)

    def chosen(self) -> List[str]:
        out = []
        for i in range(self.list.count()):
            it = self.list.item(i)
            if it.checkState() == QtCore.Qt.Checked:
                out.append(it.text())
        return out

# ----------------- Главное окно -----------------

class RowDragTable(QtWidgets.QTableWidget):
    """Таблица с перетаскиванием СТРОК (только по вертикали), без «ломания» ячеек."""
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        self.setSelectionMode(QtWidgets.QAbstractItemView.ExtendedSelection)
        self.setDragEnabled(True)
        self.setAcceptDrops(True)
        self.setDropIndicatorShown(True)
        self.setDragDropMode(QtWidgets.QAbstractItemView.DragDrop)  # без InternalMove
        self.setDefaultDropAction(QtCore.Qt.MoveAction)
        self.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        self._drag_rows = []   # внутреннее хранилище
        self.on_rows_moved = None  # callback(List[int] src_rows_sorted, int dest_row)
        self._drop_indicator_row = -1  # строка для подсветки вставки

    def startDrag(self, supportedActions):
        # фиксируем исходные строки
        rows = sorted({i.row() for i in self.selectedIndexes()})
        if not rows:
            return
        self._drag_rows = rows
        drag = QtGui.QDrag(self)
        mime = QtCore.QMimeData()
        mime.setData("application/x-rows", b",".join(str(r).encode("utf-8") for r in rows))
        drag.setMimeData(mime)
        # простая картинка для dnd
        pm = QtGui.QPixmap(self.viewport().visibleRegion().boundingRect().size())
        pm.fill(QtCore.Qt.transparent)
        painter = QtGui.QPainter(pm)
        # рисуем первую переносимую строку как превью
        if rows:
            r = rows[0]
            rect = self.visualRect(self.model().index(r, 0))
            rect = rect.adjusted(0, 0, self.viewport().width()-1, rect.height())
            painter.fillRect(rect, QtGui.QBrush(QtGui.QColor(0, 0, 0, 30)))
        painter.end()
        drag.setPixmap(pm)
        drag.exec(QtCore.Qt.MoveAction)

    def dragEnterEvent(self, e: QtGui.QDragEnterEvent):
        if e.mimeData().hasFormat("application/x-rows"):
            e.acceptProposedAction()
        else:
            e.ignore()

    def dragMoveEvent(self, e: QtGui.QDragMoveEvent):
        if e.mimeData().hasFormat("application/x-rows"):
            pos = e.position().toPoint() if hasattr(e, "position") else e.pos()
            dest_row = self.rowAt(pos.y())
            if dest_row < 0:
                dest_row = self.rowCount()
            if dest_row != self._drop_indicator_row:
                self._drop_indicator_row = dest_row
                self.viewport().update()
            e.acceptProposedAction()
        else:
            self._drop_indicator_row = -1
            self.viewport().update()
            e.ignore()

    def dragLeaveEvent(self, e):
        self._drop_indicator_row = -1
        self.viewport().update()
        super().dragLeaveEvent(e)

    def dropEvent(self, e: QtGui.QDropEvent):
        self._drop_indicator_row = -1
        self.viewport().update()
        if not e.mimeData().hasFormat("application/x-rows"):
            e.ignore(); return
        # вычисляем строку вставки
        pos = e.position().toPoint() if hasattr(e, "position") else e.pos()
        dest_row = self.rowAt(pos.y())
        if dest_row < 0:
            dest_row = self.rowCount()
        rows = self._drag_rows or []
        if not rows:
            e.ignore(); return
        # вызовем обратный вызов у владельца
        if callable(self.on_rows_moved):
            self.on_rows_moved(rows, dest_row)
        e.acceptProposedAction()

    def paintEvent(self, event):
        super().paintEvent(event)
        if self._drop_indicator_row >= 0:
            painter = QtGui.QPainter(self.viewport())
            painter.setRenderHint(QtGui.QPainter.Antialiasing, True)
            pen = QtGui.QPen(QtGui.QColor(ACCENT_ORANGE), 3)
            pen.setCapStyle(QtCore.Qt.RoundCap)
            painter.setPen(pen)
            row = self._drop_indicator_row
            if row < self.rowCount():
                rect = self.visualRect(self.model().index(row, 0))
                y = rect.top()
            else:
                if self.rowCount() > 0:
                    last_rect = self.visualRect(self.model().index(self.rowCount() - 1, 0))
                    y = last_rect.bottom()
                else:
                    y = self.viewport().height() - 2
            x1 = 4
            x2 = self.viewport().width() - 4
            painter.drawLine(x1, y, x2, y)
            arrow_size = 8
            painter.setBrush(QtGui.QBrush(QtGui.QColor(ACCENT_ORANGE)))
            painter.drawLine(x1, y - arrow_size, x1, y + arrow_size)
            painter.drawLine(x2, y - arrow_size, x2, y + arrow_size)


class MainWin(QtWidgets.QMainWindow):
    def _build_ui(self):
        self.ui = QtWidgets.QWidget(self)
        self.ui.setObjectName("centralwidget")
        self.setCentralWidget(self.ui)

        self.verticalLayout_central = QtWidgets.QVBoxLayout(self.ui)
        self.verticalLayout_central.setObjectName("verticalLayout_central")
        self.verticalLayout_central.setSpacing(8)
        self.verticalLayout_central.setContentsMargins(10, 10, 10, 10)

        # Header with back button and theme toggle
        header_h_layout = QtWidgets.QHBoxLayout()
        header_h_layout.setContentsMargins(0, 0, 0, 0)
        self._btn_back = create_back_button(self, icon_dir=ICON_DIR)
        self._btn_back.clicked.connect(lambda: go_to_main_menu(self))
        header_h_layout.addWidget(self._btn_back)
        header_h_layout.addStretch()
        self._theme_toggle = ThemeToggle(self)
        self._theme_toggle.setChecked(is_dark_theme(QtWidgets.QApplication.instance()))
        self._theme_toggle.toggled.connect(self._on_theme_toggled)
        header_h_layout.addWidget(self._theme_toggle)
        self.verticalLayout_central.addLayout(header_h_layout)

        # Top bar
        layout_top = QtWidgets.QHBoxLayout()
        layout_top.setObjectName("layoutTopBar")
        self.btnImportXml = QtWidgets.QPushButton("Импорт XML", self.ui)
        self.btnImportXml.setObjectName("btnImportXml")
        self.btnImportXml.setToolTip("Импортировать существующий Adapter.xml")
        self.btnImportExcel = QtWidgets.QPushButton("Импорт Excel", self.ui)
        self.btnImportExcel.setObjectName("btnImportExcel")
        self.btnLoadGlobal = QtWidgets.QPushButton("Загрузить общие атрибуты", self.ui)
        self.btnLoadGlobal.setObjectName("btnLoadGlobal")
        self.btnModelsWindow = QtWidgets.QPushButton("Выбор моделей", self.ui)
        self.btnModelsWindow.setObjectName("btnModelsWindow")
        self.comboProject = QtWidgets.QComboBox(self.ui)
        self.comboProject.setObjectName("comboProject")
        self.comboProject.setSizeAdjustPolicy(QtWidgets.QComboBox.AdjustToContentsOnFirstShow)
        self.comboProject.hide()

        layout_top.addWidget(self.btnModelsWindow)
        layout_top.addWidget(self.btnImportXml)
        layout_top.addWidget(self.btnImportExcel)
        layout_top.addWidget(self.btnLoadGlobal)
        layout_top.addStretch(1)
        self.verticalLayout_central.addLayout(layout_top)
        
        self._update_top_bar_icons()

        # Main split with resizable sections
        self.splitterMain = QtWidgets.QSplitter(QtCore.Qt.Horizontal, self.ui)
        self.splitterMain.setObjectName("splitterMain")
        self.splitterMain.setHandleWidth(6)
        self.splitterMain.setChildrenCollapsible(False)

        # Attributes
        self.groupAttributes = QtWidgets.QGroupBox("Атрибуты", self.ui)
        self.groupAttributes.setObjectName("groupAttributes")
        layout_attrs = QtWidgets.QVBoxLayout(self.groupAttributes)
        layout_attrs.setObjectName("layoutAttributes")

        layout_attr_search = QtWidgets.QHBoxLayout()
        layout_attr_search.setObjectName("layoutAttrSearch")
        self.labelAttrSearch = QtWidgets.QLabel("Поиск:", self.groupAttributes)
        self.labelAttrSearch.setObjectName("labelAttrSearch")
        self.editAttrSearch = QtWidgets.QLineEdit(self.groupAttributes)
        self.editAttrSearch.setObjectName("editAttrSearch")
        layout_attr_search.addWidget(self.labelAttrSearch)
        layout_attr_search.addWidget(self.editAttrSearch)
        layout_attrs.addLayout(layout_attr_search)

        self.treeAttributes = QtWidgets.QTreeView(self.groupAttributes)
        self.treeAttributes.setObjectName("treeAttributes")
        self.treeAttributes.setFrameShape(QtWidgets.QFrame.NoFrame)
        self.treeAttributes.setHeaderHidden(True)
        self.treeAttributes.setSelectionMode(QtWidgets.QAbstractItemView.ExtendedSelection)
        try:
            self.treeAttributes.setUniformRowHeights(True)
        except Exception:
            pass
        layout_attrs.addWidget(self.treeAttributes)

        self.labelAttrSource = QtWidgets.QLabel("Параметры не загружены", self.groupAttributes)
        self.labelAttrSource.setObjectName("labelAttrSource")
        layout_attrs.addWidget(self.labelAttrSource)

        # Bindings
        self.groupBindings = QtWidgets.QGroupBox("Привязки атрибута", self.ui)
        self.groupBindings.setObjectName("groupBindings")
        layout_bind = QtWidgets.QVBoxLayout(self.groupBindings)
        layout_bind.setObjectName("layoutBindings")

        self.tableBindings = RowDragTable(self.groupBindings)
        self.tableBindings.setObjectName("tableBindings")
        self.tableBindings.setFrameShape(QtWidgets.QFrame.NoFrame)
        self._bind_header = BindingsHeaderView(QtCore.Qt.Horizontal, self.tableBindings)
        self.tableBindings.setHorizontalHeader(self._bind_header)
        layout_bind.addWidget(self.tableBindings)

        layout_bind_btns = QtWidgets.QHBoxLayout()
        layout_bind_btns.setObjectName("layoutBindButtons")
        self.btnBindDelete = QtWidgets.QPushButton("Удалить выбранные", self.groupBindings)
        self.btnBindDelete.setObjectName("btnBindDelete")
        self.btnBindToggle = None
        self.btnTransform = QtWidgets.QPushButton("Преобразования", self.groupBindings)
        self.btnTransform.setObjectName("btnTransform")
        self.btnUp = QtWidgets.QPushButton("▲", self.groupBindings)
        self.btnUp.setObjectName("btnUp")
        self.btnUp.setMinimumWidth(36)
        self.btnDown = QtWidgets.QPushButton("▼", self.groupBindings)
        self.btnDown.setObjectName("btnDown")
        self.btnDown.setMinimumWidth(36)

        layout_bind_btns.addWidget(self.btnBindDelete)
        layout_bind_btns.addWidget(self.btnTransform)
        layout_bind_btns.addStretch(1)
        layout_bind_btns.addWidget(self.btnUp)
        layout_bind_btns.addWidget(self.btnDown)
        layout_bind.addLayout(layout_bind_btns)

        # Params
        self.groupParams = QtWidgets.QGroupBox("Параметры моделей", self.ui)
        self.groupParams.setObjectName("groupParams")
        layout_params = QtWidgets.QVBoxLayout(self.groupParams)
        layout_params.setObjectName("layoutParams")

        layout_param_filter = QtWidgets.QHBoxLayout()
        layout_param_filter.setObjectName("layoutParamFilter")
        self.labelFilter = QtWidgets.QLabel("Фильтр по шифру:", self.groupParams)
        self.labelFilter.setObjectName("labelFilter")
        self.editParamFilter = QtWidgets.QLineEdit(self.groupParams)
        self.editParamFilter.setObjectName("editParamFilter")
        layout_param_filter.addWidget(self.labelFilter)
        layout_param_filter.addWidget(self.editParamFilter)
        layout_param_filter.addStretch(1)
        layout_params.addLayout(layout_param_filter)

        self.tableParams = QtWidgets.QTableWidget(self.groupParams)
        self.tableParams.setObjectName("tableParams")
        self.tableParams.setFrameShape(QtWidgets.QFrame.NoFrame)
        layout_params.addWidget(self.tableParams)

        self.splitterMain.addWidget(self.groupAttributes)
        self.splitterMain.addWidget(self.groupBindings)
        self.splitterMain.addWidget(self.groupParams)
        self.splitterMain.setSizes([300, 400, 300])
        self.verticalLayout_central.addWidget(self.splitterMain, 1)

        # Bottom
        layout_bottom = QtWidgets.QHBoxLayout()
        layout_bottom.setObjectName("layoutBottom")
        layout_bottom.addStretch(1)
        self.btnSave = QtWidgets.QPushButton("Сохранить адаптер", self.ui)
        self.btnSave.setObjectName("btnSave")
        self.btnSave.setMinimumWidth(320)
        self.btnSaveGlobal = QtWidgets.QPushButton("Сохранить общие атрибуты", self.ui)
        self.btnSaveGlobal.setObjectName("btnSaveGlobal")
        self.btnSaveGlobal.setMinimumWidth(320)
        layout_bottom.addWidget(self.btnSave)
        layout_bottom.addWidget(self.btnSaveGlobal)
        layout_bottom.addStretch(1)
        self.verticalLayout_central.addLayout(layout_bottom)

        self.lblStatus = QtWidgets.QLabel("", self.ui)
        self.lblStatus.setObjectName("lblStatus")
        self.verticalLayout_central.addWidget(self.lblStatus)

    def __init__(self):
        super().__init__()
        self._pending_theme_apply = False
        self._build_ui()

        self.setWindowTitle("Larix — Адаптеры")
        self.resize(1500, 860)
        self.setMinimumSize(1160, 720)

        self.doc = AdapterDoc()
        self.projects: List[Dict[str, Any]] = []
        self.containers: List[Dict[str, Any]] = []
        self.per_model_params: Dict[int, List[Dict[str, Any]]] = {}
        self.selected_attr: Optional[str] = None
        self.attr_types: Dict[str, Optional[bool]] = {}
        self.param_models_map: Dict[int, List[str]] = {}
        self.bind_models_map: Dict[int, List[str]] = {}
        self._is_filling_bind_table: bool = False
        self._icon_bind_checked: Optional[QtGui.QIcon] = None
        self._icon_bind_unchecked: Optional[QtGui.QIcon] = None
        self._icon_bind_indeterminate: Optional[QtGui.QIcon] = None

        # Фильтры "Параметры моделей"
        self.type_filter: Optional[bool] = None     # None=оба, True=число, False=текст
        self.model_filter_set: Optional[Set[str]] = None  # None=все модели, иначе только выбранные

        # (виджет удалён из интерфейса, но логика поддерживает значение по умолчанию)
        self.chkAutoEnable = None

        # Модель для дерева атрибутов
        self.attrModel: QtGui.QStandardItemModel = QtGui.QStandardItemModel(0, 2, self.treeAttributes)
        self.treeAttributes.setModel(self.attrModel)
        self.attrModel.setHorizontalHeaderLabels(["Атрибуты", ""])
        self.treeAttributes.header().hide()
        try:
            self.treeAttributes.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
            self.treeAttributes.setAllColumnsShowFocus(False)
            self.treeAttributes.setItemDelegate(FullRowTreeSelectionDelegate(self.treeAttributes))
        except Exception:
            pass
        try:
            hdr = self.treeAttributes.header()
            hdr.setStretchLastSection(False)
            hdr.setSectionResizeMode(0, QtWidgets.QHeaderView.Stretch)
            hdr.setSectionResizeMode(1, QtWidgets.QHeaderView.Fixed)
            self.treeAttributes.setColumnWidth(1, 28)
        except Exception:
            pass

        # иконки статуса атрибутов (справа)
        try:
            app = QtWidgets.QApplication.instance()
            ok_path = resolve_icon_path("ok", ICON_DIR, app=app) or _resolve_icon_path(_ICON_FALLBACK_FILES["ok"])
            circle_path = resolve_icon_path("circle", ICON_DIR, app=app) or _resolve_icon_path(_ICON_FALLBACK_FILES["circle"])
            self._icon_attr_assigned = QtGui.QIcon(ok_path) if ok_path else _icon_ok()
            self._icon_attr_unassigned = QtGui.QIcon(circle_path) if circle_path else _icon_circle()
        except Exception:
            self._icon_attr_assigned = _icon_ok()
            self._icon_attr_unassigned = _icon_circle()

        # Иконки чекбоксов (как в других окнах: unchecked=check, checked=select)
        self._reload_bind_checkbox_icons(is_dark_theme(QtWidgets.QApplication.instance()))

        # Таблицы
        self.tableBindings.setColumnCount(4)
        self.tableBindings.setHorizontalHeaderLabels(["", "Имя параметра", "Преобразование", "Модель"])
        for col in (1, 2, 3):
            h_item = self.tableBindings.horizontalHeaderItem(col)
            if h_item:
                h_item.setTextAlignment(QtCore.Qt.AlignmentFlag.AlignLeft | QtCore.Qt.AlignmentFlag.AlignVCenter)
        self.tableBindings.horizontalHeader().setStretchLastSection(False)
        self.tableBindings.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        self.tableBindings.setSelectionMode(QtWidgets.QAbstractItemView.ExtendedSelection)
        self.tableBindings.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        try:
            self.tableBindings.setShowGrid(False)
        except Exception:
            pass
        # Размеры столбцов привязок: чекбокс фиксирован, Имя параметра растягивается, остальные по содержимому
        self.tableBindings.horizontalHeader().setSectionResizeMode(0, QtWidgets.QHeaderView.Fixed)
        self.tableBindings.horizontalHeader().setSectionResizeMode(1, QtWidgets.QHeaderView.Stretch)
        self.tableBindings.horizontalHeader().setSectionResizeMode(2, QtWidgets.QHeaderView.ResizeToContents)
        self.tableBindings.horizontalHeader().setSectionResizeMode(3, QtWidgets.QHeaderView.ResizeToContents)
        self.tableBindings.setColumnWidth(0, 40)
        self.tableBindings.horizontalHeader().setMinimumSectionSize(40)
        try:
            self.tableBindings.setIconSize(QtCore.QSize(BIND_CHECKBOX_ICON_SIZE, BIND_CHECKBOX_ICON_SIZE))
        except Exception:
            pass
        try:
            h = self.tableBindings.horizontalHeader()
            h.setSectionsClickable(True)
            h.sectionClicked.connect(self._on_bind_header_clicked)
        except Exception:
            pass
        self.tableBindings.verticalHeader().setVisible(False)

        # callback на перестановку строк для обновления порядка в данных
        def _rows_moved(src_rows: list[int], dest_row: int):
            q = self.current_queue()
            if not q or not src_rows:
                return
            n = len(q.bindings)
            src = sorted(set([r for r in src_rows if 0 <= r < n]))
            moving = [q.bindings[i] for i in src]
            remaining = [q.bindings[i] for i in range(n) if i not in src]
            # позиция вставки в массиве remaining
            # если вставка ниже исходного блока - сдвигаем индекс на количество удалённых строк до неё
            shift = sum(1 for i in src if i < dest_row)
            ins = max(0, min(len(remaining), dest_row - shift))
            new_order = remaining[:ins] + moving + remaining[ins:]
            q.bindings = new_order
            self.fill_bind_table(q)
            # выделим перемещённые строки
            for i in range(ins, ins + len(moving)):
                self.tableBindings.selectRow(i)
        self.tableBindings.on_rows_moved = _rows_moved

        # drag&drop
        self.tableBindings.setDragEnabled(True)
        self.tableBindings.setAcceptDrops(True)
        self.tableBindings.setDropIndicatorShown(True)
        self.tableBindings.setDragDropMode(QtWidgets.QAbstractItemView.InternalMove)
        self.tableBindings.setDragDropOverwriteMode(False)
        self.tableBindings.setDefaultDropAction(QtCore.Qt.MoveAction)
        self.tableBindings.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        self.tableBindings.viewport().installEventFilter(self)
        # context menu
        self.tableBindings.setContextMenuPolicy(QtCore.Qt.CustomContextMenu)
        self.tableBindings.customContextMenuRequested.connect(self._on_bind_context_menu)

        self.tableParams.setColumnCount(3)
        self.tableParams.setHorizontalHeaderLabels(["Шифр", "Тип", "Модель"])
        self.tableParams.horizontalHeader().setStretchLastSection(False)
        self.tableParams.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        self.tableParams.setSelectionMode(QtWidgets.QAbstractItemView.ExtendedSelection)
        self.tableParams.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        try:
            self.tableParams.setShowGrid(False)
        except Exception:
            pass
        header = self.tableParams.horizontalHeader()
        header.setMinimumSectionSize(40)
        header.setSectionResizeMode(0, QtWidgets.QHeaderView.Interactive)
        header.setSectionResizeMode(1, QtWidgets.QHeaderView.Interactive)
        header.setSectionResizeMode(2, QtWidgets.QHeaderView.Interactive)
        self.tableParams.verticalHeader().setVisible(False)

        # контекстное меню заголовка для фильтров (только ПКМ)
        header = self.tableParams.horizontalHeader()
        header.setSectionsClickable(False)
        header.setContextMenuPolicy(QtCore.Qt.CustomContextMenu)
        header.customContextMenuRequested.connect(self._on_param_header_menu)

        # Внешний вид
        self.apply_style()
        self.setup_icon_buttons()

        # События
        self.btnImportXml.clicked.connect(self.import_xml)
        self.btnLoadGlobal.clicked.connect(self.load_global_attrs)
        self.btnImportExcel.clicked.connect(self.import_excel)
        btnModelsWindow = getattr(self, 'btnModelsWindow', None)
        if btnModelsWindow is not None:
            btnModelsWindow.clicked.connect(self.open_models_window)
        self.editAttrSearch.textChanged.connect(lambda *_: self.refresh_attr_tree(True))
        self.treeAttributes.selectionModel().selectionChanged.connect(self.on_attr_pick)
        if self.btnBindDelete is not None:
            self.btnBindDelete.clicked.connect(self.bind_delete)
        if self.btnBindToggle is not None:
            self.btnBindToggle.clicked.connect(self.bind_toggle_enabled)
        if self.btnTransform is not None:
            self.btnTransform.clicked.connect(self.open_transform_dialog)
        if self.btnUp is not None:
            self.btnUp.clicked.connect(self.bind_move_up)
        if self.btnDown is not None:
            self.btnDown.clicked.connect(self.bind_move_down)
        self.btnSave.clicked.connect(self.export_xml)
        if self.btnSaveGlobal is not None:
            self.btnSaveGlobal.clicked.connect(self.export_global_attrs_xml)
        self.tableParams.itemDoubleClicked.connect(lambda *_: self.add_selected_params_as_bindings())
        self.tableBindings.itemDoubleClicked.connect(lambda *_: self.bind_delete())
        self.tableBindings.cellClicked.connect(self._on_bind_cell_clicked)
        self.editParamFilter.textChanged.connect(lambda *_: self.fill_param_table())

        # Старт
        self.refresh_attr_tree(True)

    def changeEvent(self, event: QtCore.QEvent) -> None:
        try:
            t = event.type()
            if t in (QtCore.QEvent.StyleChange, QtCore.QEvent.PaletteChange, QtCore.QEvent.ApplicationPaletteChange):
                if not self._pending_theme_apply:
                    self._pending_theme_apply = True
                    QtCore.QTimer.singleShot(0, self._apply_theme_extras_from_app)
        except Exception:
            pass
        super().changeEvent(event)

    def _reload_attr_status_icons(self) -> None:
        try:
            app = QtWidgets.QApplication.instance()
            ok_path = resolve_icon_path("ok", ICON_DIR, app=app) or _resolve_icon_path(_ICON_FALLBACK_FILES["ok"])
            circle_path = resolve_icon_path("circle", ICON_DIR, app=app) or _resolve_icon_path(_ICON_FALLBACK_FILES["circle"])
            self._icon_attr_assigned = QtGui.QIcon(ok_path) if ok_path else _icon_ok()
            self._icon_attr_unassigned = QtGui.QIcon(circle_path) if circle_path else _icon_circle()
        except Exception:
            self._icon_attr_assigned = _icon_ok()
            self._icon_attr_unassigned = _icon_circle()

    def _apply_theme_extras(self, dark: bool) -> None:
        self._reload_bind_checkbox_icons(bool(dark))
        self._reload_attr_status_icons()
        delegate = self.tableBindings.itemDelegate()
        if isinstance(delegate, BindingsCheckDelegate):
            delegate.set_icons(
                icon_checked=self._icon_bind_checked,
                icon_unchecked=self._icon_bind_unchecked,
                icon_indeterminate=self._icon_bind_indeterminate
            )
        try:
            self._update_top_bar_icons()
        except Exception:
            pass
        try:
            self.setup_icon_buttons()
        except Exception:
            pass
        try:
            self.refresh_attr_tree(True)
        except Exception:
            pass
        q = self.current_queue()
        if q:
            self.fill_bind_table(q)

    def _update_top_bar_icons(self) -> None:
        app = QtWidgets.QApplication.instance()
        dark = is_dark_theme(app)
        icon_size = QtCore.QSize(24, 24)
        
        xml_path = resolve_icon_path("xml", ICON_DIR, app=app, tint_in_dark=True)
        if xml_path and os.path.exists(xml_path):
            self.btnImportXml.setIcon(QtGui.QIcon(xml_path))
        self.btnImportXml.setIconSize(icon_size)
        
        excel_path = resolve_icon_path("excel1", ICON_DIR, app=app, tint_in_dark=True)
        if excel_path and os.path.exists(excel_path):
            self.btnImportExcel.setIcon(QtGui.QIcon(excel_path))
        self.btnImportExcel.setIconSize(icon_size)
        
        upload_path = resolve_icon_path("upload", ICON_DIR, app=app, tint_in_dark=True)
        if upload_path and os.path.exists(upload_path):
            self.btnLoadGlobal.setIcon(QtGui.QIcon(upload_path))
        self.btnLoadGlobal.setIconSize(icon_size)

    def _apply_theme_extras_from_app(self) -> None:
        self._pending_theme_apply = False
        try:
            app = QtWidgets.QApplication.instance()
            self._apply_theme_extras(is_dark_theme(app))
        except Exception:
            pass

    def _on_theme_toggled(self, dark: bool):
        app = QtWidgets.QApplication.instance()
        theme(app, dark, icon_dir=ICON_DIR)
        apply_dark_titlebar(self, dark)
        self._apply_theme_extras(bool(dark))

 # ----------------- Внешний вид -----------------
    def apply_style(self):
        app = QtWidgets.QApplication.instance()
        dark = is_dark_theme(app)
        theme(app, dark, icon_dir=ICON_DIR)
        sep_color = "rgba(255, 255, 255, 0.22)" if dark else "rgba(0, 0, 0, 0.14)"
        self.ui.setStyleSheet(f"""
            QGroupBox#groupAttributes,
            QGroupBox#groupBindings,
            QGroupBox#groupParams {{
                border: none;
                margin-top: 10px;
                padding-top: 6px;
            }}
            QGroupBox#groupAttributes::title,
            QGroupBox#groupBindings::title,
            QGroupBox#groupParams::title {{
                left: 8px;
                padding: 0 6px;
            }}
            QTreeView#treeAttributes,
            QTableWidget#tableBindings,
            QTableWidget#tableParams {{
                border: none;
            }}
            QTableWidget#tableBindings {{
                selection-background-color: transparent;
                selection-color: #000000;
            }}
            QTableWidget#tableBindings QTableCornerButton::section,
            QTableWidget#tableParams QTableCornerButton::section {{
                background: transparent;
                border: none;
            }}
            QTreeView#treeAttributes::item:hover,
            QTreeView#treeAttributes::item:selected,
            QTreeView#treeAttributes::item:selected:active,
            QTreeView#treeAttributes::item:selected:!active {{
                background: transparent;
                border: none;
                outline: none;
            }}
            QTableWidget#tableBindings::item:hover,
            QTableWidget#tableBindings::item:selected,
            QTableWidget#tableBindings::item:selected:active,
            QTableWidget#tableBindings::item:selected:!active,
            QTableWidget#tableParams::item:hover,
            QTableWidget#tableParams::item:selected,
            QTableWidget#tableParams::item:selected:active,
            QTableWidget#tableParams::item:selected:!active {{
                background: transparent;
                border: none;
                outline: none;
            }}
            QSplitter#splitterMain::handle {{
                background: {sep_color};
                width: 2px;
                margin: 16px 2px 10px 2px;
                border-radius: 1px;
            }}
            QSplitter#splitterMain::handle:hover {{
                background: {PALETTE.ACCENT_HOVER};
            }}
            QTableWidget#tableBindings QScrollBar:vertical,
            QTableWidget#tableParams QScrollBar:vertical {{
                background: transparent;
                width: 12px;
                margin: 16px 2px 16px 2px;
                border: none;
            }}
            QTableWidget#tableBindings QScrollBar::handle:vertical,
            QTableWidget#tableParams QScrollBar::handle:vertical {{
                background: rgba(247,146,30,0.12);
                min-height: 24px;
                border-radius: 6px;
                border: 1px solid {PALETTE.ACCENT_HOVER};
            }}
            QTableWidget#tableBindings QScrollBar::handle:vertical:hover,
            QTableWidget#tableParams QScrollBar::handle:vertical:hover {{
                background: rgba(247,146,30,0.15);
                border: 1px solid {PALETTE.ACCENT_HOVER};
            }}
            QTableWidget#tableBindings QScrollBar::handle:vertical:pressed,
            QTableWidget#tableParams QScrollBar::handle:vertical:pressed {{
                background: rgba(247,146,30,0.25);
                border: 1px solid {PALETTE.ACCENT_PRESSED};
            }}
            QTableWidget#tableBindings QScrollBar::add-line:vertical,
            QTableWidget#tableBindings QScrollBar::sub-line:vertical,
            QTableWidget#tableParams QScrollBar::add-line:vertical,
            QTableWidget#tableParams QScrollBar::sub-line:vertical {{
                background: transparent;
                height: 16px;
                subcontrol-origin: margin;
                border: none;
                border-radius: 0;
                image: none;
            }}
            QTableWidget#tableBindings QScrollBar::add-line:vertical:hover,
            QTableWidget#tableBindings QScrollBar::sub-line:vertical:hover,
            QTableWidget#tableParams QScrollBar::add-line:vertical:hover,
            QTableWidget#tableParams QScrollBar::sub-line:vertical:hover {{
                background: rgba(247,146,30,0.15);
            }}
            QTableWidget#tableBindings QScrollBar::add-line:vertical:pressed,
            QTableWidget#tableBindings QScrollBar::sub-line:vertical:pressed,
            QTableWidget#tableParams QScrollBar::add-line:vertical:pressed,
            QTableWidget#tableParams QScrollBar::sub-line:vertical:pressed {{
                background: rgba(247,146,30,0.25);
            }}
            QTableWidget#tableBindings QScrollBar::add-page:vertical,
            QTableWidget#tableBindings QScrollBar::sub-page:vertical,
            QTableWidget#tableParams QScrollBar::add-page:vertical,
            QTableWidget#tableParams QScrollBar::sub-page:vertical {{
                background: transparent;
            }}
            QTableWidget#tableBindings QScrollBar:horizontal,
            QTableWidget#tableParams QScrollBar:horizontal {{
                background: transparent;
                height: 12px;
                margin: 2px 16px 2px 16px;
                border: none;
            }}
            QTableWidget#tableBindings QScrollBar::handle:horizontal,
            QTableWidget#tableParams QScrollBar::handle:horizontal {{
                background: rgba(247,146,30,0.12);
                min-width: 24px;
                border-radius: 6px;
                border: 1px solid {PALETTE.ACCENT_HOVER};
            }}
            QTableWidget#tableBindings QScrollBar::handle:horizontal:hover,
            QTableWidget#tableParams QScrollBar::handle:horizontal:hover {{
                background: rgba(247,146,30,0.15);
                border: 1px solid {PALETTE.ACCENT_HOVER};
            }}
            QTableWidget#tableBindings QScrollBar::handle:horizontal:pressed,
            QTableWidget#tableParams QScrollBar::handle:horizontal:pressed {{
                background: rgba(247,146,30,0.25);
                border: 1px solid {PALETTE.ACCENT_PRESSED};
            }}
            QTableWidget#tableBindings QScrollBar::add-line:horizontal,
            QTableWidget#tableBindings QScrollBar::sub-line:horizontal,
            QTableWidget#tableParams QScrollBar::add-line:horizontal,
            QTableWidget#tableParams QScrollBar::sub-line:horizontal {{
                background: transparent;
                width: 16px;
                subcontrol-origin: margin;
                border: none;
                border-radius: 0;
                image: none;
            }}
            QTableWidget#tableBindings QScrollBar::add-line:horizontal:hover,
            QTableWidget#tableBindings QScrollBar::sub-line:horizontal:hover,
            QTableWidget#tableParams QScrollBar::add-line:horizontal:hover,
            QTableWidget#tableParams QScrollBar::sub-line:horizontal:hover {{
                background: rgba(247,146,30,0.15);
            }}
            QTableWidget#tableBindings QScrollBar::add-line:horizontal:pressed,
            QTableWidget#tableBindings QScrollBar::sub-line:horizontal:pressed,
            QTableWidget#tableParams QScrollBar::add-line:horizontal:pressed,
            QTableWidget#tableParams QScrollBar::sub-line:horizontal:pressed {{
                background: rgba(247,146,30,0.25);
            }}
            QTableWidget#tableBindings QScrollBar::add-page:horizontal,
            QTableWidget#tableBindings QScrollBar::sub-page:horizontal,
            QTableWidget#tableParams QScrollBar::add-page:horizontal,
            QTableWidget#tableParams QScrollBar::sub-page:horizontal {{
                background: transparent;
            }}
        """)
        
        # Установка делегата для скругленной подсветки строк
        # tableBindings: отдельный делегат, чтобы центрировать чекбоксы в первом столбце.
        self.tableBindings.setItemDelegate(BindingsCheckDelegate(
            self.tableBindings,
            icon_checked=self._icon_bind_checked,
            icon_unchecked=self._icon_bind_unchecked,
            icon_indeterminate=self._icon_bind_indeterminate
        ))
        install_viewport_row_highlighter(self.tableBindings)
        setup_hover_tracking(self.tableBindings)

        # tableParams: обычный row-hover делегат.
        self.tableParams.setItemDelegate(RowHoverDelegate(self.tableParams))
        install_viewport_row_highlighter(self.tableParams)
        setup_hover_tracking(self.tableParams)
        self._apply_theme_extras(dark)

    def _reload_bind_checkbox_icons(self, dark: bool):
        try:
            app = QtWidgets.QApplication.instance()
            off_path = resolve_icon_path("check", ICON_DIR, app=app) or _resolve_icon_path(_ICON_FALLBACK_FILES["check"])
            on_path = resolve_icon_path("select", ICON_DIR, app=app) or _resolve_icon_path(_ICON_FALLBACK_FILES["select"])
            mid_path = resolve_icon_path("poloska", ICON_DIR, app=app) or _resolve_icon_path(_ICON_FALLBACK_FILES["poloska"])
            self._icon_bind_unchecked = QtGui.QIcon(off_path) if off_path else None
            self._icon_bind_checked = QtGui.QIcon(on_path) if on_path else None
            self._icon_bind_indeterminate = QtGui.QIcon(mid_path) if mid_path else None
        except Exception:
            self._icon_bind_unchecked = None
            self._icon_bind_checked = None
            self._icon_bind_indeterminate = None

    def setup_icon_buttons(self):
        icon_key_by_filename = {
            "arrow-up.png": "arrow_up",
            "arrow-down.png": "arrow_down",
            "delete.png": "delete",
            "free-icon-setting-3288004.png": "gear",
        }
        def _norm(name: str) -> str:
            return icon_key_by_filename.get(str(name or ""), str(name or ""))

        def _apply(btn: Optional[QtWidgets.QPushButton], icon_name: str, tooltip: str):
            if btn is None:
                return
            try:
                btn.setText("")
            except Exception:
                pass
            try:
                btn.setToolTip(tooltip)
            except Exception:
                pass
            icon = None
            icon_key = _norm(icon_name)
            app = QtWidgets.QApplication.instance()
            tint = icon_key not in ("gear", "setting", "free-icon-setting-3288004.png")
            icon_path = resolve_icon_path(icon_key, ICON_DIR, app=app, tint_in_dark=tint) or _resolve_icon_path(icon_name)
            if icon_path:
                try:
                    icon = QtGui.QIcon(icon_path)
                except Exception:
                    icon = None
            else:
                try:
                    if icon_key in ("arrow_up", "arrow-up.png"):
                        icon = _icon_arrow("up")
                    elif icon_key in ("arrow_down", "arrow-down.png"):
                        icon = _icon_arrow("down")
                    elif icon_key in ("delete", "delete.png"):
                        icon = self.style().standardIcon(QtWidgets.QStyle.SP_TrashIcon)
                    else:
                        icon = self.style().standardIcon(QtWidgets.QStyle.SP_FileDialogDetailedView)
                except Exception:
                    icon = None
            if icon is not None:
                try:
                    btn.setIcon(icon)
                except Exception:
                    pass
            try:
                btn.setIconSize(QtCore.QSize(18, 18))
                btn.setMinimumSize(QtCore.QSize(34, 34))
                btn.setMaximumSize(QtCore.QSize(34, 34))
            except Exception:
                pass

        _apply(self.btnBindDelete, "delete", "Удалить выбранные")
        _apply(self.btnTransform, "gear", "Преобразования")
        _apply(self.btnUp, "arrow_up", "Вверх")
        _apply(self.btnDown, "arrow_down", "Вниз")

    # ---------- API helpers ----------
    def base(self) -> str:
        return _runtime_base_url()

    # ---------- Проекты/модели/параметры ----------
    def on_project_changed(self, idx: int):
        """При смене проекта автоматически подгружаем модели и сбрасываем параметры/фильтры."""
        if idx is None or idx < 0 or idx >= len(self.projects):
            return
        # загрузка контейнеров для выбранного проекта
        pid = int(self.projects[idx]["id"])
        try:
            self.containers = sorted(api_get_containers(self.base(), pid), key=lambda c: (c["title"] or "").lower())
        except Exception as e:
            show_error_dialog(self, "API", f"Не удалось получить модели:\n{e}")
            return
        # сбрасываем параметры и фильтры по моделям (т.к. модели другие)
        self.per_model_params.clear()
        self.model_filter_set = None
        self.fill_param_table()
        self.lblStatus.setText(f"Загружено моделей: {len(self.containers)}")
    
    def load_projects(self):
        try:
            self.projects = sorted(api_get_projects(self.base()), key=lambda p: (p["title"] or "").lower())
        except Exception as e:
            show_error_dialog(self, "API", f"Не удалось получить проекты:\n{e}")
            return False
        self.comboProject.clear()
        for p in self.projects:
            self.comboProject.addItem(p["title"])
        self.containers.clear()
        self.per_model_params.clear()
        self.tableParams.setRowCount(0)
        self.lblStatus.setText("")
        return True

    def ensure_containers_loaded(self) -> bool:
        i = self.comboProject.currentIndex()
        if i < 0:
            show_error_dialog(self, "API", "Выбери проект.")
            return False
        if self.containers:
            return True
        pid = int(self.projects[i]["id"])
        try:
            self.containers = sorted(api_get_containers(self.base(), pid), key=lambda c: (c["title"] or "").lower())
        except Exception as e:
            show_error_dialog(self, "API", f"Не удалось получить модели:\n{e}")
            return False
        return True

    def open_models_window(self):
        if not self.projects:
            if not self.load_projects():
                return
        dlg = QtWidgets.QDialog(self)
        dlg.setWindowTitle("Выбор моделей")
        dlg.resize(700, 500)
        v = QtWidgets.QVBoxLayout(dlg)

        proj_layout = QtWidgets.QHBoxLayout()
        proj_label = QtWidgets.QLabel("Проект:")
        combo_proj = QtWidgets.QComboBox()
        combo_proj.setSizeAdjustPolicy(QtWidgets.QComboBox.AdjustToContentsOnFirstShow)
        for p in self.projects:
            combo_proj.addItem(p["title"])
        if self.comboProject.currentIndex() >= 0:
            combo_proj.setCurrentIndex(self.comboProject.currentIndex())
        proj_layout.addWidget(proj_label)
        proj_layout.addWidget(combo_proj)
        proj_layout.addStretch(1)
        v.addLayout(proj_layout)

        lst = QtWidgets.QListWidget()
        lst.setSelectionMode(QtWidgets.QAbstractItemView.ExtendedSelection)
        v.addWidget(lst)

        def on_proj_change(idx: int):
            lst.clear()
            if idx is None or idx < 0 or idx >= len(self.projects):
                return
            pid = int(self.projects[idx]["id"])
            try:
                containers = sorted(api_get_containers(self.base(), pid), key=lambda c: (c["title"] or "").lower())
                for c in containers:
                    lst.addItem(c["title"])
                lst._containers = containers
            except Exception as e:
                show_error_dialog(dlg, "API", f"Не удалось получить модели:\n{e}")
        
        lst._containers = []
        combo_proj.currentIndexChanged.connect(on_proj_change)
        if combo_proj.currentIndex() >= 0:
            on_proj_change(combo_proj.currentIndex())

        h = QtWidgets.QHBoxLayout()
        v.addLayout(h)
        btnAll = QtWidgets.QPushButton("Выделить все")
        btnNone = QtWidgets.QPushButton("Снять выделение")
        btnLoad = QtWidgets.QPushButton("Загрузить параметры")
        h.addWidget(btnAll); h.addWidget(btnNone); h.addStretch(1); h.addWidget(btnLoad)

        btnAll.clicked.connect(lambda: lst.selectAll())
        btnNone.clicked.connect(lambda: lst.clearSelection())

        def do_load():
            sels = lst.selectedIndexes()
            if not sels:
                show_error_dialog(dlg, "API", "Выдели минимум одну модель.")
                return
            idx = combo_proj.currentIndex()
            if idx >= 0:
                self.comboProject.setCurrentIndex(idx)
            containers = getattr(lst, '_containers', [])
            if not containers:
                pid = int(self.projects[idx]["id"])
                try:
                    containers = sorted(api_get_containers(self.base(), pid), key=lambda c: (c["title"] or "").lower())
                except Exception as e:
                    show_error_dialog(dlg, "API", f"Не удалось получить модели:\n{e}")
                    return
            self.containers = containers
            self.per_model_params.clear()
            self.model_filter_set = None
            errs = 0
            for i in sels:
                row = i.row()
                cid = int(containers[row]["id"])
                try:
                    self.per_model_params[cid] = api_get_params_for_container(self.base(), cid)
                except Exception as e:
                    errs += 1
                    show_error_dialog(dlg, "API", f"Не получил параметры для «{containers[row]['title']}»:\n{e}")
            self.fill_param_table()
            self.lblStatus.setText(f"Загружено моделей: {len(sels)}; ошибок: {errs}")
            dlg.accept()

        btnLoad.clicked.connect(do_load)
        dlg.exec()

    def current_models(self) -> List[str]:
        title_by_cid = {c["id"]: c["title"] for c in self.containers}
        return sorted([title_by_cid.get(cid, f"ID {cid}") for cid in self.per_model_params.keys()], key=str.lower)

    def fill_param_table(self):
        title_by_cid = {c["id"]: c["title"] for c in self.containers}
        grouped: Dict[str, Dict[str, Any]] = {}

        code_q = (self.editParamFilter.text() or "").lower().strip()

        for cid, items in self.per_model_params.items():
            model_title = title_by_cid.get(cid, f"ID {cid}")
            for r in items:
                code = r["code"]
                # 1) по шифру
                if code_q and code_q not in code.lower():
                    continue
                # 2) по типу
                if self.type_filter is not None and bool(r["isNumeric"]) != bool(self.type_filter):
                    continue
                # 3) по моделям
                if self.model_filter_set is not None and model_title not in self.model_filter_set:
                    continue
                g = grouped.setdefault(code, {"isnum": r["isNumeric"], "models": set()})
                g["isnum"] = g["isnum"] or r["isNumeric"]
                g["models"].add(model_title)

        rows = []
        for code, data in grouped.items():
            models = sorted(list(data["models"]), key=str.lower)
            isnum = "Число" if data["isnum"] else "Текст"
            rows.append((code, isnum, models))

        rows.sort(key=lambda x: (x[0].lower(), x[2][0].lower() if x[2] else ""))

        self.tableParams.setRowCount(0)
        self.param_models_map.clear()
        for i, (code, isnum, models) in enumerate(rows):
            self.tableParams.insertRow(i)
            self.tableParams.setItem(i, 0, QtWidgets.QTableWidgetItem(code))
            self.tableParams.setItem(i, 1, QtWidgets.QTableWidgetItem(isnum))
            if len(models) > 1:
                display_text = f"{models[0]}..."
            else:
                display_text = models[0] if models else ""
            it2 = QtWidgets.QTableWidgetItem(display_text)
            it2.setToolTip("\n".join(models))
            self.tableParams.setItem(i, 2, it2)
            self.param_models_map[i] = models
        self.tableParams.resizeColumnToContents(0)
        self.tableParams.resizeColumnToContents(1)
        self.tableParams.resizeColumnToContents(2)

    # ---------- Фильтры заголовков ----------
    
    def on_param_header_clicked(self, section: int):
        if section == 1:
            self._show_type_filter_menu()
        elif section == 2:
            # по клику на "Модель" открываем окно выбора моделей
            self.open_model_filter_dialog()

    def _on_param_header_menu(self, pos: QtCore.QPoint):
        header: QtWidgets.QHeaderView = self.tableParams.horizontalHeader()
        section = int(header.logicalIndexAt(pos))
        if section not in (1, 2):
            return
        if section == 1:
            self._show_type_filter_menu(header.mapToGlobal(pos))
        else:
            self._show_model_filter_menu(header.mapToGlobal(pos))

    def _show_type_filter_menu(self, global_pos: Optional[QtCore.QPoint] = None):
        menu = QtWidgets.QMenu(self)
        act_all = QtGui.QAction("Оба", menu)
        act_text = QtGui.QAction("Текст", menu)
        act_num = QtGui.QAction("Число", menu)
        ag = QtGui.QActionGroup(menu)
        for a in (act_all, act_text, act_num):
            a.setCheckable(True)
            ag.addAction(a); menu.addAction(a)
        if self.type_filter is None:
            act_all.setChecked(True)
        elif self.type_filter is False:
            act_text.setChecked(True)
        else:
            act_num.setChecked(True)

        def apply_choice(a: QtGui.QAction):
            if a is act_all:
                self.type_filter = None
            elif a is act_text:
                self.type_filter = False
            else:
                self.type_filter = True
            self.fill_param_table()
        ag.triggered.connect(apply_choice)
        menu.exec(global_pos or QtGui.QCursor.pos())

    def _show_model_filter_menu(self, global_pos: QtCore.QPoint):
        models = self.current_models()
        if not models:
            show_error_dialog(self, "Фильтр по моделям", "Сначала загрузите параметры через «Выбор моделей».")
            return

        current = set(models) if self.model_filter_set is None else set(self.model_filter_set)

        popup = QtWidgets.QFrame(self, QtCore.Qt.Popup)
        popup.setFrameShape(QtWidgets.QFrame.Shape.StyledPanel)
        popup.setStyleSheet("QFrame { background: palette(window); border: 1px solid palette(mid); border-radius: 6px; }")
        layout = QtWidgets.QVBoxLayout(popup)
        layout.setContentsMargins(8, 8, 8, 8)
        layout.setSpacing(4)

        list_widget = QtWidgets.QListWidget(popup)
        list_widget.setSelectionMode(QtWidgets.QAbstractItemView.SelectionMode.NoSelection)
        list_widget.setVerticalScrollBarPolicy(QtCore.Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        list_widget.setHorizontalScrollBarPolicy(QtCore.Qt.ScrollBarPolicy.ScrollBarAlwaysOff)

        item_all = QtWidgets.QListWidgetItem("Все модели")
        item_all.setFlags(item_all.flags() | QtCore.Qt.ItemFlag.ItemIsUserCheckable)
        item_all.setCheckState(QtCore.Qt.CheckState.Checked if self.model_filter_set is None else QtCore.Qt.CheckState.Unchecked)
        list_widget.addItem(item_all)

        for model in models:
            item = QtWidgets.QListWidgetItem(model)
            item.setFlags(item.flags() | QtCore.Qt.ItemFlag.ItemIsUserCheckable)
            item.setCheckState(QtCore.Qt.CheckState.Checked if model in current else QtCore.Qt.CheckState.Unchecked)
            list_widget.addItem(item)

        list_widget.itemClicked.connect(lambda it: None)
        layout.addWidget(list_widget)

        def apply_filter():
            all_checked = item_all.checkState() == QtCore.Qt.CheckState.Checked
            if all_checked:
                self.model_filter_set = None
            else:
                selected = set()
                for i in range(1, list_widget.count()):
                    it = list_widget.item(i)
                    if it.checkState() == QtCore.Qt.CheckState.Checked:
                        selected.add(it.text())
                self.model_filter_set = None if len(selected) == len(models) else selected
            self.fill_param_table()

        def on_item_changed(it: QtWidgets.QListWidgetItem):
            if it is item_all:
                state = it.checkState()
                for i in range(1, list_widget.count()):
                    list_widget.item(i).setCheckState(state)
            else:
                all_items_checked = all(
                    list_widget.item(i).checkState() == QtCore.Qt.CheckState.Checked
                    for i in range(1, list_widget.count())
                )
                item_all.setCheckState(QtCore.Qt.CheckState.Checked if all_items_checked else QtCore.Qt.CheckState.Unchecked)
            apply_filter()

        list_widget.itemChanged.connect(on_item_changed)

        btn_close = QtWidgets.QPushButton("Закрыть", popup)
        btn_close.clicked.connect(popup.close)
        layout.addWidget(btn_close)

        popup.setLayout(layout)
        popup.move(global_pos)
        popup.show()
        popup.setFocus()

    def open_model_filter_dialog(self):
        models = self.current_models()
        if not models:
            show_error_dialog(self, "Фильтр по моделям", "Сначала загрузите параметры через «Выбор моделей».")
            return
        dlg = ModelFilterDialog(self, models, self.model_filter_set)
        if dlg.exec() == QtWidgets.QDialog.Accepted:
            chosen = dlg.chosen()
            self.model_filter_set = None if len(chosen) == len(models) else set(chosen)
            self.fill_param_table()

    # ---------- Атрибуты слева ----------
    def _attr_tree_state(self):
        expanded = set()
        for i in range(self.attrModel.rowCount()):
            item = self.attrModel.item(i, 0)
            if item and item.hasChildren() and self.treeAttributes.isExpanded(item.index()):
                expanded.add(item.text())
        sel = None
        indexes = self.treeAttributes.selectionModel().selectedIndexes()
        if indexes:
            it = self.attrModel.itemFromIndex(indexes[0])
            data = it.data()
            if data and data[0] == "leaf":
                sel = data[1]
        return expanded, sel

    def _restore_attr_tree_state(self, expanded, selected):
        for i in range(self.attrModel.rowCount()):
            item = self.attrModel.item(i, 0)
            if item and item.text() in expanded:
                self.treeAttributes.setExpanded(item.index(), True)
        if selected:
            def walk(parent=None):
                rng = range(self.attrModel.rowCount()) if parent is None else range(parent.rowCount())
                for i in rng:
                    it = self.attrModel.item(i,0) if parent is None else parent.child(i)
                    if not it: continue
                    data = it.data()
                    if data and data[0]=="leaf" and data[1]==selected:
                        self.treeAttributes.setCurrentIndex(it.index()); return True
                    if it.hasChildren() and walk(it): return True
                return False
            walk()

    def refresh_attr_tree(self, preserve_state: bool = True):
        expanded, selected = self._attr_tree_state() if preserve_state else (set(), None)
        q = (self.editAttrSearch.text() or "").lower().strip()
        self.attrModel.removeRows(0, self.attrModel.rowCount())
        buckets: Dict[str, List[str]] = {}
        for a in self.doc.list_attrs():
            full = a.attribute_full_name.strip()
            if q and q not in full.lower(): continue
            if "." in full:
                grp, leaf = full.split(".", 1)
            else:
                grp, leaf = "", full
            buckets.setdefault(grp, []).append(leaf)

        for grp in sorted(buckets.keys(), key=str.lower):
            if grp:
                parent = QtGui.QStandardItem(grp)
                parent.setEditable(False)
                parent_status = QtGui.QStandardItem("")
                parent_status.setEditable(False)
                try:
                    parent_status.setTextAlignment(QtCore.Qt.AlignCenter)
                except Exception:
                    pass
                self.attrModel.appendRow([parent, parent_status])
                for leaf in sorted(set(buckets[grp]), key=str.lower):
                    full = f"{grp}.{leaf}"
                    text = self.decorate_leaf(full)
                    child = QtGui.QStandardItem(text)
                    child.setData(("leaf", full))
                    child.setEditable(False)
                    status = QtGui.QStandardItem("")
                    status.setEditable(False)
                    status.setData(("leaf", full))
                    try:
                        status.setTextAlignment(QtCore.Qt.AlignCenter)
                    except Exception:
                        pass
                    try:
                        qobj = self.doc.queues.get(full)
                        has_bindings = bool(qobj and qobj.bindings)
                        if has_bindings and getattr(self, "_icon_attr_assigned", None):
                            status.setIcon(self._icon_attr_assigned)  # type: ignore[arg-type]
                        elif (not has_bindings) and getattr(self, "_icon_attr_unassigned", None):
                            status.setIcon(self._icon_attr_unassigned)  # type: ignore[arg-type]
                    except Exception:
                        pass
                    parent.appendRow([child, status])
            else:
                for leaf in sorted(set(buckets.get("", [])), key=str.lower):
                    full = leaf
                    text = self.decorate_leaf(full)
                    item = QtGui.QStandardItem(text)
                    item.setData(("leaf", full))
                    item.setEditable(False)
                    status = QtGui.QStandardItem("")
                    status.setEditable(False)
                    status.setData(("leaf", full))
                    try:
                        status.setTextAlignment(QtCore.Qt.AlignCenter)
                    except Exception:
                        pass
                    try:
                        qobj = self.doc.queues.get(full)
                        has_bindings = bool(qobj and qobj.bindings)
                        if has_bindings and getattr(self, "_icon_attr_assigned", None):
                            status.setIcon(self._icon_attr_assigned)  # type: ignore[arg-type]
                        elif (not has_bindings) and getattr(self, "_icon_attr_unassigned", None):
                            status.setIcon(self._icon_attr_unassigned)  # type: ignore[arg-type]
                    except Exception:
                        pass
                    self.attrModel.appendRow([item, status])
        self._restore_attr_tree_state(expanded, selected)

    def decorate_leaf(self, full: str) -> str:
        return full

    def on_attr_pick(self):
        indexes = self.treeAttributes.selectionModel().selectedIndexes()
        if not indexes:
            self.selected_attr = None; self.tableBindings.setRowCount(0); return
        item: QtGui.QStandardItem = self.attrModel.itemFromIndex(indexes[0])
        data = item.data()
        if not data or data[0] != "leaf":
            self.selected_attr = None; self.tableBindings.setRowCount(0); return
        name = data[1]
        if "." not in name:
            self.selected_attr = None; self.tableBindings.setRowCount(0); return
        self.selected_attr = name
        q = self.doc.ensure_attr(name)
        self.fill_bind_table(q)

    def attrs_toggle_default(self):
        changed = False
        for idx in self.treeAttributes.selectionModel().selectedIndexes():
            item = self.attrModel.itemFromIndex(idx); data = item.data()
            if not data or data[0] != "leaf": continue
            name = data[1]; q = self.doc.ensure_attr(name)
            q.default_is_enabled = not q.default_is_enabled; changed = True
        if changed:
            self.refresh_attr_tree(True)

    def attrs_mark_with_bindings(self):
        changed = False
        for name, q in self.doc.queues.items():
            if q.bindings and not q.default_is_enabled:
                q.default_is_enabled = True; changed = True
        if changed:
            self.refresh_attr_tree(True)
            self.lblStatus.setText("Отмечены атрибуты с привязками.")
        else:
            self.lblStatus.setText("Нет изменений.")

    # ---------- Таблица привязок ----------
    def fill_bind_table(self, q: BindingQueue):
        self._is_filling_bind_table = True
        try:
            self.tableBindings.setRowCount(0)
            dst_isnum = self.attr_types.get(q.attribute_full_name, None)
            dst_s = "Число" if dst_isnum else "Текст" if dst_isnum is not None else "?"
            self.bind_models_map.clear()
            for i, b in enumerate(q.bindings):
                src_s = "Число" if b.src_is_numeric else "Текст" if b.src_is_numeric is not None else "?"
                conv = f"{src_s} -> {dst_s}"
                self.tableBindings.insertRow(i)

                it0 = QtWidgets.QTableWidgetItem("")
                it0.setData(QtCore.Qt.UserRole, b.uid)
                it0.setData(QtCore.Qt.UserRole + 1, bool(b.is_enabled))
                it0.setTextAlignment(QtCore.Qt.AlignCenter)
                it0.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                if b.is_enabled:
                    if self._icon_bind_checked is not None:
                        it0.setIcon(self._icon_bind_checked)
                    else:
                        it0.setIcon(_icon_checkbox(True, size=BIND_CHECKBOX_ICON_SIZE))
                else:
                    if self._icon_bind_unchecked is not None:
                        it0.setIcon(self._icon_bind_unchecked)
                    else:
                        it0.setIcon(_icon_checkbox(False, size=BIND_CHECKBOX_ICON_SIZE))
                self.tableBindings.setItem(i, 0, it0)

                it1 = QtWidgets.QTableWidgetItem(b.parameter_code)
                it1.setToolTip(b.parameter_code)
                it1.setTextAlignment(QtCore.Qt.AlignmentFlag.AlignLeft | QtCore.Qt.AlignmentFlag.AlignVCenter)
                self.tableBindings.setItem(i, 1, it1)
                it2 = QtWidgets.QTableWidgetItem(conv)
                it2.setTextAlignment(QtCore.Qt.AlignmentFlag.AlignLeft | QtCore.Qt.AlignmentFlag.AlignVCenter)
                self.tableBindings.setItem(i, 2, it2)
                models_list = [m.strip() for m in (b.src_model_title or "").split(",") if m.strip()]
                if len(models_list) > 1:
                    display_text = f"{models_list[0]}..."
                else:
                    display_text = models_list[0] if models_list else ""
                it3 = QtWidgets.QTableWidgetItem(display_text)
                it3.setTextAlignment(QtCore.Qt.AlignmentFlag.AlignLeft | QtCore.Qt.AlignmentFlag.AlignVCenter)
                if models_list:
                    it3.setToolTip("\n".join(models_list))
                self.tableBindings.setItem(i, 3, it3)
                self.bind_models_map[i] = models_list
            self._update_bind_header_checkbox(q)
        finally:
            self._is_filling_bind_table = False

    def _update_bind_header_checkbox(self, q: BindingQueue):
        try:
            header_item = self.tableBindings.horizontalHeaderItem(0)
            if header_item is None:
                header_item = QtWidgets.QTableWidgetItem("")
                self.tableBindings.setHorizontalHeaderItem(0, header_item)
            header_item.setText("")
            if not q.bindings:
                icon = self._icon_bind_unchecked or _icon_checkbox(False, size=BIND_CHECKBOX_ICON_SIZE)
            else:
                enabled = [bool(b.is_enabled) for b in q.bindings]
                if all(enabled):
                    icon = self._icon_bind_checked or _icon_checkbox(True, size=BIND_CHECKBOX_ICON_SIZE)
                elif any(enabled):
                    icon = self._icon_bind_indeterminate or (self._icon_bind_unchecked or _icon_checkbox(False, size=BIND_CHECKBOX_ICON_SIZE))
                else:
                    icon = self._icon_bind_unchecked or _icon_checkbox(False, size=BIND_CHECKBOX_ICON_SIZE)
            header_item.setIcon(icon)
            if getattr(self, "_bind_header", None) is not None:
                self._bind_header.set_checkbox_icon(icon)
        except Exception:
            pass

    def _on_bind_header_clicked(self, section: int):
        if section != 0:
            return
        if self._is_filling_bind_table:
            return
        q = self.current_queue()
        if not q or not q.bindings:
            return
        enable = not all(bool(b.is_enabled) for b in q.bindings)
        for b in q.bindings:
            b.is_enabled = enable
        self.fill_bind_table(q)
        self.refresh_attr_tree(True)

    def _on_bind_cell_clicked(self, row: int, col: int):
        if self._is_filling_bind_table:
            return
        if col != 0:
            return
        q = self.current_queue()
        if not q:
            return
        if not (0 <= row < len(q.bindings)):
            return
        q.bindings[row].is_enabled = not q.bindings[row].is_enabled
        self.fill_bind_table(q)
        self._reselect_rows([row])

    def current_queue(self) -> Optional[BindingQueue]:
        if not self.selected_attr: return None
        return self.doc.ensure_attr(self.selected_attr)

    def add_selected_params_as_bindings(self):
        q = self.current_queue()
        if not q:
            show_error_dialog(self, "Атрибут", "Выбери атрибут-лист слева."); return
        sels = self.tableParams.selectionModel().selectedRows()
        if not sels:
            show_error_dialog(self, "Параметры", "Отметь строки справа."); return
        for m in sels:
            r = m.row()
            code = self.tableParams.item(r, 0).text()
            isnum = self.tableParams.item(r, 1).text() == "Число"
            models = self.param_models_map.get(r, [])
            enabled = self.chkAutoEnable.isChecked() if self.chkAutoEnable is not None else True
            b = Binding(
                parameter_code=code,
                src_is_numeric=isnum,
                is_enabled=enabled,
                src_model_title=", ".join(models)
            )
            q.bindings.append(b)
        self.fill_bind_table(q)
        self.refresh_attr_tree(True)

    def bind_delete(self):
        q = self.current_queue()
        if not q: return
        sels = sorted([m.row() for m in self.tableBindings.selectionModel().selectedRows()], reverse=True)
        for r in sels:
            if 0 <= r < len(q.bindings): q.bindings.pop(r)
        self.fill_bind_table(q); self.refresh_attr_tree(True)

    def bind_toggle_enabled(self):
        q = self.current_queue()
        if not q: return
        sels = [m.row() for m in self.tableBindings.selectionModel().selectedRows()]
        for r in sels:
            if 0 <= r < len(q.bindings):
                q.bindings[r].is_enabled = not q.bindings[r].is_enabled
        self.fill_bind_table(q)

    def bind_move_up(self):
        q = self.current_queue()
        if not q: return
        sels = sorted([m.row() for m in self.tableBindings.selectionModel().selectedRows()])
        if not sels or sels[0] == 0: return
        for r in sels: q.bindings[r-1], q.bindings[r] = q.bindings[r], q.bindings[r-1]
        self.fill_bind_table(q); self._reselect_rows([r-1 for r in sels])

    def bind_move_down(self):
        q = self.current_queue()
        if not q: return
        sels = sorted([m.row() for m in self.tableBindings.selectionModel().selectedRows()], reverse=True)
        if not sels or sels[0] >= len(q.bindings)-1: return
        for r in sels: q.bindings[r], q.bindings[r+1] = q.bindings[r+1], q.bindings[r]
        self.fill_bind_table(q); self._reselect_rows([r+1 for r in reversed(sels)])

    def _reselect_rows(self, rows: List[int]):
        sel = self.tableBindings.selectionModel(); sel.clearSelection()
        for r in rows:
            if 0 <= r < self.tableBindings.rowCount():
                self.tableBindings.selectRow(r)
        if rows:
            self.tableBindings.scrollTo(self.tableBindings.model().index(rows[0], 0))

    def _on_bind_context_menu(self, pos: QtCore.QPoint):
        idx = self.tableBindings.indexAt(pos)
        if not idx.isValid(): return
        self.tableBindings.selectRow(idx.row())
        q = self.current_queue()
        if not q: return
        r = idx.row()
        if not (0 <= r < len(q.bindings)): return
        b = q.bindings[r]
        dlg = TransformDialog(self, b)
        if dlg.exec() == QtWidgets.QDialog.Accepted:
            self.fill_bind_table(q)

    # drag&drop sync

    def open_transform_dialog(self):
        q = self.current_queue()
        if not q:
            return
        sels = [m.row() for m in self.tableBindings.selectionModel().selectedRows()]
        if not sels:
            show_error_dialog(self, "Преобразования", "Выбери строку привязки.")
            return
        r = sels[0]
        if not (0 <= r < len(q.bindings)):
            return
        b = q.bindings[r]
        dlg = TransformDialog(self, b)
        if dlg.exec() == QtWidgets.QDialog.Accepted:
            self.fill_bind_table(q)

    # ---------- Импорт/Экспорт/Глобальные атрибуты ----------
    def load_global_attrs(self):
        try:
            gc = api_get_global_component(self.base(), comp_type=1)
        except Exception as e:
            show_error_dialog(self, "API", f"Не удалось получить общие атрибуты:\n{e}")
            return
        pairs = flatten_global_attributes_with_types(gc)
        self.doc = AdapterDoc()
        self.attr_types.clear()
        self.selected_attr = None
        for full, isnum in pairs:
            if not full: continue
            self.doc.ensure_attr(full)
            self.attr_types[full] = isnum
        self.tableBindings.setRowCount(0)
        self.labelAttrSource.setText("Параметры загружены из общих атрибутов")
        self.refresh_attr_tree(True)
        show_error_dialog(self, "Общие атрибуты", f"Получено: {len(pairs)}")

    def import_xml(self):
        path, _ = QtWidgets.QFileDialog.getOpenFileName(self, "Открыть Adapter.xml", "", "Adapter XML (*.xml);;All files (*.*)")
        if not path: return
        try:
            self.doc = AdapterDoc.from_xml(path)
            self.attr_types.clear()
            self.selected_attr = None
            self.tableBindings.setRowCount(0)
            self.labelAttrSource.setText("Параметры загружены из импортированного файла")
            self.refresh_attr_tree(True)
            show_info_dialog(self, "Импорт", f"Загружено атрибутов: {len(self.doc.queues)}")
        except Exception as e:
            show_error_dialog(self, "Импорт", f"Не удалось импортировать:\n{e}")

    def export_xml(self):
        if not self.doc.queues:
            show_error_dialog(self, "Экспорт", "Нет данных для сохранения."); return
        path, _ = QtWidgets.QFileDialog.getSaveFileName(self, "Сохранить Adapter.xml", "Adapter.xml", "Adapter XML (*.xml);;All files (*.*)")
        if not path: return
        try:
            tree = self.doc.to_xml(); tree.write(path, encoding="utf-8", xml_declaration=True)
            show_info_dialog(self, "Экспорт", f"Сохранено:\n{path}")
        except Exception as e:
            show_error_dialog(self, "Экспорт", f"Ошибка сохранения:\n{e}")

    def export_global_attrs_xml(self):
        names = sorted((self.doc.queues or {}).keys(), key=lambda s: (s or "").lower())
        if not names:
            show_error_dialog(self, "Общие атрибуты", "Нет общих атрибутов для сохранения.")
            return

        path, _ = QtWidgets.QFileDialog.getSaveFileName(
            self,
            "Сохранить общие атрибуты.xml",
            "Общие атрибуты.xml",
            "XML (*.xml);;All files (*.*)",
        )
        if not path:
            return

        try:
            root = ET.Element(
                "EC.Entities.AttributeTree",
                {
                    "xmlns:xsd": "http://www.w3.org/2001/XMLSchema",
                    "xmlns:xsi": "http://www.w3.org/2001/XMLSchema-instance",
                },
            )
            sec = ET.SubElement(root, "Section")
            
            groups_map: dict[str, list[tuple[str, str]]] = {}
            for nm in names:
                group = (self.attr_groups or {}).get(nm, "")
                if not group and "." in nm:
                    group = nm.split(".", 1)[0]
                if not group:
                    group = "General"
                short_name = nm.split(".", 1)[-1] if "." in nm else nm
                if group not in groups_map:
                    groups_map[group] = []
                groups_map[group].append((nm, short_name))
            
            attr_id = 1
            for group_name in sorted(groups_map.keys(), key=str.lower):
                sections = ET.SubElement(sec, "Sections", {"Name": group_name, "IsComitted": "false"})
                ET.SubElement(sections, "Section")
                attrs = ET.SubElement(sections, "Attributes")
                for full_name, short_name in groups_map[group_name]:
                    is_num = bool(self.attr_types.get(full_name)) if self.attr_types else False
                    ET.SubElement(
                        attrs,
                        "Attribute",
                        {
                            "Name": short_name,
                            "IsComitted": "false",
                            "Id": str(attr_id),
                            "Title": "",
                            "Description": "",
                            "Uom": "",
                            "IsNumeric": _b(is_num),
                            "ReportColumnType": "Text",
                        },
                    )
                    attr_id += 1
            ET.SubElement(sec, "Attributes")

            tree = ET.ElementTree(root)
            tree.write(path, encoding="utf-8", xml_declaration=True)
            show_error_dialog(self, "Общие атрибуты", f"Сохранено:\n{path}")
        except Exception as e:
            show_error_dialog(self, "Общие атрибуты", f"Ошибка сохранения:\n{e}")

def main():
    app = QtWidgets.QApplication(sys.argv)
    try:
        if os.path.exists(TITLEBAR_ICON_PATH):
            app.setWindowIcon(QtGui.QIcon(TITLEBAR_ICON_PATH))
    except Exception:
        pass
    try:
        theme(app, load_saved_theme(False), icon_dir=ICON_DIR, persist=False)
        enable_theme_sync(app, ICON_DIR)
    except Exception:
        pass
    win = MainWin()
    win.show()
    apply_dark_titlebar(win, is_dark_theme(app))
    sys.exit(app.exec())
# === HOTFIX: импорт параметров из Excel (вставить перед if __name__ == "__main__":) ===

class _ExcelSheetSelectDialog(QtWidgets.QDialog):
    def __init__(self, sheet_names: list[str], parent=None):
        super().__init__(parent)
        self.setWindowTitle("Выбор листов Excel")
        self.setMinimumWidth(320)
        self.setMinimumHeight(300)
        self._sheet_names = sheet_names
        self._check_states: dict[str, bool] = {name: False for name in sheet_names}
        
        layout = QtWidgets.QVBoxLayout(self)
        
        label = QtWidgets.QLabel("Выберите листы для импорта:")
        layout.addWidget(label)
        
        self.list_widget = QtWidgets.QListWidget()
        self.list_widget.setSelectionMode(QtWidgets.QAbstractItemView.NoSelection)
        for name in sheet_names:
            item = QtWidgets.QListWidgetItem(name)
            item.setFlags(item.flags() | QtCore.Qt.ItemIsUserCheckable)
            item.setCheckState(QtCore.Qt.Unchecked)
            self.list_widget.addItem(item)
        self.list_widget.itemChanged.connect(self._on_item_changed)
        layout.addWidget(self.list_widget)
        
        btn_layout = QtWidgets.QHBoxLayout()
        
        self.btn_all = QtWidgets.QPushButton("Выбрать все")
        self.btn_all.clicked.connect(self._select_all)
        btn_layout.addWidget(self.btn_all)
        
        self.btn_none = QtWidgets.QPushButton("Снять все")
        self.btn_none.clicked.connect(self._select_none)
        btn_layout.addWidget(self.btn_none)
        
        layout.addLayout(btn_layout)
        
        btn_box = QtWidgets.QDialogButtonBox(
            QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel
        )
        wire_dialog_button_box(btn_box, self.accept, self.reject)
        layout.addWidget(btn_box)
    
    def _on_item_changed(self, item: QtWidgets.QListWidgetItem):
        self._check_states[item.text()] = item.checkState() == QtCore.Qt.Checked
    
    def _select_all(self):
        for i in range(self.list_widget.count()):
            self.list_widget.item(i).setCheckState(QtCore.Qt.Checked)
    
    def _select_none(self):
        for i in range(self.list_widget.count()):
            self.list_widget.item(i).setCheckState(QtCore.Qt.Unchecked)
    
    def get_selected_sheets(self) -> list[str]:
        return [name for name, checked in self._check_states.items() if checked]


def _parse_excel_sheet(ws, default_group_from_cell: str = "") -> tuple[dict, dict, dict, list[str]]:
    """
    Парсит один лист Excel и возвращает данные для объединения.
    
    Returns:
        tuple: (queues_dict, attr_types, attr_groups, errors)
        - queues_dict: dict[str, BindingQueue] - словарь атрибутов
        - attr_types: dict[str, bool] - типы атрибутов (число/строка)
        - attr_groups: dict[str, str] - группы атрибутов
        - errors: list[str] - список ошибок при парсинге
    """
    from PySide6 import QtWidgets
    
    def _val(r, c):
        v = ws.cell(row=r, column=c).value
        return (str(v).strip() if v is not None else "")
    
    def _split_list(s: str):
        if not s:
            return []
        import csv, io
        reader = csv.reader(io.StringIO(s), delimiter=',', quotechar='"', doublequote=True, escapechar=None, skipinitialspace=True)
        try:
            row = next(reader)
            return [t for t in row if t is not None and t != '']
        except Exception:
            s = s.replace(';', ',')
            return [t.strip() for t in s.split(',') if t.strip()]
    
    def _is_valid_param_name(name: str) -> bool:
        name = (name or "").strip()
        if not name:
            return False
        return all(ch == "_" or ch.isalnum() for ch in name)
    
    header_row = None
    cols = {"name": None, "type": None, "list": None, "rep_from": None, "rep_to": None, "group": None}
    WANT = {
        "name": ["наименование параметра"],
        "type": ["тип параметра"],
        "list": ["список параметров"],
        "rep_from": ["замена с", "замена c"],
        "rep_to": ["замена на"],
        "group": ["группа параметров"],
    }
    
    for r in range(1, min(ws.max_row, 30) + 1):
        row_vals = [(_val(r, c)).lower() for c in range(1, ws.max_column + 1)]
        found = {}
        for key, variants in WANT.items():
            for c, txt in enumerate(row_vals, start=1):
                if any(v in txt for v in variants):
                    found[key] = c
                    break
        if {"name", "type", "list"}.issubset(found.keys()):
            header_row = r
            for k, v in found.items():
                cols[k] = v
            break
    
    if not header_row:
        return {}, {}, {}, ["Не найдена строка заголовков (ожидаются: Наименование параметра / Тип параметра / Список параметров)"]
    
    default_group = default_group_from_cell
    if not default_group:
        for rr in range(1, min(ws.max_row, 8) + 1):
            for cc in range(1, min(ws.max_column, 8) + 1):
                raw = ws.cell(row=rr, column=cc).value
                if not raw:
                    continue
                s = str(raw).strip()
                low = s.lower()
                if "групп" in low:
                    val = None
                    if ":" in s:
                        after = s.split(":", 1)[1].strip()
                        val = after if after else None
                    if not val:
                        right = ws.cell(row=rr, column=cc+1).value
                        if right:
                            val = str(right).strip()
                    if val:
                        default_group = val
                        break
            if default_group:
                break
    
    queues_dict: dict = {}
    attr_types: dict = {}
    attr_groups: dict = {}
    errors: list[str] = []
    
    r = header_row + 1
    while r <= ws.max_row:
        name = _val(r, cols["name"]) if cols["name"] else ""
        ptype = (_val(r, cols["type"]) if cols["type"] else "").lower()
        plist = _val(r, cols["list"]) if cols["list"] else ""
        rep_from = _val(r, cols["rep_from"]) if cols["rep_from"] else ""
        rep_to = _val(r, cols["rep_to"]) if cols["rep_to"] else ""
        grp = _val(r, cols["group"]) if cols["group"] else ""
        
        if name and not _is_valid_param_name(name):
            errors.append(f"Строка {r}: недопустимое имя параметра \"{name}\"")
            r += 1
            continue
        
        if not (name or plist or ptype or grp):
            r += 1
            continue
        
        isnum = True if "чис" in ptype else False
        full_name = f"{(grp or default_group)}.{name}" if (grp or default_group) else name
        
        if full_name not in queues_dict:
            queues_dict[full_name] = BindingQueue(attribute_full_name=full_name)
            queues_dict[full_name].bindings = []
        
        attr_types[full_name] = isnum
        if (grp or default_group):
            attr_groups[full_name] = (grp or default_group)
        
        lefts = _split_list(rep_from)
        rights = _split_list(rep_to)
        pairs = [(l, rights[i]) for i, l in enumerate(lefts) if i < len(rights)]
        
        params = _split_list(plist)
        for code in params:
            b = Binding(parameter_code=code, src_is_numeric=None, is_enabled=True, src_model_title="Excel")
            if pairs:
                t = TransformSettings()
                t.replaces = pairs
                b.transform = t
            queues_dict[full_name].bindings.append(b)
        r += 1
    
    return queues_dict, attr_types, attr_groups, errors


def _deamon_import_excel(self):
    from PySide6 import QtWidgets
    try:
        import openpyxl
    except Exception:
        show_error_dialog(self, "Импорт Excel", "Не установлен openpyxl.\nУстанови: pip install openpyxl")
        return

    path, _ = QtWidgets.QFileDialog.getOpenFileName(self, "Открыть Excel", "", "Excel (*.xlsx *.xlsm)")
    if not path:
        return

    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        sheet_names = wb.sheetnames
        if not sheet_names:
            show_error_dialog(self, "Импорт Excel", "В книге нет листов.")
            return
        
        if len(sheet_names) == 1:
            selected_sheets = sheet_names
        else:
            dlg = _ExcelSheetSelectDialog(sheet_names, self)
            if dlg.exec() != QtWidgets.QDialog.Accepted:
                return
            selected_sheets = dlg.get_selected_sheets()
            if not selected_sheets:
                show_error_dialog(self, "Импорт Excel", "Не выбрано ни одного листа.")
                return
        
        if not hasattr(self, "attr_types"):
            self.attr_types = {}
        if not hasattr(self, "attr_groups"):
            self.attr_groups = {}
        
        new_doc = AdapterDoc()
        self.attr_types.clear()
        self.attr_groups.clear()
        self.selected_attr = None
        
        loaded_sheets: list[str] = []
        skipped_sheets: list[tuple[str, str]] = []
        total_bindings = 0
        
        for sheet_name in selected_sheets:
            ws = wb[sheet_name]
            queues_dict, attr_types, attr_groups, errors = _parse_excel_sheet(ws)
            
            if errors:
                skipped_sheets.append((sheet_name, errors[0]))
                continue
            
            if not queues_dict:
                skipped_sheets.append((sheet_name, "Нет данных для импорта"))
                continue
            
            for full_name, queue in queues_dict.items():
                existing_q = new_doc.ensure_attr(full_name)
                existing_q.bindings.extend(queue.bindings)
                total_bindings += len(queue.bindings)
            
            for name, isnum in attr_types.items():
                self.attr_types[name] = isnum
            for name, grp in attr_groups.items():
                self.attr_groups[name] = grp
            
            loaded_sheets.append(sheet_name)
        
        if not loaded_sheets:
            msg = "Ни один лист не был загружен.\n\n"
            for name, reason in skipped_sheets:
                msg += f"• {name}: {reason}\n"
            show_error_dialog(self, "Импорт Excel", msg)
            return
        
        self.doc = new_doc
        self.tableBindings.setRowCount(0)
        self.labelAttrSource.setText(f"Параметры загружены из Excel ({len(loaded_sheets)} листов)")
        
        self.refresh_attr_tree(False)
        
        selected_index = None
        for i in range(self.attrModel.rowCount()):
            root_item = self.attrModel.item(i, 0)
            if not root_item:
                continue
            data = root_item.data()
            if data and data[0] == "leaf":
                selected_index = root_item.index()
                break
            if root_item.hasChildren():
                child = root_item.child(0)
                if child:
                    cdata = child.data()
                    if cdata and cdata[0] == "leaf":
                        selected_index = child.index()
                        break
        if selected_index:
            self.treeAttributes.setCurrentIndex(selected_index)
            self.on_attr_pick()
            try:
                name_item = self.attrModel.itemFromIndex(selected_index)
                data = name_item.data() if name_item else None
                if data and data[0] == 'leaf':
                    q = self.doc.ensure_attr(data[1])
                    self.fill_bind_table(q)
            except Exception:
                pass
            if self.tableBindings.rowCount() > 0:
                self.tableBindings.selectRow(0)
                self.tableBindings.setFocus()
        
        result_msg = f"Выбрано листов: {len(selected_sheets)}\n"
        result_msg += f"Загружено: {len(loaded_sheets)}\n"
        result_msg += f"Атрибутов: {len(self.doc.queues)}\n"
        result_msg += f"Привязок: {total_bindings}"
        if skipped_sheets:
            result_msg += f"\n\nПропущено листов:\n"
            for name, reason in skipped_sheets:
                result_msg += f"• {name}: {reason}\n"
        
        show_error_dialog(self, "Импорт Excel", result_msg)
    except Exception as e:
        show_error_dialog(self, "Импорт Excel", f"Не удалось импортировать Excel:\n{e}")
# Подмешиваем метод в класс, если его не было
MainWin.import_excel = _deamon_import_excel
# === /HOTFIX ===

if __name__ == "__main__":
    main()
