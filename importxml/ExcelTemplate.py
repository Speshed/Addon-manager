# -*- coding: utf-8 -*-
"""
Модуль для создания шаблона Excel для импорта параметров.

Функции модуля:
- create_excel_template(): Создаёт книгу Excel с шаблоном данных

Структура шаблона:
- Заголовки: ID | Параметр | Значение | Тип данных
- Примеры данных для демонстрации формата
"""

import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo


# ============================================================================
# КОНСТАНТЫ
# ============================================================================

# Заголовки столбцов шаблона Excel
EXCEL_HEADERS = ["ID", "Параметр", "Значение", "Тип данных"]

# Ширины столбцов (в символах Excel)
COL_WIDTHS = {
    "A": 12,  # ID
    "B": 25,  # Параметр
    "C": 25,  # Значение
    "D": 15   # Тип данных
}

# Высота строки заголовков
HEADER_ROW_HEIGHT = 30

# Стиль таблицы в Excel
TABLE_STYLE_NAME = "TableStyleMedium2"

# Примеры данных для шаблона
EXAMPLE_DATA = [
    ["322307", "\\Комментарии", "Первая итерация", "Текст"],
    ["338530", "\\Комментарии", "Первая итерация", "Текст"],
    ["321352", "\\Комментарии", "Первая итерация", "Текст"],
    ["397152", "\\Комментарии", "Первая итерация", "Текст"],
    ["319613", "\\Комментарии", "Первая итерация", "Текст"],
    ["323274", "\\Новый параметр", "Айбим", "Текст"],
]


# ============================================================================
# ФУНКЦИИ
# ============================================================================

def create_excel_template():
    """
    Создаёт книгу Excel с шаблоном для импорта параметров.

    Шаблон содержит:
    - Строку заголовков с названиями колонок
    - Примеры данных
    - Форматированную таблицу с автофильтрами

    Returns:
        openpyxl.Workbook: Объект книги Excel, готовый к сохранению

    Побочные эффекты:
        Нет
    """
    # 1. Создаём новую книгу Excel
    wb = openpyxl.Workbook()
    ws = wb.active

    # 2. Заполняем строку заголовков
    # Цикл: заполняем ячейки первой строки заголовками из константы EXCEL_HEADERS
    for col_idx, header in enumerate(EXCEL_HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        # Устанавливаем шрифт заголовка (жирный, размер 11)
        cell.font = Font(bold=True, size=11)
        # Устанавливаем выравнивание (по левому краю, по центру вертикали)
        cell.alignment = Alignment(horizontal="left", vertical="center")

    # 3. Устанавливаем ширину колонок из словаря COL_WIDTHS
    for col_letter, width in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    # 4. Устанавливаем высоту строки заголовков
    ws.row_dimensions[1].height = HEADER_ROW_HEIGHT

    # 5. Заполняем примеры данных
    # Цикл по строкам: заполняем каждую строку примером данных
    for row_idx, data in enumerate(EXAMPLE_DATA, 2):  # начинаем с 2-й строки (после заголовка)
        # Цикл по колонкам: заполняем ячейки в текущей строке
        for col_idx, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            # Устанавливаем шрифт для данных (обычный, размер 10)
            cell.font = Font(size=10)
            # Устанавливаем выравнивание (по левому краю, по центру вертикали)
            cell.alignment = Alignment(horizontal="left", vertical="center")

    # 6. Создаём объект таблицы Excel для автофильтров и стилей
    # Определяем диапазон таблицы (от A1 до последнего примера данных)
    table_ref = f"A1:D{1 + len(EXAMPLE_DATA)}"

    # Создаём таблицу Excel
    table = Table(displayName="ImportTable", ref=table_ref)

    # Настраиваем стиль таблицы
    table_style = TableStyleInfo(
        name=TABLE_STYLE_NAME,
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,    # Чередование цветов строк
        showColumnStripes=False,
    )
    table.tableStyleInfo = table_style

    # 7. Добавляем таблицу на лист
    ws.add_table(table)

    return wb
