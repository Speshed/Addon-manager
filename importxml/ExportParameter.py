# -*- coding: utf-8 -*-
"""
Модуль диалогового окна для экспорта параметров в Excel.

Основные классы:
- ExportParamsDialog: Диалог для выбора проекта, модели и параметров для экспорта

Используется для:
- Выбора проекта из списка
- Выбора одной или нескольких моделей
- Выбора одного или нескольких параметров
- Экспорта значений выбранных параметров для всех элементов выбранных моделей в Excel
"""

import os
from typing import Any, Dict, List, Optional

# PySide6 для GUI (основная библиотека)
try:
    from PySide6 import QtWidgets, QtGui, QtCore
except Exception:
    # Fallback на PyQt5 если PySide6 недоступен
    from PyQt5 import QtWidgets, QtGui, QtCore

# Библиотека для HTTP запросов
try:
    import requests
except Exception:
    requests = None

# Путь к директории с иконками
try:
    from Larix_set import ICON_DIR
except Exception:
    ICON_DIR = "icon"


# ============================================================================
# КОНСТАНТЫ
# ============================================================================

# Таймауты для HTTP запросов (в секундах)
REQUEST_TIMEOUT_SHORT = 30   # Для запросов списка проектов, контейнеров, параметров
REQUEST_TIMEOUT_LONG = 60    # Для запросов списка элементов

# HTTP коды успеха
HTTP_SUCCESS = 200

# Размеры окна диалога
DIALOG_WIDTH = 1000
DIALOG_HEIGHT = 700

# Типы данных параметров
PARAM_TYPE_NUMERIC = "число"
PARAM_TYPE_TEXT = "текст"


# ============================================================================
# КЛАССЫ
# ============================================================================

class ExportParamsDialog(QtWidgets.QDialog):
    """
    Диалоговое окно для экспорта параметров в Excel.

    Функционал:
    1. Выбор проекта из списка
    2. Выбор одной или нескольких моделей
    3. Выбор одного или нескольких параметров
    4. Экспорт значений параметров в Excel файл

    Args:
        base_url (str): Базовый URL API (например, "http://localhost:5000")
        parent (QtWidgets.QWidget, optional): Родительский виджет

    Побочные эффекты:
        - Отправляет HTTP запросы к API
        - Сохраняет Excel файл на диск
    """

    def __init__(self, base_url: str, parent=None):
        """
        Инициализация диалога экспорта параметров.

        Args:
            base_url: Базовый URL API
            parent: Родительский виджет
        """
        super().__init__(parent)
        self.base_url = base_url

        # Списки данных из API
        self._projects: List[Dict[str, Any]] = []   # Список проектов
        self._containers: List[Dict[str, Any]] = []  # Список контейнеров (моделей)
        self._param_data: List[Dict[str, Any]] = []  # Все параметры контейнеров
        self._selected_params: List[Dict[str, Any]] = []  # Выбранные параметры для экспорта

        # Строим UI и загружаем проекты
        self._build_ui()
        self._load_projects()

    def _build_ui(self):
        """
        Строит пользовательский интерфейс диалога.

        Создаёт следующие элементы:
        - Выпадающий список проектов
        - Список моделей с чекбоксами
        - Поле фильтрации параметров
        - Список параметров с чекбоксами
        - Кнопку экспорта
        """
        # Настраиваем окно диалога
        self.setWindowTitle("Экспорт параметров")
        self.resize(DIALOG_WIDTH, DIALOG_HEIGHT)

        # Основной вертикальный layout
        layout = QtWidgets.QVBoxLayout(self)
        layout.setSpacing(12)  # Отступ между элементами

        # 1. Группа выбора проекта
        proj_group = QtWidgets.QGroupBox("Выбор проекта")
        layout.addWidget(proj_group)
        proj_layout = QtWidgets.QVBoxLayout(proj_group)

        # Слой с выпадающим списком и кнопкой обновления
        proj_row = QtWidgets.QHBoxLayout()
        proj_row.addWidget(QtWidgets.QLabel("Проект:"))
        self.cmb_project = QtWidgets.QComboBox()
        proj_row.addWidget(self.cmb_project, 1)  # stretch=1 - занимает всё доступное место
        self.btn_refresh_proj = QtWidgets.QPushButton("Обновить")
        self.btn_refresh_proj.clicked.connect(self._load_projects)
        proj_row.addWidget(self.btn_refresh_proj)
        proj_layout.addLayout(proj_row)

        # 2. Группа выбора модели
        model_group = QtWidgets.QGroupBox("Выбор модели")
        layout.addWidget(model_group)
        model_layout = QtWidgets.QVBoxLayout(model_group)

        # Список моделей (каждая с чекбоксом)
        self.lst_models = QtWidgets.QListWidget()
        # Отключаем выделение строк, только чекбоксы
        self.lst_models.setSelectionMode(QtWidgets.QAbstractItemView.NoSelection)
        self.lst_models.setFocusPolicy(QtCore.Qt.NoFocus)
        # Стили для подсветки при наведении
        self.lst_models.setStyleSheet("""
            QListWidget::item { padding: 8px; }
            QListWidget::item:hover { background: #FFE3C2; color: #000000; }
        """)
        model_layout.addWidget(self.lst_models)

        # Кнопки выбора всех моделей и снятия выделения
        row_models = QtWidgets.QHBoxLayout()
        row_models.addStretch(1)
        self.btn_select_all_models = QtWidgets.QPushButton("Выбрать все")
        self.btn_select_all_models.clicked.connect(self._select_all_models)
        row_models.addWidget(self.btn_select_all_models)
        self.btn_clear_models = QtWidgets.QPushButton("Снять выделение")
        self.btn_clear_models.clicked.connect(self._clear_models)
        row_models.addWidget(self.btn_clear_models)
        row_models.addStretch(1)
        model_layout.addLayout(row_models)

        # 3. Группа выбора параметров
        params_group = QtWidgets.QGroupBox("Параметры для экспорта")
        layout.addWidget(params_group, 1)  # stretch=1 - занимает всё доступное место
        params_layout = QtWidgets.QVBoxLayout(params_group)

        # Слой с фильтром и кнопкой загрузки параметров
        row_filter = QtWidgets.QHBoxLayout()
        row_filter.addWidget(QtWidgets.QLabel("Фильтр:"))
        self.ed_filter = QtWidgets.QLineEdit()
        self.ed_filter.textChanged.connect(self._filter_params)
        row_filter.addWidget(self.ed_filter, 1)
        self.btn_load_params = QtWidgets.QPushButton("Загрузить параметры")
        self.btn_load_params.clicked.connect(self._load_params)
        row_filter.addWidget(self.btn_load_params)
        params_layout.addLayout(row_filter)

        # Список параметров (каждый с чекбоксом)
        self.lst_params = QtWidgets.QListWidget()
        self.lst_params.setSelectionMode(QtWidgets.QAbstractItemView.NoSelection)
        self.lst_params.setFocusPolicy(QtCore.Qt.NoFocus)
        self.lst_params.setAutoScroll(False)  # Отключаем автоскролл для производительности
        self.lst_params.setStyleSheet("""
            QListWidget::item { padding: 8px; }
            QListWidget::item:hover { background: #FFE3C2; color: #000000; }
        """)
        params_layout.addWidget(self.lst_params)

        # Кнопки выбора всех параметров и снятия выделения
        row_select_all = QtWidgets.QHBoxLayout()
        row_select_all.addStretch(1)
        self.btn_select_all = QtWidgets.QPushButton("Выбрать все")
        self.btn_select_all.clicked.connect(self._select_all)
        row_select_all.addWidget(self.btn_select_all)
        self.btn_clear_selection = QtWidgets.QPushButton("Снять выделение")
        self.btn_clear_selection.clicked.connect(self._clear_selection)
        row_select_all.addWidget(self.btn_clear_selection)
        row_select_all.addStretch(1)
        params_layout.addLayout(row_select_all)

        # 4. Кнопка экспорта
        row_export = QtWidgets.QHBoxLayout()
        row_export.addStretch(1)
        self.btn_export = QtWidgets.QPushButton("Экспортировать")
        self.btn_export.clicked.connect(self._export_to_excel)
        self.btn_export.setEnabled(False)  # Сначала отключаем
        row_export.addWidget(self.btn_export)
        row_export.addStretch(1)
        layout.addLayout(row_export)

        # 5. Статус бар (низ диалога)
        self.status_label = QtWidgets.QLabel()
        layout.addWidget(self.status_label)

        # 6. Подключаем сигналы (события)
        self.cmb_project.currentIndexChanged.connect(self._on_project_changed)
        self.lst_params.itemChanged.connect(self._on_param_selection_changed)

    def _load_projects(self):
        """
        Загружает список проектов из API.

        Выполняет GET запрос к /api/project/projects
        и заполняет выпадающий список проектов.

        Побочные эффекты:
            - Обновляет self._projects
            - Обновляет self.cmb_project
            - Обновляет self.status_label
        """
        try:
            # Проверка наличия библиотеки requests
            if requests is None:
                raise RuntimeError("Нужен requests: pip install requests")

            # Формируем URL для получения списка проектов
            url = f"{self.base_url.rstrip('/')}/api/project/projects"

            # Отправляем GET запрос к API
            resp = requests.get(url, timeout=REQUEST_TIMEOUT_SHORT)
            resp.raise_for_status()  # Вызывает исключение при статусе != 200

            # Парсим JSON ответ
            data = resp.json() or []

            # Цикл по списку: формируем список проектов с полями id и title
            self._projects = [
                {
                    "id": x.get("id"),
                    "title": x.get("title") or x.get("name") or f"ID {x.get('id')}"
                }
                for x in data
            ]

            # Очищаем и заполняем выпадающий список
            self.cmb_project.clear()
            for p in self._projects:
                self.cmb_project.addItem(p["title"])

            # Выбираем первый проект, если список не пустой
            if self._projects:
                self.cmb_project.setCurrentIndex(0)

            # Обновляем статус бар
            self.status_label.setText(f"Загружено {len(self._projects)} проектов")

        except Exception as e:
            # Обработка ошибок загрузки
            self.status_label.setText(f"Ошибка загрузки проектов: {e}")
            QtWidgets.QMessageBox.critical(self, "Ошибка", f"Не удалось загрузить проекты:\n{e}")

    def _on_project_changed(self):
        """
        Обработчик изменения выбранного проекта.

        При смене проекта загружает список моделей (контейнеров)
        для выбранного проекта.

        Побочные эффекты:
            - Вызывает _load_models()
        """
        idx = self.cmb_project.currentIndex()
        if idx < 0 or not self._projects:
            return

        # Получаем ID выбранного проекта
        project_id = self._projects[idx]["id"]
        self._load_models(project_id)

    def _load_models(self, project_id: int):
        """
        Загружает список моделей (контейнеров) проекта из API.

        Args:
            project_id: ID проекта для получения контейнеров

        Выполняет GET запрос к /api/imcContainer/getProjectImcContainers/{project_id}
        и заполняет список моделей с чекбоксами.

        Побочные эффекты:
            - Обновляет self._containers
            - Обновляет self.lst_models
            - Очищает список параметров
        """
        try:
            if requests is None:
                raise RuntimeError("Нужен requests: pip install requests")

            # Формируем URL для получения контейнеров проекта
            url = f"{self.base_url.rstrip('/')}/api/imcContainer/getProjectImcContainers/{project_id}"

            # Отправляем GET запрос
            resp = requests.get(url, timeout=REQUEST_TIMEOUT_SHORT)
            resp.raise_for_status()

            # Парсим JSON ответ
            data = resp.json() or []

            # Формируем список контейнеров с полями id и title
            self._containers = [
                {"id": x.get("id"), "title": x.get("title") or f"ID {x.get('id')}"}
                for x in data
            ]

            # Очищаем и заполняем список моделей
            self.lst_models.clear()
            # Цикл по контейнерам: создаём элементы списка с чекбоксами
            for c in self._containers:
                # Создаём элемент списка с чекбоксом
                item = QtWidgets.QListWidgetItem(c["title"])
                # Устанавливаем флаги: чекбокс активен и элемент доступен
                item.setFlags(QtCore.Qt.ItemIsUserCheckable | QtCore.Qt.ItemIsEnabled)
                # Устанавливаем начальное состояние: не отмечен
                item.setCheckState(QtCore.Qt.Unchecked)
                # Сохраняем данные контейнера в элементе списка
                item.setData(QtCore.Qt.UserRole, c)
                self.lst_models.addItem(item)

            # Обновляем статус бар
            self.status_label.setText(f"Загружено {len(self._containers)} моделей")

            # Очищаем список параметров, так как они зависят от модели
            self._clear_params()

        except Exception as e:
            self.status_label.setText(f"Ошибка загрузки моделей: {e}")
            QtWidgets.QMessageBox.critical(self, "Ошибка", f"Не удалось загрузить модели:\n{e}")

    def _load_params(self):
        """
        Загружает список параметров для выбранных моделей из API.

        Собирает параметры из всех отмеченных моделей.

        Побочные эффекты:
            - Обновляет self._param_data
            - Обновляет self.lst_params
            - Применяет фильтр к списку параметров
        """
        # Собираем ID выбранных контейнеров
        selected_containers = []
        for i in range(self.lst_models.count()):
            item = self.lst_models.item(i)
            if item.checkState() == QtCore.Qt.Checked:
                container = item.data(QtCore.Qt.UserRole)
                selected_containers.append(container)

        # Проверяем, что выбран хотя бы один контейнер
        if not selected_containers:
            self.status_label.setText("Выберите модель")
            QtWidgets.QMessageBox.warning(self, "Внимание", "Выберите модель для загрузки параметров")
            return

        try:
            if requests is None:
                raise RuntimeError("Нужен requests: pip install requests")

            # Загружаем параметры из всех выбранных контейнеров
            all_params = []
            # Цикл по выбранным контейнерам: загружаем параметры для каждого
            for container in selected_containers:
                container_id = container["id"]
                # Формируем URL для получения параметров контейнера
                url = f"{self.base_url.rstrip('/')}/api/imcParameterDefinition/imcParameterDefinitions"
                # Отправляем GET запрос с параметром containerIds
                resp = requests.get(url, params=[("containerIds", container_id)], timeout=REQUEST_TIMEOUT_SHORT)
                resp.raise_for_status()

                # Парсим JSON ответ
                data = resp.json() or []
                all_params.extend(data)

            # Сохраняем все параметры
            self._param_data = all_params

            # Применяем фильтр к списку параметров
            self._filter_params()

            # Обновляем статус бар
            self.status_label.setText(f"Загружено {len(all_params)} параметров из {len(selected_containers)} моделей")

        except Exception as e:
            self.status_label.setText(f"Ошибка загрузки параметров: {e}")
            QtWidgets.QMessageBox.critical(self, "Ошибка", f"Не удалось загрузить параметры:\n{e}")

    def _filter_params(self):
        """
        Фильтрует список параметров по тексту фильтра.

        Показывает только те параметры, у которых название или код
        содержит текст из поля фильтра.

        Побочные эффекты:
            - Обновляет self.lst_params
        """
        # Получаем текст фильтра в нижнем регистре
        filter_text = self.ed_filter.text().lower()

        # Фильтруем параметры по тексту
        filtered = []
        for p in self._param_data:
            # Получаем название параметра (title, name или code)
            name = (p.get("title") or p.get("name") or p.get("code") or "").lower()
            # Если фильтр пуст или содержится в названии - добавляем в список
            if not filter_text or filter_text in name:
                filtered.append(p)

        # Очищаем и заполняем список параметров
        self.lst_params.clear()
        for p in filtered:
            # Отображаем код параметра в списке
            display_name = p.get("code") or p.get("title") or p.get("name") or ""
            item = QtWidgets.QListWidgetItem(display_name)
            # Устанавливаем флаги: чекбокс активен и элемент доступен
            item.setFlags(QtCore.Qt.ItemIsUserCheckable | QtCore.Qt.ItemIsEnabled)
            # Устанавливаем начальное состояние: не отмечен
            item.setCheckState(QtCore.Qt.Unchecked)
            # Сохраняем данные параметра в элементе списка
            item.setData(QtCore.Qt.UserRole, p)
            self.lst_params.addItem(item)

    def _clear_params(self):
        """
        Очищает список параметров и сбрасывает состояние.

        Побочные эффекты:
            - Очищает self._param_data
            - Очищает self.lst_params
            - Очищает self._selected_params
            - Отключает кнопку экспорта
        """
        self._param_data = []
        self.lst_params.clear()
        self._selected_params = []
        self.btn_export.setEnabled(False)

    def _select_all(self):
        """
        Отмечает все параметры в списке.

        Побочные эффекты:
            - Устанавливает состояние Checked для всех элементов списка
        """
        for i in range(self.lst_params.count()):
            item = self.lst_params.item(i)
            item.setCheckState(QtCore.Qt.Checked)

    def _clear_selection(self):
        """
        Снимает выделение со всех параметров.

        Побочные эффекты:
            - Устанавливает состояние Unchecked для всех элементов списка
        """
        for i in range(self.lst_params.count()):
            item = self.lst_params.item(i)
            item.setCheckState(QtCore.Qt.Unchecked)

    def _on_param_selection_changed(self):
        """
        Обработчик изменения выделения параметров.

        Собирает все отмеченные параметры и включает кнопку экспорта,
        если выбран хотя бы один параметр.

        Побочные эффекты:
            - Обновляет self._selected_params
            - Обновляет состояние кнопки экспорта
        """
        # Собираем все отмеченные параметры
        self._selected_params = []
        for i in range(self.lst_params.count()):
            item = self.lst_params.item(i)
            if item.checkState() == QtCore.Qt.Checked:
                self._selected_params.append(item.data(QtCore.Qt.UserRole))

        # Включаем кнопку экспорта, если выбран хотя бы один параметр
        self.btn_export.setEnabled(len(self._selected_params) > 0)

    def _select_all_models(self):
        """
        Отмечает все модели в списке.

        Побочные эффекты:
            - Устанавливает состояние Checked для всех элементов списка
        """
        for i in range(self.lst_models.count()):
            item = self.lst_models.item(i)
            item.setCheckState(QtCore.Qt.Checked)

    def _clear_models(self):
        """
        Снимает выделение со всех моделей.

        Побочные эффекты:
            - Устанавливает состояние Unchecked для всех элементов списка
        """
        for i in range(self.lst_models.count()):
            item = self.lst_models.item(i)
            item.setCheckState(QtCore.Qt.Unchecked)

    def _export_to_excel(self):
        """
        Экспортирует выбранные параметры в Excel файл.

        Для каждой выбранной модели:
        1. Загружает все элементы модели
        2. Для каждого элемента загружает значения всех выбранных параметров
        3. Собирает данные в таблицу: ID | Параметр | Значение | Тип данных

        Сохраняет Excel файл с помощью openpyxl.

        Побочные эффекты:
            - Создаёт Excel файл на диске
            - Показывает диалог сохранения файла
        """
        # Проверяем, что выбраны параметры
        if not self._selected_params:
            return

        # Собираем выбранные модели
        selected_containers = []
        for i in range(self.lst_models.count()):
            item = self.lst_models.item(i)
            if item.checkState() == QtCore.Qt.Checked:
                container = item.data(QtCore.Qt.UserRole)
                selected_containers.append(container)

        # Проверяем, что выбраны модели
        if not selected_containers:
            QtWidgets.QMessageBox.warning(self, "Внимание", "Выберите модель")
            return

        # Показываем диалог сохранения файла
        path, _ = QtWidgets.QFileDialog.getSaveFileName(
            self,
            "Сохранить в Excel",
            "",
            "Excel (*.xlsx *.xls)"
        )

        # Если пользователь отменил диалог - выходим
        if not path:
            return

        # Проверяем наличие openpyxl
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment
            from openpyxl.worksheet.table import Table, TableStyleInfo
        except ImportError:
            QtWidgets.QMessageBox.critical(self, "Ошибка", "Нужен openpyxl: pip install openpyxl")
            return

        try:
            if requests is None:
                raise RuntimeError("Нужен requests: pip install requests")

            # Список для сбора данных
            data = []

            # Сначала подсчитываем общее количество элементов (для прогресса)
            total_elements = 0
            # Цикл по выбранным контейнерам: считаем элементы
            for container in selected_containers:
                container_id = container["id"]
                # Загружаем элементы контейнера
                url = f"{self.base_url.rstrip('/')}/api/imcElement/imcElements/{container_id}"
                resp = requests.get(url, timeout=REQUEST_TIMEOUT_LONG)
                resp.raise_for_status()
                elements_data = resp.json() or []
                elements = elements_data if isinstance(elements_data, list) else elements_data.get("items", [])

                # Накапливаем общее количество элементов
                total_elements += len(elements)

            # Теперь собираем данные
            processed_elements = 0

            # Цикл по выбранным контейнерам: загружаем элементы и их параметры
            for container in selected_containers:
                container_id = container["id"]

                # Загружаем элементы контейнера
                url = f"{self.base_url.rstrip('/')}/api/imcElement/imcElements/{container_id}"
                resp = requests.get(url, timeout=REQUEST_TIMEOUT_LONG)
                resp.raise_for_status()
                elements_data = resp.json() or []
                elements = elements_data if isinstance(elements_data, list) else elements_data.get("items", [])

                # Обновляем статус бар (прогресс)
                self.status_label.setText(f"Обработка {processed_elements + len(elements)}/{total_elements} элементов...")
                QtWidgets.QApplication.processEvents()  # Обновляем UI

                # Обрабатываем каждый элемент
                for el in elements:
                    # Получаем идентификаторы элемента
                    native_id = el.get("nativeId") or el.get("NativeId") or el.get("Id") or el.get("id")
                    real_id = el.get("Id") or el.get("id")

                    # Проверяем, что идентификаторы не пустые
                    if not native_id or not real_id:
                        continue

                    processed_elements += 1

                    # Получаем все параметры элемента
                    try:
                        param_url = f"{self.base_url.rstrip('/')}/api/imcParameterValue/imcParameterValues/{real_id}"
                        param_resp = requests.get(param_url, timeout=REQUEST_TIMEOUT_SHORT)
                        param_resp.raise_for_status()
                        param_data_el = param_resp.json() or []
                        param_values_list = param_data_el if isinstance(param_data_el, list) else param_data_el.get("items", [])

                        # Для каждого выбранного параметра проверяем значение у элемента
                        for param in self._selected_params:
                            param_id = param.get("Id") or param.get("id")
                            param_code = param.get("code")
                            param_name = param_code or param.get("title") or param.get("name") or str(param_id) or ""
                            is_numeric = bool(param.get("isNumeric", False))

                            param_value = ""

                            # Ищем значение параметра в списке значений элемента
                            for p in param_values_list:
                                p_param_def_id = p.get("parameterDefinitionId") or ""

                                # Проверяем по parameterDefinitionId
                                if p_param_def_id and str(p_param_def_id) == str(param_id):
                                    param_value = p.get("alternativeValueString") or p.get("originValueString") or p.get("value") or p.get("displayValue") or ""
                                    break

                                # Проверяем по code
                                p_code = p.get("code") or ""
                                if p_code == param_code:
                                    param_value = p.get("alternativeValueString") or p.get("originValueString") or p.get("value") or p.get("displayValue") or ""
                                    break

                            # Добавляем запись в список данных
                            data.append({
                                "ID": str(native_id),
                                "Параметр": param_name,
                                "Значение": param_value,
                                "Тип данных": PARAM_TYPE_NUMERIC if is_numeric else PARAM_TYPE_TEXT
                            })

                    except Exception as e:
                        # Логируем ошибку, но продолжаем обработку
                        print(f"Ошибка при получении параметров элемента {native_id}: {e}")
                        continue

            # Создаём Excel файл
            wb = openpyxl.Workbook()
            ws = wb.active

            # Устанавливаем заголовки
            headers = ["ID", "Параметр", "Значение", "Тип данных"]
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True, size=11)
                cell.alignment = Alignment(horizontal="left", vertical="center")

            # Устанавливаем ширину колонок
            ws.column_dimensions["A"].width = 12  # ID
            ws.column_dimensions["B"].width = 25  # Параметр
            ws.column_dimensions["C"].width = 25  # Значение
            ws.column_dimensions["D"].width = 15  # Тип данных

            # Устанавливаем высоту для заголовков
            ws.row_dimensions[1].height = 30

            # Заполняем данные
            for row_idx, row_data in enumerate(data, 2):
                for col_idx, value in enumerate(row_data.values(), 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.font = Font(size=10)
                    cell.alignment = Alignment(horizontal="left", vertical="center")

            # Оформляем таблицу с фильтрами
            table_ref = f"A1:D{1 + len(data)}"
            table = Table(displayName="ExportTable", ref=table_ref)
            table_style = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            table.tableStyleInfo = table_style
            ws.add_table(table)

            # Сохраняем файл
            wb.save(path)

            # Обновляем статус бар и показываем сообщение об успехе
            self.status_label.setText(f"Экспортировано {len(data)} записей")
            QtWidgets.QMessageBox.information(
                self,
                "Успех",
                f"Файл сохранён:\n{path}\n\nЗаписей: {len(data)}"
            )

        except Exception as e:
            # Обработка ошибок экспорта
            self.status_label.setText(f"Ошибка экспорта: {e}")
            QtWidgets.QMessageBox.critical(self, "Ошибка", f"Не удалось сохранить файл:\n{e}")
