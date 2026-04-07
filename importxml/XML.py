#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
XML change list uploader for BIM-Info API.
Reads a change list XML, lets the user pick local containers, and updates only changed values.
"""

from __future__ import annotations

import argparse
import json
import os
import sys
import xml.etree.ElementTree as ET
import re
import tempfile
from dataclasses import dataclass
from typing import Dict, Iterable, List, Optional, Tuple

import requests

try:
    import openpyxl
except Exception:
    openpyxl = None

try:
    from PySide6 import QtWidgets, QtCore
    from PySide6.QtGui import QIcon, QPixmap, QPalette, QColor, QPainter
    from PySide6.QtWidgets import (
        QApplication,
        QFileDialog,
        QHBoxLayout,
        QLabel,
        QLineEdit,
        QMainWindow,
        QMessageBox,
        QPushButton,
        QCheckBox,
        QComboBox,
        QScrollArea,
        QTabWidget,
        QTableWidget,
        QTableWidgetItem,
        QVBoxLayout,
        QWidget,
        QGroupBox,
        QFrame,
        QHeaderView,
        QAbstractItemView,
    )
    PYSIDE = True
except Exception:
    PYSIDE = False

APP_DIR = os.path.abspath(os.path.dirname(__file__))
ICON_DIR = os.path.join(APP_DIR, "icon")
LOGO_ICO_PATH = os.path.join(ICON_DIR, "logo.ico")

if PYSIDE:
    _BASE_FONT_FAMILY = "Segoe UI"
    _BASE_FONT_SIZE_PT = 10
    _CACHE_SUBDIR = "nik_style_cache"

    class _Palette:
        BG_LIGHT = "#FFFFFF"
        BG_DARK = "#1E1E1E"
        FG_LIGHT = "#000000"
        FG_DARK = "#FFFFFF"
        ACCENT = "#F7921E"
        ACCENT_HOVER = "#FFA74B"
        ACCENT_PRESSED = "#E07E12"
        GRAY = "#D9D9D9"
        GRAY_DARK = "#3A3A3A"
        BORDER_LIGHT = "#dcdcdc"
        BORDER_DARK = "#3A3A3A"
        SOFT_HOVER = "#FFE3C2"
        SELECTED = "#FFC37A"
        SCROLL_TRACK_LIGHT = "#FAFAFA"
        SCROLL_TRACK_DARK = "#252525"
        CHECKBOX_HOVER = "#E5E5E5"

    PALETTE = _Palette()

    _DEKSTOP_ICON_FILES = {
        "logo": ["Manager-scaled.png"],
        "logo_white": ["Manager-scaled_white.png"],
        "login": ["free-icon-login-2623062.png"],
        "folder": ["folder_icon_variant_1.png"],
        "save": ["free-icon-download-126488.png"],
        "plus": ["free-icon-plus-3303893.png"],
        "gear": ["free-icon-setting-3288004.png"],
        "eye_open": ["free-icon-eye-2455724.png"],
        "eye_closed": ["free-icon-hide-11238328.png"],
        "no_folder": ["no folder.png"],
        "delete": ["delete.png"],
        "edit": ["edit.png"],
        "refresh": ["free-icon-refresh-5234214.png"],
        "cad": ["free-icon-cad-8304395.png"],
        "filter": ["filter.png"],
        "flash": ["flash.png"],
        "warning": ["warning.png"],
        "structure": ["structure.png"],
        "insert": ["insert.png"],
        "back": ["back.png"],
        "sync": ["sync.png"],
        "arrow_down": ["arrow-down.png"],
        "arrow_up": ["arrow-up.png"],
        "arrow_right": ["arrow-right.png"],
        "arrow_left": ["arrow-left.png"],
        "sort_up": ["arrow-up.png"],
        "sort_down": ["arrow-down.png"],
        "arrow_down_free": ["arrow-down.png"],
        "rotate_left": ["rotate-left.png", "free-icon-rotate-left.png"],
        "rotate_right": ["rotate-right.png", "free-icon-rotate-right.png"],
        "sun": ["sun.png"],
        "moon": ["moon.png"],
        "check": ["check.png"],
        "select": ["select.png"],
        "circle2": ["krug.png"],
        "circle_dot": ["krug_galka.png"],
        "poloska": ["poloska.png"],
        "ok": ["ok.png"],
        "none": ["none.png"],
        "1": ["1.png"],
        "2": ["2.png"],
        "extend": ["extend.png"],
        "arrow_oba": ["arrow-oba.png"],
        "navigation": ["navigation.png"],
        "move": ["move.png"],
        "compare": ["compare.png"],
    }

    def _cache_dir(icon_dir: str) -> str:
        base = os.path.abspath(icon_dir or ICON_DIR)
        root = os.path.join(tempfile.gettempdir(), _CACHE_SUBDIR)
        os.makedirs(root, exist_ok=True)
        sub = os.path.join(root, os.path.basename(base) or "icon")
        os.makedirs(sub, exist_ok=True)
        return sub

    def _qss_url(path: str) -> str:
        return (path or "").replace("\\", "/")

    def _tint_pixmap(pm: QPixmap, color: QColor) -> QPixmap:
        if pm.isNull():
            return pm
        tinted = QPixmap(pm.size())
        tinted.fill(QtCore.Qt.transparent)
        painter = QPainter(tinted)
        painter.setCompositionMode(QPainter.CompositionMode_Source)
        painter.drawPixmap(0, 0, pm)
        painter.setCompositionMode(QPainter.CompositionMode_SourceIn)
        painter.fillRect(tinted.rect(), color)
        painter.end()
        return tinted

    def _ensure_color_copy(src_path: str, icon_dir: str, color: QColor, suffix: str) -> str:
        if not src_path or not os.path.exists(src_path):
            return src_path
        cache = _cache_dir(icon_dir)
        base = os.path.basename(src_path)
        name, ext = os.path.splitext(base)
        dst = os.path.join(cache, f"{name}_{suffix}{ext}")
        if os.path.exists(dst):
            return dst
        pm = QPixmap(src_path)
        if pm.isNull():
            return src_path
        pm = _tint_pixmap(pm, color)
        try:
            pm.save(dst)
            return dst
        except Exception:
            return src_path

    def _ensure_white_copy(src_path: str, icon_dir: str) -> str:
        return _ensure_color_copy(src_path, icon_dir, QColor("#FFFFFF"), "white")

    def _ensure_black_copy(src_path: str, icon_dir: str) -> str:
        return _ensure_color_copy(src_path, icon_dir, QColor("#000000"), "black")

    def resolve_icon_path(name: str, icon_dir: str = ICON_DIR, app=None) -> str:
        if not name:
            return ""
        if os.path.exists(name):
            path = os.path.abspath(name)
        else:
            path = ""
        if not path:
            candidates = _DEKSTOP_ICON_FILES.get(name, [])
            for fname in candidates:
                candidate = os.path.join(icon_dir, fname)
                if os.path.exists(candidate):
                    path = candidate
                    break
        if not path:
            direct = os.path.join(icon_dir, f"{name}.png")
            if os.path.exists(direct):
                path = direct
        if not path:
            return ""
        app = app or QtWidgets.QApplication.instance()
        if is_dark_theme(app) and name not in ("logo", "logo_white"):
            return _ensure_white_copy(path, icon_dir)
        return path

    def is_dark_theme(app: QtWidgets.QApplication | None) -> bool:
        if app is None:
            app = QtWidgets.QApplication.instance()
        if app is None:
            return False
        try:
            theme_prop = str(app.property("nik_theme") or "").strip().lower()
            if theme_prop in ("dark", "night"):
                return True
            if theme_prop in ("light", "day"):
                return False
        except Exception:
            pass
        try:
            col = app.palette().color(QPalette.Window)
            return col.lightness() < 128
        except Exception:
            return False

    def _qss_common(
        ar_down: str,
        ar_up: str,
        ar_left: str,
        ar_right: str,
        cmb_down: str,
        chk_off: str,
        chk_on: str,
        chk_mid: str,
        rchk_off: str,
        rchk_on: str,
        list_hover_off: str,
        list_hover_on: str,
        list_hover_mid: str,
        *,
        dark: bool,
    ) -> str:
        bg = PALETTE.BG_DARK if dark else PALETTE.BG_LIGHT
        fg = PALETTE.FG_DARK if dark else PALETTE.FG_LIGHT
        border = PALETTE.BORDER_DARK if dark else PALETTE.BORDER_LIGHT
        track_bg = bg if dark else PALETTE.SCROLL_TRACK_LIGHT

        ar_down = _qss_url(ar_down)
        ar_up = _qss_url(ar_up)
        ar_left = _qss_url(ar_left)
        ar_right = _qss_url(ar_right)
        cmb_down = _qss_url(cmb_down)

        chk_off = _qss_url(chk_off)
        chk_on = _qss_url(chk_on)
        chk_mid = _qss_url(chk_mid)
        rchk_off = _qss_url(rchk_off)
        rchk_on = _qss_url(rchk_on)
        list_hover_off = _qss_url(list_hover_off)
        list_hover_on = _qss_url(list_hover_on)
        list_hover_mid = _qss_url(list_hover_mid)

        hover_text = "#000000"

        qss = f"""
        * {{
            font-family: '{_BASE_FONT_FAMILY}';
            font-size: {_BASE_FONT_SIZE_PT}pt;
            color: {fg};
            selection-color: #000000;
        }}
        QWidget {{ background: {bg}; }}
        QStatusBar, QMenuBar, QToolBar, QMenu, QDockWidget::title {{ background: {bg}; border: 1px solid {border}; }}
        QTabBar::pane {{ background: {bg}; border: none; }}
        #header {{ background: {bg}; border: none; }}
        QTabWidget::pane {{ border: none; border-radius: 12px; margin-top: 8px; }}
        QTabBar::tab {{
            background: {bg};
            color: {fg};
            border: 1px solid {border};
            border-bottom-color: {border};
            padding: 6px 14px;
            border-top-left-radius: 10px;
            border-top-right-radius: 10px;
            margin: 0 4px;
        }}
        QTabBar::tab:hover {{
            background: {PALETTE.SOFT_HOVER};
            color: {hover_text};
            border-color: {PALETTE.ACCENT};
        }}
        QTabBar::tab:selected {{
            background: {PALETTE.SELECTED};
            color: {hover_text};
            border-color: {PALETTE.ACCENT};
        }}
        QTabBar::tab:!selected {{ margin-top: 6px; }}

        QPushButton {{
            background: {bg};
            color: {fg};
            border: 1px solid {border};
            border-radius: 14px;
            padding: 6px 12px;
        }}
        QPushButton:hover {{
            background: rgba(247, 146, 30, 0.15);
            color: {fg};
            border: 1px solid {PALETTE.ACCENT};
        }}
        QPushButton:pressed {{
            background: rgba(247, 146, 30, 0.25);
            border: 1px solid {PALETTE.ACCENT};
        }}
        QPushButton:disabled {{
            background: {'#2A2A2A' if dark else '#f0f0f0'};
            color: {'#8f8f8f' if dark else '#9b9b9b'};
        }}

        QPushButton[largeButton="true"] {{
            font-size: {_BASE_FONT_SIZE_PT + 2}pt;
            font-weight: 600;
            padding: 8px 14px;
        }}

        .btn-secondary {{
            background: {bg};
            color: {fg};
            border: 1px solid {border};
            border-radius: 12px;
            padding: 5px 10px;
        }}
        .btn-secondary:hover {{ background: rgba(247, 146, 30, 0.15); border: 1px solid {PALETTE.ACCENT}; }}

        QComboBox {{
            border: 1px solid {border};
            border-radius: 12px;
            padding: 3px 28px 3px 8px;
            background: {bg};
            selection-background-color: {PALETTE.SELECTED};
            selection-color: #000000;
        }}
        QComboBox::drop-down {{ width: 26px; border: none; }}
        QComboBox::down-arrow {{ image: url('{cmb_down}'); width: 12px; height: 12px; }}

        QLineEdit,
        QSpinBox, QDoubleSpinBox,
        QDateEdit, QTimeEdit, QDateTimeEdit {{
            border: 1px solid {border};
            border-radius: 12px;
            padding: 3px 8px;
            background: {bg};
            selection-background-color: {PALETTE.SELECTED};
        }}
        QLineEdit:hover,
        QSpinBox:hover, QDoubleSpinBox:hover,
        QDateEdit:hover, QTimeEdit:hover, QDateTimeEdit:hover {{
            background: {bg};
            color: {fg};
        }}
        QLineEdit:disabled,
        QSpinBox:disabled, QDoubleSpinBox:disabled,
        QDateEdit:disabled, QTimeEdit:disabled, QDateTimeEdit:disabled {{
            background: {'#2A2A2A' if dark else '#f0f0f0'};
            color: {'#8f8f8f' if dark else '#9b9b9b'};
        }}

        QComboBox QAbstractItemView {{
            background: {bg};
            border: 1px solid {border};
            outline: none;
            selection-background-color: {PALETTE.SELECTED};
        }}
        QComboBox QAbstractItemView::item {{ padding: 4px 8px; }}
        QComboBox QAbstractItemView::item:hover {{ background: {PALETTE.SOFT_HOVER}; color: {hover_text}; }}
        QComboBox QAbstractItemView::item:selected {{ background: {PALETTE.SELECTED}; color: {hover_text}; }}

        QListView, QListWidget {{
            background: {bg};
            border: 1px solid {border};
            outline: none;
            selection-background-color: {PALETTE.SELECTED};
            selection-color: {hover_text};
        }}
        QListView::item, QListWidget::item {{
            padding: 6px 8px;
            border: none;
        }}
        QListView::item:hover, QListWidget::item:hover {{
            background: {PALETTE.SOFT_HOVER};
            color: {hover_text};
        }}
        QListView::item:selected, QListWidget::item:selected {{
            background: {PALETTE.SELECTED};
            color: {hover_text};
        }}
        QListView::item:selected:active, QListWidget::item:selected:active {{
            background: {PALETTE.SELECTED};
        }}
        QListView::indicator, QListWidget::indicator {{
            width: 18px;
            height: 18px;
        }}
        QListView::indicator:unchecked, QListWidget::indicator:unchecked {{
            image: url('{chk_off}');
        }}
        QListView::indicator:checked, QListWidget::indicator:checked {{
            image: url('{chk_on}');
        }}
        QListView::indicator:indeterminate, QListWidget::indicator:indeterminate {{
            image: url('{chk_mid}');
        }}
        QListView::indicator:hover, QListWidget::indicator:hover {{
            background: transparent;
            border-radius: 0;
        }}

        QCheckBox {{ padding: 2px; color: {fg}; }}
        QCheckBox::indicator {{ width: 18px; height: 18px; }}
        QCheckBox::indicator:unchecked {{ image: url('{chk_off}'); }}
        QCheckBox::indicator:checked   {{ image: url('{chk_on}'); }}
        QCheckBox::indicator:indeterminate {{ image: url('{chk_mid}'); }}
        QCheckBox::indicator:hover {{
            background: transparent;
            border-radius: 0;
        }}
        QCheckBox[round="true"]::indicator {{ width: 18px; height: 18px; }}
        QCheckBox[round="true"]::indicator:hover {{
            background: transparent;
            border-radius: 0;
        }}
        QCheckBox:hover {{ color: {fg}; background: transparent; }}
        QRadioButton:hover {{ color: {fg}; background: transparent; }}

        QAbstractSpinBox::up-arrow {{ image: url('{ar_up}'); width: 12px; height: 12px; }}
        QAbstractSpinBox::down-arrow {{ image: url('{ar_down}'); width: 12px; height: 12px; }}
        QAbstractSpinBox::up-button, QAbstractSpinBox::down-button {{ background: transparent; border: none; }}

        QDateEdit::drop-down, QDateTimeEdit::drop-down {{ background: transparent; border: none; width: 22px; }}
        QDateEdit::down-arrow, QDateTimeEdit::down-arrow {{ image: url('{ar_down}'); width: 12px; height: 12px; margin-right: 4px; }}

        QHeaderView::section {{
            background: {bg};
            color: {fg};
            border: 1px solid transparent;
            padding: 6px 8px;
        }}
        QTableView, QTreeView {{
            gridline-color: transparent;
            outline: none;
            border: none;
            selection-background-color: {PALETTE.SELECTED};
        }}
        QHeaderView::up-arrow {{ image: url('{ar_up}'); width: 12px; height: 12px; }}
        QHeaderView::down-arrow {{ image: url('{ar_down}'); width: 12px; height: 12px; }}
        QHeaderView::section:hover {{
            background: #FFF3E6;
            color: {hover_text};
            border-color: #FFD1A0;
        }}
        QHeaderView::section:pressed {{
            background: #ffca91;
            color: {hover_text};
            border-color: #FFA74B;
        }}
        QHeaderView::section:hover,
        QHeaderView::section:pressed {{ border-radius: 8px; }}
        QHeaderView::section:first:hover,
        QHeaderView::section:first:pressed {{ border-top-left-radius: 8px; }}
        QHeaderView::section:last:hover,
        QHeaderView::section:last:pressed {{ border-top-right-radius: 8px; }}

        QScrollBar:vertical {{ background: {track_bg}; width: 12px; margin: 16px 0 16px 0; border: none; }}
        QScrollBar::handle:vertical {{ background: {PALETTE.SELECTED}; min-height: 24px; border-radius: 6px; border: none; }}
        QScrollBar::handle:vertical:hover {{ background: {PALETTE.ACCENT_HOVER}; }}
        QScrollBar::handle:vertical:pressed {{ background: {PALETTE.ACCENT_PRESSED}; }}
        QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{ background: transparent; height: 16px; }}
        QScrollBar::add-page:vertical, QScrollBar::sub-page:vertical {{ background: transparent; }}
        QScrollBar:horizontal {{ background: {track_bg}; height: 12px; margin: 0 16px 0 16px; border: none; }}
        QScrollBar::handle:horizontal {{ background: {PALETTE.SELECTED}; min-width: 24px; border-radius: 6px; border: none; }}
        QScrollBar::handle:horizontal:hover {{ background: {PALETTE.ACCENT_HOVER}; }}
        QScrollBar::handle:horizontal:pressed {{ background: {PALETTE.ACCENT_PRESSED}; }}
        QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal {{ background: transparent; width: 16px; }}
        QScrollBar::add-page:horizontal, QScrollBar::sub-page:horizontal {{ background: transparent; }}

        QScrollArea {{ border: none; }}

        QGroupBox {{
            border: 1px solid {border};
            border-radius: 12px;
            margin-top: 8px;
            padding: 8px;
        }}
        QGroupBox::title {{
            subcontrol-origin: margin;
            left: 10px;
            padding: 0 4px;
        }}
        QFrame[thinLine="true"] {{
            background: {border};
            min-height: 1px;
            max-height: 1px;
        }}
        QToolTip {{
            background: {PALETTE.SELECTED};
            color: #000000;
            border: 1px solid {PALETTE.ACCENT};
            padding: 4px 6px;
            border-radius: 6px;
        }}
        QLineEdit[search="true"] {{
            padding-left: 26px;
            background-image: url('{ar_right}');
            background-repeat: no-repeat;
            background-position: 6px 50%;
        }}
        """
        if dark:
            qss += f"""
            QListView::indicator:unchecked:hover, QListWidget::indicator:unchecked:hover {{ image: url('{list_hover_off}'); }}
            QListView::indicator:checked:hover,   QListWidget::indicator:checked:hover   {{ image: url('{list_hover_on}'); }}
            QListView::indicator:indeterminate:hover, QListWidget::indicator:indeterminate:hover {{ image: url('{list_hover_mid}'); }}
            """
        else:
            qss += f"""
            QListView::indicator:unchecked:hover, QListWidget::indicator:unchecked:hover {{ image: url('{chk_off}'); }}
            QListView::indicator:checked:hover,   QListWidget::indicator:checked:hover   {{ image: url('{chk_on}'); }}
            QListView::indicator:indeterminate:hover, QListWidget::indicator:indeterminate:hover {{ image: url('{chk_mid}'); }}
            """

        qss += f"""
        *[noCheckHoverRecolor="true"] QListView::indicator:unchecked:hover,
        *[noCheckHoverRecolor="true"] QListWidget::indicator:unchecked:hover {{ image: url('{chk_off}'); }}
        *[noCheckHoverRecolor="true"] QListView::indicator:checked:hover,
        *[noCheckHoverRecolor="true"] QListWidget::indicator:checked:hover {{ image: url('{chk_on}'); }}
        *[noCheckHoverRecolor="true"] QListView::indicator:indeterminate:hover,
        *[noCheckHoverRecolor="true"] QListWidget::indicator:indeterminate:hover {{ image: url('{chk_mid}'); }}
        """
        return qss

    def apply_app_style(app: QtWidgets.QApplication, *, theme: str | None = None, icon_dir: str = ICON_DIR) -> None:
        if theme is not None:
            app.setProperty("nik_theme", theme.lower())
        dark = is_dark_theme(app)

        ar_down = resolve_icon_path("arrow_down", icon_dir, app=app)
        if dark and ar_down:
            ar_down = _ensure_white_copy(ar_down, icon_dir)
        ar_up = resolve_icon_path("arrow_up", icon_dir, app=app)
        ar_right = resolve_icon_path("arrow_right", icon_dir, app=app)
        ar_left = resolve_icon_path("arrow_left", icon_dir, app=app)
        if dark and ar_left:
            ar_left = _ensure_white_copy(ar_left, icon_dir)

        cmb_down = resolve_icon_path("arrow_down", icon_dir, app=app) or ar_down
        if dark and cmb_down:
            cmb_down = _ensure_white_copy(cmb_down, icon_dir)

        chk_off = resolve_icon_path("check", icon_dir, app=app)
        chk_on = resolve_icon_path("select", icon_dir, app=app)
        chk_mid = resolve_icon_path("poloska", icon_dir, app=app)
        if dark:
            if chk_off:
                chk_off = _ensure_white_copy(chk_off, icon_dir)
            if chk_on:
                chk_on = _ensure_white_copy(chk_on, icon_dir)
            if chk_mid:
                chk_mid = _ensure_white_copy(chk_mid, icon_dir)
        else:
            if chk_off:
                chk_off = _ensure_black_copy(chk_off, icon_dir)
            if chk_on:
                chk_on = _ensure_black_copy(chk_on, icon_dir)
            if chk_mid:
                chk_mid = _ensure_black_copy(chk_mid, icon_dir)

        rchk_off = resolve_icon_path("circle2", icon_dir, app=app)
        rchk_on = resolve_icon_path("circle_dot", icon_dir, app=app)
        if dark:
            if rchk_off:
                rchk_off = _ensure_white_copy(rchk_off, icon_dir)
            if rchk_on:
                rchk_on = _ensure_white_copy(rchk_on, icon_dir)

        if dark:
            list_hover_off = _ensure_black_copy(chk_off, icon_dir) if chk_off else ""
            list_hover_on = _ensure_black_copy(chk_on, icon_dir) if chk_on else ""
            list_hover_mid = _ensure_black_copy(chk_mid, icon_dir) if chk_mid else ""
        else:
            list_hover_off, list_hover_on, list_hover_mid = chk_off or "", chk_on or "", chk_mid or ""

        app.setStyleSheet(_qss_common(
            ar_down or "", ar_up or "", ar_left or "", ar_right or "",
            cmb_down or "",
            chk_off or "", chk_on or "", chk_mid or "",
            rchk_off or "", rchk_on or "",
            list_hover_off, list_hover_on, list_hover_mid,
            dark=dark,
        ))

    def load_logo(icon_dir: str = ICON_DIR) -> QPixmap:
        app = QtWidgets.QApplication.instance()
        dark = is_dark_theme(app)
        name = "logo_white" if dark else "logo"
        p = resolve_icon_path(name, icon_dir)
        if not p:
            p = resolve_icon_path("logo", icon_dir)
        pm = QPixmap(p) if p else QPixmap()
        return pm if not pm.isNull() else QPixmap()

    def set_header_logo(label: QLabel, icon_dir: str, height: int = 48) -> None:
        pm = load_logo(icon_dir)
        if not pm.isNull():
            pm = pm.scaledToHeight(height, QtCore.Qt.SmoothTransformation)
            label.setPixmap(pm)
            label.setAlignment(QtCore.Qt.AlignCenter)

    class ThemeSwitch(QtWidgets.QAbstractButton):
        toggledTheme = QtCore.Signal(str) if hasattr(QtCore, "Signal") else QtCore.pyqtSignal(str)
        _anim_t = 0.0

        def __init__(self, parent=None, icon_dir: str = ICON_DIR):
            super().__init__(parent)
            self.setObjectName("themeToggle")
            self.setCheckable(True)
            self._icon_dir = icon_dir
            self.setCursor(QtCore.Qt.PointingHandCursor)
            self.setFocusPolicy(QtCore.Qt.NoFocus)
            self.setToolTip("Toggle theme")
            self.setMinimumSize(58, 26)
            self.setMaximumHeight(28)
            self._anim = QtCore.QPropertyAnimation(self, b"anim_t", self)
            self._anim.setDuration(180)
            self._anim.setEasingCurve(QtCore.QEasingCurve.OutCubic)
            self.toggled.connect(self._on_toggled)
            app = QtWidgets.QApplication.instance()
            self.setChecked(is_dark_theme(app))
            self._sun_icon = self._load_icon_pm("sun")
            self._moon_icon = self._load_icon_pm("moon")

        @QtCore.Property(float)
        def anim_t(self):
            return self._anim_t

        @anim_t.setter
        def anim_t(self, v: float):
            self._anim_t = max(0.0, min(1.0, float(v)))
            self.update()

        def _on_toggled(self, checked: bool):
            self._anim.stop()
            self._anim.setStartValue(self._anim_t)
            self._anim.setEndValue(1.0 if checked else 0.0)
            self._anim.start()
            app = QtWidgets.QApplication.instance()
            if app is None:
                return
            new_theme = "dark" if checked else "light"
            apply_app_style(app, theme=new_theme, icon_dir=self._icon_dir)
            self.toggledTheme.emit(new_theme)

        def sizeHint(self):
            return QtCore.QSize(66, 28)

        def minimumSizeHint(self):
            return QtCore.QSize(58, 26)

        def _load_icon_pm(self, name: str) -> QPixmap:
            p = resolve_icon_path(name, self._icon_dir)
            if not p:
                candidate = os.path.join(self._icon_dir, f"{name}.png")
                if os.path.exists(candidate):
                    p = candidate
            if p and os.path.exists(p):
                pm = QPixmap(p)
                if not pm.isNull():
                    return pm
            pm = QPixmap(24, 24)
            pm.fill(QtCore.Qt.transparent)
            return pm

        def paintEvent(self, e):
            painter = QPainter(self)
            try:
                painter.setRenderHint(QPainter.Antialiasing)
                rect = self.rect()
                track_rect = QtCore.QRectF(rect.adjusted(1, 1, -1, -1))
                dark = self.isChecked()

                track_color = QColor("#555555") if dark else QColor("#e8e8e8")
                painter.setPen(QtCore.Qt.NoPen)
                painter.setBrush(track_color)
                radius = track_rect.height() / 2.0
                painter.drawRoundedRect(track_rect, radius, radius)

                knob_margin = 3
                knob_d = track_rect.height() - knob_margin * 2
                knob_x = track_rect.left() + knob_margin + (track_rect.width() - 2 * knob_margin - knob_d) * self._anim_t
                knob_rect = QtCore.QRectF(knob_x, track_rect.top() + knob_margin, knob_d, knob_d)
                knob_color = QColor("#101010") if dark else QColor("#fdfdfd")
                painter.setBrush(knob_color)
                painter.drawEllipse(knob_rect)

                icon_size = int(track_rect.height() * 0.5)
                center_y = track_rect.center().y()
                left_x = track_rect.left() + 6
                right_x = track_rect.right() - icon_size - 6
                sun_on_left = self._anim_t >= 0.5
                sun_x = left_x if sun_on_left else right_x
                moon_x = right_x if sun_on_left else left_x
                if icon_size > 0:
                    sun_pm = QPixmap()
                    moon_pm = QPixmap()
                    tint = QColor(QtCore.Qt.white) if dark else QColor("#222222")
                    if not self._sun_icon.isNull():
                        _sun = self._sun_icon.scaled(icon_size, icon_size, QtCore.Qt.KeepAspectRatio, QtCore.Qt.SmoothTransformation)
                        sun_pm = _tint_pixmap(_sun, tint)
                    if not self._moon_icon.isNull():
                        _moon = self._moon_icon.scaled(icon_size, icon_size, QtCore.Qt.KeepAspectRatio, QtCore.Qt.SmoothTransformation)
                        moon_pm = _tint_pixmap(_moon, tint)

                    painter.save()
                    painter.setPen(QtCore.Qt.NoPen)
                    if not dark and not sun_pm.isNull():
                        highlight_size = icon_size + 8
                        painter.setBrush(QColor("#F7921E"))
                        painter.drawEllipse(QtCore.QRectF(
                            sun_x - (highlight_size - icon_size) / 2,
                            center_y - highlight_size / 2,
                            highlight_size,
                            highlight_size,
                        ))
                    elif dark and not moon_pm.isNull():
                        highlight_size = icon_size + 8
                        painter.setBrush(QColor("#F7921E"))
                        painter.drawEllipse(QtCore.QRectF(
                            moon_x - (highlight_size - icon_size) / 2,
                            center_y - highlight_size / 2,
                            highlight_size,
                            highlight_size,
                        ))
                    painter.restore()

                    if not sun_pm.isNull():
                        painter.drawPixmap(int(sun_x), int(center_y - sun_pm.height() / 2), sun_pm)
                    if not moon_pm.isNull():
                        painter.drawPixmap(int(moon_x), int(center_y - moon_pm.height() / 2), moon_pm)
            finally:
                painter.end()
else:
    apply_app_style = None
    ThemeSwitch = None
    set_header_logo = None


# ============================================================================
# КОНСТАНТЫ API
# ============================================================================

# Базовый URL API по умолчанию
INTERNAL_BASE_URL = "http://localhost:5000/api"

# Заголовки HTTP запросов по умолчанию
INTERNAL_HEADERS = {"accept": "*/*", "Content-Type": "application/json"}

# URL эндпоинтов API
INTERNAL_PROJECTS_URL = "/project/projects"
INTERNAL_CONTAINERS_URL = "/imcContainer/getProjectImcContainers/{project_id}"
INTERNAL_ELEMENTS_URL = "/imcElement/imcElements/{container_id}"
INTERNAL_PARAM_DEFS_URL = "/imcParameterDefinition/imcParameterDefinitions"
INTERNAL_PARAM_DEF_CREATE_URL = "/imcParameterDefinition/imcParameterDefinition"
INTERNAL_PARAM_VALUES_URL = "/imcParameterValue/setAlternateValueByElements"
INTERNAL_PARAM_VALUES_GET_URL = "/imcParameterValue/imcParameterValues/{element_id}"

# ============================================================================
# КОНСТАНТЫ ДЛЯ ПАРАМЕТРОВ
# ============================================================================

# Значение layer для создания параметра
PARAMETER_LAYER = 2

# Значение reportColumnType для создания параметра
PARAMETER_REPORT_COLUMN_TYPE = 3

# ============================================================================
# РЕГУЛЯРНЫЕ ВЫРАЖЕНИЯ
# ============================================================================

# Регулярное выражение для GUID с суффиксом (формат: xxxxxxxx-xxxx-xxxx-xxxx-xxxxxxxx-xxxxx)
_GUID_SUFFIX_RE = re.compile(r"^[0-9a-fA-F-]{36}-[0-9a-fA-F]{8}$")


def _normalize_base_url(raw: str) -> str:
    value = (raw or "").strip()
    if not value:
        return INTERNAL_BASE_URL
    if value.isdigit():
        return f"http://localhost:{value}/api"
    if value.startswith("http://") or value.startswith("https://"):
        return value.rstrip("/")
    return f"http://{value}".rstrip("/")


def _set_internal_base_url(raw: str) -> None:
    global INTERNAL_BASE_URL
    INTERNAL_BASE_URL = _normalize_base_url(raw)


def _set_dark_title_bar(widget, enable: bool = True) -> None:
    if sys.platform != "win32":
        return
    try:
        import ctypes
    except Exception:
        return
    try:
        hwnd = int(widget.winId())
        value = ctypes.c_int(1 if enable else 0)
        size = ctypes.sizeof(value)
        for attr in (20, 19):
            try:
                ctypes.windll.dwmapi.DwmSetWindowAttribute(hwnd, attr, ctypes.byref(value), size)
            except Exception:
                continue
    except Exception:
        pass


def _is_dark_theme(app: QtWidgets.QApplication | None) -> bool:
    if app is None:
        return False
    try:
        theme_prop = str(app.property("nik_theme") or "").strip().lower()
        if theme_prop in ("dark", "night"):
            return True
        if theme_prop in ("light", "day"):
            return False
    except Exception:
        pass
    try:
        col = app.palette().color(QPalette.Window)
        return col.lightness() < 128
    except Exception:
        return False


@dataclass(frozen=True)
class Change:
    identifier: str
    code: str
    old_value: str
    new_value: str
    is_numeric: bool
    uom: str


@dataclass(frozen=True)
class ParameterDef:
    code: str
    title: str
    is_numeric: bool
    uom: str


@dataclass(frozen=True)
class ReportItem:
    model: str
    identifier: str
    code: str
    old_value: str
    new_value: str


@dataclass(frozen=True)
class ReportData:
    updated: List[ReportItem]
    skipped_same: List[ReportItem]
    not_found: List[ReportItem]
    failed_update: List[ReportItem]


def api_get(url: str, headers: Dict[str, str], params: Optional[Dict[str, str]] = None, timeout: int = 10):
    """
    Выполняет GET запрос к API.

    Args:
        url: Полный URL запроса
        headers: Заголовки HTTP запроса
        params: Параметры запроса (query string)
        timeout: Таймаут в секундах

    Returns:
        Распарсенный JSON ответ или None при ошибке

    Побочные эффекты:
        Нет
    """
    try:
        # Отправляем GET запрос к API
        r = requests.get(url, headers=headers, params=params, timeout=timeout)
        r.raise_for_status()  # Вызывает исключение при статусе != 2xx
        return r.json()
    except Exception:
        return None


def api_post_json(url: str, headers: Dict[str, str], payload: Dict, timeout: int = 20) -> Tuple[bool, str]:
    """
    Выполняет POST запрос к API с JSON payload.

    Args:
        url: Полный URL запроса
        headers: Заголовки HTTP запроса
        payload: Данные для отправки в теле запроса (будут сериализованы в JSON)
        timeout: Таймаут в секундах

    Returns:
        Кортеж (success, error_message):
        - success: True если запрос успешен (статус 200, 201 или 204), иначе False
        - error_message: Пустая строка при успехе, иначе текст ошибки

    Побочные эффекты:
        Отправляет HTTP запрос к серверу
    """
    try:
        # Отправляем POST запрос с JSON payload
        r = requests.post(url, headers=headers, data=json.dumps(payload, ensure_ascii=False), timeout=timeout)
        if r.status_code in (200, 201, 204):
            return True, ""
        return False, r.text
    except Exception as exc:
        return False, str(exc)


def choose_item(items: List[Dict], label_key: str, id_key: str) -> Tuple[int, str]:
    """
    Интерактивный выбор одного элемента из списка через консоль.

    Args:
        items: Список словарей с элементами
        label_key: Ключ для отображаемого названия элемента
        id_key: Ключ для получения идентификатора элемента

    Returns:
        Кортеж (item_id, item_label) выбранного элемента

    Побочные эффекты:
        Выводит список элементов в консоль и ожидает ввод пользователя
    """
    # Цикл: выводим все элементы с номерами для выбора
    for idx, item in enumerate(items, start=1):
        label = str(item.get(label_key, ""))
        item_id = item.get(id_key)
        print(f"{idx}. {label} | {item_id}")
    
    # Цикл: ожидаем корректный ввод номера элемента
    while True:
        raw = input("Select number: ").strip()
        if not raw.isdigit():
            print("Enter a number.")
            continue
        pos = int(raw)
        if 1 <= pos <= len(items):
            chosen = items[pos - 1]
            return chosen[id_key], str(chosen.get(label_key, ""))
        print("Out of range.")


def read_xml_changes(xml_path: str) -> List[Change]:
    """
    Читает изменения параметров из XML файла.

    Формат XML:
    <root>
        <pvc enuid="element_id" pc="param_code">
            <v1>old_value</v1>
            <v2>new_value</v2>
        </pvc>
        ...
    </root>

    Args:
        xml_path: Путь к XML файлу

    Returns:
        Список объектов Change с изменениями параметров

    Raises:
        ValueError: Если XML не содержит элементов pvc или отсутствуют обязательные атрибуты

    Побочные эффекты:
        Читает файл с диска
    """
    # Читаем и парсим XML файл
    raw = open(xml_path, "rb").read()
    root = ET.fromstring(raw)
    changes: List[Change] = []
    
    # Цикл: ищем все элементы <pvc> и извлекаем атрибуты
    for pvc in root.findall(".//pvc"):
        enuid = (pvc.attrib.get("enuid") or "").strip()
        code = (pvc.attrib.get("pc") or "").strip()
        old_value = (pvc.findtext("v1") or "").strip()
        new_value = (pvc.findtext("v2") or "").strip()
        
        # Пропускаем элементы без обязательных атрибутов
        if not enuid or not code:
            continue
        changes.append(Change(identifier=enuid, code=code, old_value=old_value, new_value=new_value, is_numeric=False, uom=""))
    
    if not changes:
        raise ValueError("XML не содержит элементов pvc или отсутствуют обязательные атрибуты enuid и pc")
    return changes


def _normalize_header(value: object) -> str:
    """
    Нормализует заголовок столбца Excel.

    Преобразует заголовок к нижнему регистру, убирает пробелы.

    Args:
        value: Исходное значение заголовка

    Returns:
        Нормализованная строка заголовка

    Побочные эффекты:
        Нет
    """
    return str(value or "").strip().lower()


def _find_header_index(header_map: Dict[str, int], candidates: Iterable[str]) -> Optional[int]:
    """
    Находит индекс столбца по списку кандидатов названий.

    Args:
        header_map: Словарь {нормализованный_заголовок: индекс_колонки}
        candidates: Список кандидатов для поиска заголовка

    Returns:
        Индекс найденной колонки или None если не найдено

    Побочные эффекты:
        Нет
    """
    # Цикл: перебираем кандидатов пока не найдём совпадение в header_map
    for candidate in candidates:
        idx = header_map.get(candidate)
        if idx is not None:
            return idx
    return None


def _parse_bool(value: object) -> bool:
    """
    Парсит булево значение из строки.

    Args:
        value: Исходное значение

    Returns:
        True если значение означает "истина", иначе False

    Побочные эффекты:
        Нет
    """
    val = str(value or "").strip().lower()
    return val in ("1", "true", "yes", "y", "да")


def _parse_numeric(value: object) -> bool:
    """
    Определяет, является ли строка числовым типом.

    Args:
        value: Исходное значение

    Returns:
        True если значение означает числовой тип, иначе False

    Побочные эффекты:
        Нет
    """
    val = str(value or "").strip().lower()
    if not val:
        return False
    if _parse_bool(val):
        return True
    # Проверяем наличие токенов, указывающих на числовой тип
    numeric_tokens = ("number", "numeric", "int", "integer", "float", "double", "?????", "??????")
    return any(token in val for token in numeric_tokens)


def _parse_float(value: object) -> Optional[float]:
    """
    Парсит дробное число из значения.

    Args:
        value: Исходное значение (может быть числом или строкой)

    Returns:
        Число с плавающей точкой или None если не удаётся распарсить

    Побочные эффекты:
        Нет
    """
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    # Заменяем запятую на точку (европейский формат чисел)
    val = str(value).strip().replace(",", ".")
    if not val:
        return None
    try:
        return float(val)
    except Exception:
        return None


def _normalize_parameter_code(raw: str) -> str:
    """
    Нормализует код параметра из Excel.
    Сохраняет ведущий обратный слеш если он есть (например, \Комментарии).
    Убирает обрамляющие кавычки, экранирование и невидимые символы.
    """
    code = (raw or "").strip()
    
    # Убираем невидимые символы (BOM, non-breaking spaces и т.д.)
    code = code.replace('\u200b', '').replace('\ufeff', '').replace('\xa0', ' ')
    
    # Убираем обрамляющие кавычки
    if len(code) >= 2 and code[0] == '"' and code[-1] == '"':
        code = code[1:-1]
    elif len(code) >= 2 and code[0] == "'" and code[-1] == "'":
        code = code[1:-1]
    
    # Превращаем \\ в \ (экранированный слеш)
    code = code.replace("\\\\", "\\")
    
    # Финальный trim
    code = code.strip()
    
    # Если пусто после нормализации - возвращаем как есть
    if not code:
        return (raw or "").strip()
    
    return code


def _find_parameter_code_in_system(code_normalized: str, params: List[Dict]) -> str:
    """
    Находит реальный код параметра в системе.
    Учитывает вариант с ведущим слешем и без него.
    Возвращает код который реально есть в системе.
    """
    print(f"🔍 Поиск параметра в системе: {repr(code_normalized)}")
    
    # Сначала ищем точное совпадение
    for p in params:
        if p.get("code") == code_normalized:
            print(f"   ✅ Найден точное совпадение: {repr(code_normalized)}")
            return code_normalized
    
    # Если код начинается с \, ищем без слеша
    if code_normalized.startswith("\\"):
        code_without_slash = code_normalized[1:]
        print(f"   Проверяем вариант без слеша: {repr(code_without_slash)}")
        for p in params:
            if p.get("code") == code_without_slash:
                print(f"   ✅ Найден вариант без слеша: {repr(code_without_slash)}")
                return code_without_slash
        print(f"   ❌ Вариант без слеша не найден")
    
    # Если код НЕ начинается с \, ищем со слешем
    if not code_normalized.startswith("\\"):
        code_with_slash = "\\" + code_normalized
        print(f"   Проверяем вариант со слешем: {repr(code_with_slash)}")
        for p in params:
            if p.get("code") == code_with_slash:
                print(f"   ✅ Найден вариант со слешем: {repr(code_with_slash)}")
                return code_with_slash
        print(f"   ❌ Вариант со слешем не найден")
    
    # Не нашли - возвращаем исходный
    print(f"   ❌ Параметр не найден, возвращаем исходный: {repr(code_normalized)}")
    return code_normalized


def _validate_parameter_payload(code: str, is_numeric: bool, value: str, element_ids: List[int]) -> Tuple[bool, str, str]:
    """
    Валидирует payload перед отправкой.
    Возвращает (is_valid, error_message, normalized_value).
    """
    if not code or not code.strip():
        return False, "Parameter code is empty", value
    
    if not element_ids:
        return False, "elementIds list is empty", value
    
    if any(eid is None or eid == 0 for eid in element_ids):
        return False, f"elementIds contains None or 0: {element_ids}", value
    
    # Проверка числового значения
    if is_numeric:
        try:
            parsed = float(value)
            return True, "", str(parsed)
        except ValueError:
            return False, f"isNumeric=true but value is not a number: {value}", value
    else:
        # Для строкового параметра
        if not value or not value.strip():
            # Пустое значение допустимо
            return True, "", ""
        return True, "", value.strip()


def read_excel_payload(excel_path: str) -> Tuple[List[Change], List[ParameterDef]]:
    if openpyxl is None:
        raise ImportError("openpyxl не установлен")
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        raise ValueError("Excel не содержит строку заголовков")
    headers = [_normalize_header(cell) for cell in header_row]
    header_map = {header: idx for idx, header in enumerate(headers) if header}
    
    # Формат с типом данных (ID, Параметр, Значение, Тип данных)
    id_idx = _find_header_index(header_map, ("id",))
    param_idx = _find_header_index(header_map, ("параметр", "parameter", "code"))
    value_idx = _find_header_index(header_map, ("значение", "value", "v2"))
    data_type_idx = _find_header_index(header_map, ("тип данных", "типданных", "datatype", "type", "data_type", "data type"))
    
    if id_idx is not None and param_idx is not None and value_idx is not None:
        # Формат с типом данных
        changes: List[Change] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            elem_id = str(row[id_idx] or "").strip()
            code_raw = str(row[param_idx] or "")
            code = _normalize_parameter_code(code_raw)
            new_value = str(row[value_idx] or "").strip()
            
            # Определяем тип данных из столбца
            is_numeric = False
            if data_type_idx is not None:
                type_value = str(row[data_type_idx] or "").strip().lower()
                is_numeric = type_value in ("число", "числовой", "numeric", "number", "integer", "int", "float", "double", "decimal")
            
            # Лог нормализации (для диагностики)
            if code_raw != code:
                print(f"🔄 Normalized code: {repr(code_raw)} -> {repr(code)}")
            
            if not elem_id or not code:
                continue
            changes.append(Change(identifier=elem_id, code=code, old_value="", new_value=new_value, is_numeric=is_numeric, uom=""))
        if not changes:
            raise ValueError("Excel не содержит данных для импорта")
        return changes, []
    
    # Формат без типа данных (ID, Параметр, Значение)
    if id_idx is not None and param_idx is not None and value_idx is not None:
        changes: List[Change] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            elem_id = str(row[id_idx] or "").strip()
            code_raw = str(row[param_idx] or "")
            code = _normalize_parameter_code(code_raw)
            new_value = str(row[value_idx] or "").strip()
            
            # Лог нормализации
            if code_raw != code:
                print(f"🔄 Normalized code: {repr(code_raw)} -> {repr(code)}")
            
            if not elem_id or not code:
                continue
            # Тип данных по умолчанию — текстовый
            changes.append(Change(identifier=elem_id, code=code, old_value="", new_value=new_value, is_numeric=False, uom=""))
        if not changes:
            raise ValueError("Excel не содержит данных для импорта")
        return changes, []
    
    if param_idx is None:
        raise ValueError("Excel не содержит столбцов для параметров (Параметр/Parameter/Code)")
    
    title_idx = _find_header_index(header_map, ("название", "наименование", "title", "name"))
    uom_idx = _find_header_index(header_map, ("ед", "ед.изм", "едизм", "uom", "unit", "units"))
    numeric_idx = _find_header_index(header_map, ("числовой", "numeric", "isnumeric", "тип", "datatype"))
    
    param_defs: List[ParameterDef] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        code_raw = str(row[param_idx] or "")
        code = _normalize_parameter_code(code_raw)
        if not code:
            continue
        title = str(row[title_idx] or "").strip() if title_idx is not None else ""
        uom = str(row[uom_idx] or "").strip() if uom_idx is not None else ""
        
        # Определяем тип данных из столбца
        is_numeric = False
        if numeric_idx is not None:
            type_value = str(row[numeric_idx] or "").strip().lower()
            is_numeric = type_value in ("число", "числовой", "numeric", "number", "integer", "int", "float", "double", "decimal")
        
        # Лог нормализации
        if code_raw != code:
            print(f"🔄 Normalized code: {repr(code_raw)} -> {repr(code)}")
        
        if not title:
            title = f"Created from Excel: {code}"
        param_defs.append(ParameterDef(code=code, title=title, is_numeric=is_numeric, uom=uom))
    
    return [], param_defs


def choose_items(items: List[Dict], label_key: str, id_key: str) -> List[int]:
    """
    Интерактивный выбор нескольких элементов из списка через консоль.

    Args:
        items: Список словарей с элементами
        label_key: Ключ для отображаемого названия элемента
        id_key: Ключ для получения идентификатора элемента

    Returns:
        Список ID выбранных элементов

    Побочные эффекты:
        Выводит список элементов в консоль и ожидает ввод пользователя
    """
    # Цикл: выводим все элементы с номерами для выбора
    for idx, item in enumerate(items, start=1):
        label = str(item.get(label_key, ""))
        item_id = item.get(id_key)
        print(f"{idx}. {label} | {item_id}")
    
    # Цикл: ожидаем корректный ввод номеров элементов через запятую
    while True:
        raw = input("Введите номера через запятую: ").strip()
        if not raw:
            print("Введите хотя бы один номер.")
            continue
        parts = [p.strip() for p in raw.split(",") if p.strip()]
        if not all(p.isdigit() for p in parts):
            print("Введите только цифры, разделенные запятыми.")
            continue
        idxs = [int(p) for p in parts]
        if any(i < 1 or i > len(items) for i in idxs):
            print("Номер вне диапазона.")
            continue
        return [items[i - 1][id_key] for i in idxs]


def load_local_projects() -> List[Dict]:
    """
    Загружает список всех проектов из API.

    Returns:
        Список словарей с проектами

    Побочные эффекты:
        Выполняет GET запрос к API
    """
    data = api_get(f"{INTERNAL_BASE_URL}{INTERNAL_PROJECTS_URL}", {"accept": "application/json"}) or []
    return data if isinstance(data, list) else [data]


def load_local_containers(project_id: int) -> List[Dict]:
    """
    Загружает список контейнеров проекта.

    Args:
        project_id: ID проекта

    Returns:
        Список словарей с контейнерами

    Побочные эффекты:
        Выполняет GET запрос к API
    """
    url = INTERNAL_CONTAINERS_URL.format(project_id=project_id)
    data = api_get(f"{INTERNAL_BASE_URL}{url}", {"accept": "application/json"}) or []
    return data if isinstance(data, list) else [data]


def load_elements(container_id: int) -> Dict[str, int]:
    """
    Загружает все элементы контейнера и создаёт маппинг идентификаторов.

    Для каждого элемента создаёт несколько ключей в маппинге:
    - nativeId (оригинальный и в нижнем регистре)
    - enuid (оригинальный и в нижнем регистре)
    - uniqueId (оригинальный и в нижнем регистре)
    - nid (оригинальный и в нижнем регистре)
    - Если это GUID с суффиксом, то и базовая часть GUID

    Args:
        container_id: ID контейнера для загрузки элементов

    Returns:
        Словарь {идентификатор: элемент_id} для быстрого поиска

    Побочные эффекты:
        Выполняет GET запрос к API
    """
    url = INTERNAL_ELEMENTS_URL.format(container_id=container_id)
    elements = api_get(f"{INTERNAL_BASE_URL}{url}", {"accept": "application/json"}) or []
    if not isinstance(elements, list):
        return {}
    mapping: Dict[str, int] = {}
    
    # Цикл: перебираем все элементы контейнера
    for el in elements:
        # Получаем все возможные идентификаторы элемента
        native_id = el.get("nativeId") or el.get("NativeId")
        enuid = el.get("enuid") or el.get("Enuid")
        unique_id = el.get("uniqueId") or el.get("UniqueId")
        nid = el.get("nid") or el.get("Nid")
        elem_id = el.get("id") or el.get("Id")
        if elem_id is None:
            continue
        
        candidates = []
        # Цикл: собираем все варианты идентификаторов для поиска
        for value in (native_id, enuid, unique_id, nid):
            if value is None:
                continue
            val = str(value).strip()
            if not val:
                continue
            candidates.append(val)
            candidates.append(val.lower())
            # Если это GUID с суффиксом, добавляем базовую часть
            if _GUID_SUFFIX_RE.match(val):
                base = val.rsplit("-", 1)[0]
                candidates.append(base)
                candidates.append(base.lower())
        
        # Цикл: заполняем маппинг всеми вариантами ключей
        for key in candidates:
            mapping.setdefault(key, elem_id)
    return mapping


def element_id_from_enuid(element_map: Dict[str, int], enuid: str) -> Optional[int]:
    """
    Ищет элемент по идентификатору в element_map.
    Поддерживает форматы:
    - GUID (xxxxxxxx-xxxx-xxxx-xxxx-xxxxxxxx-xxxxx)
    - nativeId/uniqueId/enuid/nid
    - IFC формат: IfcElement/GUID, IfcWallStandardCase/GUID и т.д.
    """
    if not enuid:
        return None
    val = str(enuid).strip()
    if not val:
        return None
    
    # 1. Проверяем прямое совпадение
    for key in (val, val.lower()):
        if key in element_map:
            return element_map[key]
    
    # 2. Обработка IFC формата: IfcElement/GUID -> пробуем GUID
    if "/" in val:
        guid = val.split("/")[-1]
        for key in (guid, guid.lower()):
            if key in element_map:
                return element_map[key]
    
    # 3. Обработка GUID формата с суффиксом
    if _GUID_SUFFIX_RE.match(val):
        base = val.rsplit("-",1)[0]
        for key in (base, base.lower()):
            if key in element_map:
                return element_map[key]
    
    return None


def load_parameter_definitions(container_id: int) -> List[Dict]:
    """
    Загружает список определений параметров контейнера.

    Args:
        container_id: ID контейнера

    Returns:
        Список словарей с определениями параметров

    Побочные эффекты:
        Выполняет GET запрос к API
    """
    return api_get(
        f"{INTERNAL_BASE_URL}{INTERNAL_PARAM_DEFS_URL}",
        {"accept": "application/json"},
        {"containerIds": container_id},
    ) or []


def load_element_param_values(element_id: int) -> List[Dict]:
    """
    Загружает значения всех параметров для элемента.

    Args:
        element_id: ID элемента

    Returns:
        Список значений параметров элемента

    Побочные эффекты:
        Выполняет GET запрос к API
    """
    url = INTERNAL_PARAM_VALUES_GET_URL.format(element_id=element_id)
    data = api_get(f"{INTERNAL_BASE_URL}{url}", {"accept": "application/json"}) or []
    # Некоторые API возвращают данные в формате {"value": [...]}
    if isinstance(data, dict) and "value" in data:
        data = data.get("value") or []
    return data if isinstance(data, list) else []


def extract_param_value(item: Dict) -> str:
    """
    Извлекает значение параметра из объекта API.

    Выбирает значение в зависимости от флага onAlternative:
    - onAlternative=True: alternativeValueNumeric / alternativeValueString
    - onAlternative=False: originValueNumeric / originValueString

    Args:
        item: Словарь со значением параметра из API

    Returns:
        Строковое представление значения параметра

    Побочные эффекты:
        Нет
    """
    on_alt = bool(item.get("onAlternative"))
    is_numeric = bool(item.get("isNumeric"))
    if on_alt:
        value = item.get("alternativeValueNumeric") if is_numeric else item.get("alternativeValueString")
    else:
        value = item.get("originValueNumeric") if is_numeric else item.get("originValueString")
    if value is None:
        return ""
    return str(value).strip()


def ensure_parameter_definition(container_id: int, code: str, element_ids: List[int], value: str, existing_codes: set, is_numeric: bool, uom: str) -> Tuple[bool, str]:
    """
    Создает параметр если его нет в системе.
    Использует нормализованный code и корректный payload.
    """
    print("=" * 80)
    print(f"📋 СОЗДАНИЕ ПАРАМЕТРА")
    print("=" * 80)
    
    # 1. Валидация входных данных
    is_valid, error_msg, normalized_value = _validate_parameter_payload(code, is_numeric, value, element_ids)
    if not is_valid:
        print(f"❌ Ошибка валидации: {error_msg}")
        print(f"   Code: {repr(code)}")
        print(f"   isNumeric: {is_numeric}")
        print(f"   Value: {repr(value)}")
        print(f"   Element IDs: {element_ids}")
        print("=" * 80)
        return False, error_msg
    
    # 2. Проверяем существование параметра в кэше
    if code in existing_codes:
        print(f"📌 Параметр '{code}' уже существует в existing_codes")
        print("=" * 80)
        return True, ""
    
    # 3. Загружаем параметры из API
    params = api_get(
        f"{INTERNAL_BASE_URL}/imcParameterDefinition/imcParameterDefinitions",
        {"accept": "application/json"},
        {"containerIds": container_id}
    ) or []
    
    print(f"🔢 Всего параметров в контейнере {container_id}: {len(params)}")
    
    # 4. Находим реальный код параметра в системе (учитывая ведущий слеш)
    real_code = _find_parameter_code_in_system(code, params)
    
    if real_code in existing_codes:
        print(f"📌 Параметр '{real_code}' уже существует в existing_codes (реальный код)")
        print("=" * 80)
        return True, ""
    
    if any(p.get("code") == real_code for p in params):
        print(f"📌 Параметр '{real_code}' уже существует в API (пропускаем создание)")
        print(f"   Исходный код: {repr(code)}")
        print(f"   Реальный код: {repr(real_code)}")
        print("=" * 80)
        existing_codes.add(real_code)
        return True, ""
    
    # 5. Формируем payload - КОРРЕКТНО
    # ИСПОЛЬЗУЕМ real_code для создания, чтобы совпадало с существующими параметрами
    create_data = {
        "code": real_code,
        "isNumeric": is_numeric,
        "layer": PARAMETER_LAYER,
        "reportColumnType": PARAMETER_REPORT_COLUMN_TYPE,
        "title": "Создан через Excel",
        "elementMaps": [
            {
                "containerId": container_id,
                "elementIds": element_ids
            }
        ]
    }
    
    # Добавляем только нужное поле значения
    if is_numeric:
        create_data["numericValue"] = float(normalized_value)
    else:
        create_data["stringValue"] = normalized_value
    
    # НЕ добавляем uom если пустой
    if uom and uom.strip():
        create_data["uom"] = uom.strip()
    
    # 6. Диагностический лог
    print(f"   Code (из Excel): {repr(code)}")
    print(f"   Real code (для создания): {repr(real_code)}")
    print(f"   Коды совпадают? {code == real_code}")
    print(f"   isNumeric: {is_numeric}")
    print(f"   Value: {repr(normalized_value)}")
    print(f"   Container ID: {container_id}")
    print(f"   Element IDs: {element_ids}")
    print(f"   Payload: {json.dumps(create_data, ensure_ascii=False, indent=2)}")
    
    # 7. Отправляем запрос
    try:
        create_resp = requests.post(
            f"{INTERNAL_BASE_URL}{INTERNAL_PARAM_DEF_CREATE_URL}",
            headers={"Content-Type": "application/json"},
            data=json.dumps(create_data, ensure_ascii=False)
        )
        
        print(f"   Status: {create_resp.status_code}")
        
        if create_resp.status_code in (200, 201):
            print(f"✅ Параметр '{real_code}' создан успешно")
            # Добавляем В ОБА кода (нормализованный и реальный)
            existing_codes.add(real_code)
            existing_codes.add(code)
            print(f"   Добавлен в existing_codes: {repr(real_code)}")
            print("=" * 80)
            return True, ""
        else:
            print(f"❌ Ошибка создания параметра '{real_code}':")
            print(f"   Status: {create_resp.status_code}")
            print(f"   Response: {create_resp.text}")
            print("=" * 80)
            return False, f"Status {create_resp.status_code}: {create_resp.text}"
    except Exception as exc:
        print(f"❌ Исключение при создании параметра '{real_code}': {exc}")
        print("=" * 80)
        return False, str(exc)


def update_parameter_values(
    container_id: int,
    code: str,
    element_ids: List[int],
    value: str,
    is_numeric: bool
) -> Tuple[bool, str]:
    """
    Обновляет значение параметра для указанных элементов.
    Использует setAlternateValueByElements.
    """
    print("=" * 80)
    print(f"🔄 ОБНОВЛЕНИЕ ЗНАЧЕНИЯ ПАРАМЕТРА")
    print("=" * 80)
    
    # 1. Валидация
    is_valid, error_msg, normalized_value = _validate_parameter_payload(code, is_numeric, value, element_ids)
    if not is_valid:
        print(f"❌ Ошибка валидации: {error_msg}")
        print(f"   Code: {repr(code)}")
        print(f"   isNumeric: {is_numeric}")
        print(f"   Value: {repr(value)}")
        print(f"   Element IDs: {element_ids}")
        print("=" * 80)
        return False, error_msg
    
    # 2. Проверяем что параметр существует
    params = api_get(
        f"{INTERNAL_BASE_URL}/imcParameterDefinition/imcParameterDefinitions",
        {"accept": "application/json"},
        {"containerIds": container_id}
    ) or []
    
    print(f"📊 Всего параметров в контейнере {container_id}: {len(params)}")
    print(f"🔍 Ищем код: {repr(code)}")
    
    # Показываем первые 5 параметров для диагностики
    if params and len(params) <= 10:
        print(f"📋 Все параметры:")
        for p in params:
            print(f"   - code={repr(p.get('code'))}, isNumeric={p.get('isNumeric')}")
    elif params:
        print(f"📋 Первые 5 параметров:")
        for p in params[:5]:
            print(f"   - code={repr(p.get('code'))}, isNumeric={p.get('isNumeric')}")
    
    real_code = _find_parameter_code_in_system(code, params)
    
    print(f"🎯 Найден реальный код: {repr(real_code)}")
    print(f"   Совпадает с искомым? {real_code == code}")
    
    if not any(p.get("code") == real_code for p in params):
        print(f"❌ Параметр '{real_code}' не существует в контейнере {container_id}")
        print(f"   Исходный код: {repr(code)}")
        print(f"   Реальный код: {repr(real_code)}")
        print("=" * 80)
        return False, f"Parameter '{real_code}' not found"
    
    # 3. Формируем payload
    update_data = {
        "containerIds": [
            {
                "containerId": container_id,
                "elementIds": element_ids
            }
        ],
        "parameterCode": real_code,  # Используем реальный код
        "isNumeric": is_numeric
    }
    
    # Добавляем только нужное поле значения
    if is_numeric:
        update_data["numericValue"] = float(normalized_value)
    else:
        update_data["stringValue"] = normalized_value
    
    # 4. Диагностический лог
    print(f"   Code: {repr(code)}")
    print(f"   Real code (from API): {repr(real_code)}")
    print(f"   isNumeric: {is_numeric}")
    print(f"   Value: {repr(normalized_value)}")
    print(f"   Container ID: {container_id}")
    print(f"   Element IDs: {element_ids}")
    print(f"   Payload: {json.dumps(update_data, ensure_ascii=False, indent=2)}")
    
    # 5. Отправляем запрос
    try:
        update_resp = requests.post(
            f"{INTERNAL_BASE_URL}{INTERNAL_PARAM_VALUES_URL}",
            headers=INTERNAL_HEADERS,
            data=json.dumps(update_data, ensure_ascii=False)
        )
        
        print(f"   Status: {update_resp.status_code}")
        
        if update_resp.status_code in (200, 201):
            print(f"✅ Значение параметра '{real_code}' обновлено")
            print("=" * 80)
            return True, ""
        else:
            print(f"❌ Ошибка обновления параметра '{real_code}':")
            print(f"   Status: {update_resp.status_code}")
            print(f"   Response: {update_resp.text}")
            print("=" * 80)
            return False, f"Status {update_resp.status_code}: {update_resp.text}"
    except Exception as exc:
        print(f"❌ Исключение при обновлении параметра '{real_code}': {exc}")
        print("=" * 80)
        return False, str(exc)


def _first_element_id(container_id: int) -> Optional[int]:
    """
    Находит ID первого элемента в контейнере.

    Используется для создания параметра без привязки к элементам.

    Args:
        container_id: ID контейнера

    Returns:
        ID первого элемента или None если контейнер пустой

    Побочные эффекты:
        Выполняет GET запрос к API
    """
    url = INTERNAL_ELEMENTS_URL.format(container_id=container_id)
    elements = api_get(f"{INTERNAL_BASE_URL}{url}", {"accept": "application/json"}) or []
    if not isinstance(elements, list):
        return None
    
    # Цикл: ищем первый элемент с валидным ID
    for item in elements:
        elem_id = item.get("id") or item.get("Id")
        if elem_id is not None:
            try:
                return int(elem_id)
            except Exception:
                return elem_id
    return None


def create_parameter_definition_from_table(
    container_id: int,
    param_def: ParameterDef,
    element_id: Optional[int],
) -> Tuple[bool, str]:
    """
    Создаёт определение параметра из Excel таблицы.

    Сначала пытается создать параметр без elementMaps.
    Если не удаётся и element_id передан - пробует с elementMaps.

    Args:
        container_id: ID контейнера
        param_def: Определение параметра
        element_id: ID элемента для привязки (опционально)

    Returns:
        Кортеж (success, error_message)

    Побочные эффекты:
        Отправляет POST запрос к API
    """
    create_data = {
        "code": param_def.code,
        "isNumeric": param_def.is_numeric,
        "layer": PARAMETER_LAYER,
        "reportColumnType": PARAMETER_REPORT_COLUMN_TYPE,
        "title": param_def.title,
        "uom": param_def.uom,
    }
    
    # Первая попытка: создаём параметр без elementMaps
    ok, err = api_post_json(
        f"{INTERNAL_BASE_URL}{INTERNAL_PARAM_DEF_CREATE_URL}",
        {"Content-Type": "application/json"},
        create_data,
    )
    if ok:
        return True, ""
    
    # Если неудача и есть element_id - пробуем с elementMaps
    if element_id is None:
        return False, err
    create_data["elementMaps"] = [{"containerId": container_id, "elementIds": [element_id]}]
    return api_post_json(
        f"{INTERNAL_BASE_URL}{INTERNAL_PARAM_DEF_CREATE_URL}",
        {"Content-Type": "application/json"},
        create_data,
    )


def format_report(report: ReportData) -> str:
    """
    Форматирует отчёт в виде текстовой таблицы.

    Args:
        report: Данные отчёта

    Returns:
        Отформатированная строка отчёта

    Побочные эффекты:
        Нет
    """
    lines = [
        "Отчет:",
        f"Обновлено: {len(report.updated)}",
        f"Пропущено (без изменений): {len(report.skipped_same)}",
        f"Не найдено (элемент): {len(report.not_found)}",
        f"Ошибка обновления: {len(report.failed_update)}",
    ]

    def add_block(title: str, items: Iterable[ReportItem]):
        """
        Добавляет блок с таблицей элементов в отчёт.
        """
        items = list(items)
        if not items:
            return
        lines.append("")
        lines.append(f"{title}:")
        headers = ["Модель", "Идентификатор", "Параметр", "Старое значение", "Новое значение"]
        rows = [
            [it.model, it.identifier, it.code, it.old_value, it.new_value]
            for it in items
        ]
        
        # Цикл: вычисляем ширину каждой колонки для красивой таблицы
        col_widths = []
        for col_idx, header in enumerate(headers):
            max_len = max(len(header), max(len(str(row[col_idx])) for row in rows))
            col_widths.append(max_len)
        
        header_row = " | ".join(header.ljust(col_widths[i]) for i, header in enumerate(headers))
        sep_row = "-+-".join("-" * col_widths[i] for i in range(len(headers)))
        lines.append(header_row)
        lines.append(sep_row)
        
        # Цикл: выводим строки таблицы
        for row in rows:
            line = " | ".join(str(row[i]).ljust(col_widths[i]) for i in range(len(headers)))
            lines.append(line)

    add_block("Обновлено", report.updated)
    add_block("Пропущено", report.skipped_same)
    add_block("Не найдено (элемент)", report.not_found)
    add_block("Ошибка обновления", report.failed_update)

    return "\n".join(lines)


def read_excel_changes(excel_path: str) -> List[Change]:
    """
    Читает изменения из Excel файла (только изменения, без определений параметров).

    Args:
        excel_path: Путь к Excel файлу

    Returns:
        Список изменений параметров

    Побочные эффекты:
        Читает файл с диска
    """
    changes, param_defs = read_excel_payload(excel_path)
    return changes


def detect_file_format(file_path: str) -> str:
    """
    Определяет формат файла по расширению.

    Args:
        file_path: Путь к файлу

    Returns:
        "excel" или "xml"

    Raises:
        ValueError: Если формат файла не поддерживается

    Побочные эффекты:
        Нет
    """
    ext = os.path.splitext(file_path)[1].lower()
    if ext in (".xlsx", ".xls"):
        return "excel"
    elif ext == ".xml":
        return "xml"
    else:
        raise ValueError(f"Неподдерживаемый формат файла: {ext}. Используйте XML или Excel (.xlsx, .xls)")


def run_upload(
    file_path: str,
    containers: List[Tuple[int, str]],
) -> Tuple[int, str, ReportData]:
    """
    Основная функция загрузки изменений из файла в контейнеры.

    Поддерживает два режима:
    1. Excel с определениями параметров (param_defs не пустые) - создаёт параметры
    2. XML или Excel с изменениями (changes не пустые) - обновляет значения параметров

    Args:
        file_path: Путь к файлу (XML или Excel)
        containers: Список контейнеров [(container_id, container_title), ...]

    Returns:
        Кортеж (exit_code, report_text, report_data):
        - exit_code: 0 при успехе, 1 при ошибке
        - report_text: Текстовый отчёт
        - report_data: Структурированные данные отчёта

    Побочные эффекты:
        Выполняет HTTP запросы к API для создания параметров и обновления значений
    """
    file_format = detect_file_format(file_path)
    is_excel = file_format == "excel"
    
    try:
        # Читаем данные из файла
        if is_excel:
            changes, param_defs = read_excel_payload(file_path)
        else:
            changes = read_xml_changes(file_path)
            param_defs = []
    except Exception as e:
        empty_report = ReportData([], [], [], [])
        return 1, str(e), empty_report

    # Режим 1: Создание определений параметров из Excel
    if is_excel and param_defs:
        updated: List[ReportItem] = []
        skipped_same: List[ReportItem] = []
        not_found: List[ReportItem] = []
        failed_update: List[ReportItem] = []
        first_element_cache: Dict[int, Optional[int]] = {}

        # Цикл: перебираем все выбранные контейнеры
        for container_id, container_title in containers:
            param_defs_existing = load_parameter_definitions(container_id)
            existing_codes = {p.get("code") for p in param_defs_existing if p.get("code")}

            # Кэшируем первый элемент контейнера для создания параметров
            if container_id not in first_element_cache:
                first_element_cache[container_id] = _first_element_id(container_id)
            element_id = first_element_cache[container_id]

            # Цикл: создаём каждое определение параметра из Excel
            for param_def in param_defs:
                if param_def.code in existing_codes:
                    skipped_same.append(ReportItem(container_title, "-", param_def.code, "", "exists"))
                    continue
                ok, err = create_parameter_definition_from_table(container_id, param_def, element_id)
                if ok:
                    updated.append(ReportItem(container_title, "-", param_def.code, "", param_def.title))
                    existing_codes.add(param_def.code)
                else:
                    failed_update.append(ReportItem(container_title, "-", param_def.code, "", err or "create failed"))

        report = ReportData(updated, skipped_same, not_found, failed_update)
        return 0, format_report(report), report

    # Режим 2: Обновление значений параметров (из XML или Excel)
    if not changes:
        empty_report = ReportData([], [], [], [])
        source = "Excel" if is_excel else "XML"
        return 0, f"Нет изменений в файле {source}.", empty_report

    updated: List[ReportItem] = []
    skipped_same: List[ReportItem] = []
    not_found: List[ReportItem] = []
    failed_update: List[ReportItem] = []

    def fmt_entry(
        container_title: str,
        ch: Change,
        new_value: Optional[str] = None,
        old_value: Optional[str] = None,
    ) -> ReportItem:
        """
        Создаёт элемент отчёта из изменения.
        """
        return ReportItem(
            model=container_title,
            identifier=ch.identifier,
            code=ch.code,
            old_value=ch.old_value if old_value is None else old_value,
            new_value=ch.new_value if new_value is None else new_value,
        )

    pending_changes = []
    # Цикл: разделяем изменения на ожидающие обработки и пропущенные (без изменений)
    for ch in changes:
        if ch.old_value == ch.new_value:
            skipped_same.append(fmt_entry("", ch))
        else:
            pending_changes.append(ch)

    # Цикл: перебираем все контейнеры для обработки изменений
    for container_id, container_title in containers:
        # Загружаем element_map для обоих случаев (Excel и XML)
        # Для Excel: ID может быть nativeId/enuid/GUID вида "IfcElement/8b..."
        print("=" * 80)
        print(f"📦 Обработка контейнера: {container_title} (ID: {container_id})")
        print("=" * 80)
        element_map = load_elements(container_id)
        if not element_map:
            print(f"⚠️ Не удалось загрузить элементы для контейнера {container_id} ({container_title})")
            print("=" * 80)
            continue
        print(f"📊 Загружено элементов: {len(element_map)}")
        print(f"📋 Пример идентификаторов: {list(element_map.keys())[:5]}")
        print("=" * 80)

        per_container_changes: List[Tuple[Change, int]] = []
        element_values_cache: Dict[int, Dict[str, str]] = {}
        
        # Цикл: связываем изменения с элементами контейнера
        for ch in pending_changes:
            # Для обоих случаев (Excel и XML) ищем элемент по identifier
            # Identifier может быть: nativeId, enuid, uniqueId, nid, GUID вида "IfcElement/8b..."
            element_id = element_id_from_enuid(element_map, ch.identifier)
            if element_id is None:
                print(f"⚠️ Элемент НЕ НАЙДЕН:")
                print(f"   identifier из Excel: {repr(ch.identifier)}")
                print(f"   Поиск по: direct, lowercase, IFC format (IfcElement/GUID), GUID base")
                print(f"   Всего в element_map: {len(element_map)} ключей")
                not_found.append(fmt_entry(container_title, ch))
                continue
            print(f"✅ Элемент найден: identifier={repr(ch.identifier)} -> element_id={element_id}")
            
            # Кэшируем значения параметров элемента
            if element_id not in element_values_cache:
                params = load_element_param_values(element_id)
                cache = {}
                # Цикл: кэшируем все значения параметров с вариантами слешей
                for p in params:
                    param_code_raw = str(p.get("code", "")).strip()
                    if not param_code_raw:
                        continue
                    param_value = extract_param_value(p)
                    cache[param_code_raw] = param_value
                    # Добавляем вариант без ведущего слеша
                    if param_code_raw.startswith("\\"):
                        cache.setdefault(param_code_raw[1:], param_value)
                    # Добавляем вариант с ведущим слешем
                    else:
                        cache.setdefault("\\" + param_code_raw, param_value)
                element_values_cache[element_id] = cache
            
            current_value = element_values_cache[element_id].get(ch.code)
            if current_value is not None and current_value == ch.new_value:
                skipped_same.append(fmt_entry(container_title, ch, current_value, current_value))
                continue
            per_container_changes.append((ch, element_id))

        if not per_container_changes:
            continue

        # Загружаем определения параметров контейнера
        param_defs = load_parameter_definitions(container_id)
        existing_codes = {p.get("code") for p in param_defs if p.get("code")}

        # Группируем изменения по коду параметра
        code_to_element_ids: Dict[str, set] = {}
        for ch, element_id in per_container_changes:
            code_to_element_ids.setdefault(ch.code, set()).add(element_id)

        # Создаём параметры которые ещё не существуют
        create_errors: Dict[str, str] = {}
        for code, element_ids in code_to_element_ids.items():
            value_sample = next((c.new_value for c, _ in per_container_changes if c.code == code), "")
            is_numeric = next((c.is_numeric for c, _ in per_container_changes if c.code == code), False)
            uom = next((c.uom for c, _ in per_container_changes if c.code == code), "")
            ok, err = ensure_parameter_definition(
                container_id,
                code,
                sorted(element_ids),
                value_sample,
                existing_codes,
                is_numeric,
                uom,
            )
            if not ok:
                create_errors[code] = err or "create failed"

        # Группируем изменения по коду параметра, значению и типу
        grouped: Dict[Tuple[str, str, bool], List[Tuple[Change, int]]] = {}
        for ch, element_id in per_container_changes:
            grouped.setdefault((ch.code, ch.new_value, ch.is_numeric), []).append((ch, element_id))

        # Цикл: обрабатываем каждую группу изменений
        for (code, value, is_numeric), items in grouped.items():
            if code in create_errors:
                for ch, _ in items:
                    failed_update.append(fmt_entry(container_title, ch, f"{ch.new_value} (ошибка: {create_errors[code]})"))
                continue
            element_ids = sorted({eid for _, eid in items})
            
            # Сначала создаем/проверяем параметр
            ok, err = ensure_parameter_definition(
                container_id,
                code,
                element_ids,
                value,
                existing_codes,
                is_numeric,
                "",
            )
            if ok:
                # Получаем real_code для обновления
                params = api_get(
                    f"{INTERNAL_BASE_URL}/imcParameterDefinition/imcParameterDefinitions",
                    {"accept": "application/json"},
                    {"containerIds": container_id}
                ) or []
                real_code = _find_parameter_code_in_system(code, params)
                
                print(f"🔄 Обновление параметра: code={repr(code)}, real_code={repr(real_code)}")
                print(f"   Коды совпадают? {code == real_code}")
                
                # Параметр создан или существует, теперь обновляем значение
                # Используем real_code для обновления!
                update_ok, update_err = update_parameter_values(container_id, real_code, element_ids, value, is_numeric)
                if update_ok:
                    for ch, _ in items:
                        updated.append(fmt_entry(container_title, ch))
                else:
                    err_text = update_err or "update failed"
                    for ch, _ in items:
                        failed_update.append(fmt_entry(container_title, ch, f"{ch.new_value} (ошибка: {err_text})"))
            else:
                for ch, _ in items:
                    failed_update.append(fmt_entry(container_title, ch, f"{ch.new_value} (ошибка: {err})"))

    report = ReportData(updated, skipped_same, not_found, failed_update)
    return 0, format_report(report), report


if PYSIDE:
    class OutputWindow(QMainWindow):
        def __init__(self, parent=None):
            super().__init__(parent)
            self.setWindowTitle("XML Uploader - Report")
            self.setMinimumWidth(720)
            try:
                if os.path.exists(LOGO_ICO_PATH):
                    self.setWindowIcon(QIcon(LOGO_ICO_PATH))
            except Exception:
                pass

            central = QWidget()
            layout = QVBoxLayout(central)
            self.tabs = QTabWidget()
            self.tables = {
                "Обновлено": self._create_table(),
                "Пропущено": self._create_table(),
                "Не найдено": self._create_table(),
                "Ошибка обновления": self._create_table(),
            }
            self.tab_order = ["Обновлено", "Пропущено", "Не найдено", "Ошибка обновления"]
            for title in self.tab_order:
                self.tabs.addTab(self.tables[title], title)
            layout.addWidget(self.tabs, stretch=1)
            self.setCentralWidget(central)
            self.apply_theme()
            QtCore.QTimer.singleShot(
                0,
                lambda: _set_dark_title_bar(
                    self,
                    _is_dark_theme(QtWidgets.QApplication.instance()),
                ),
            )

        def _create_table(self) -> QTableWidget:
            table = QTableWidget()
            headers = ["Модель", "Идентификатор", "Параметр", "Старое значение", "Новое значение"]
            table.setColumnCount(len(headers))
            table.setHorizontalHeaderLabels(headers)
            table.setEditTriggers(QAbstractItemView.NoEditTriggers)
            table.setSelectionBehavior(QAbstractItemView.SelectRows)
            table.setSelectionMode(QAbstractItemView.SingleSelection)
            table.setAlternatingRowColors(True)
            header = table.horizontalHeader()
            header.setSectionResizeMode(QHeaderView.Stretch)
            header.setStretchLastSection(False)
            table.verticalHeader().setVisible(False)
            return table

        def _fill_table(self, table: QTableWidget, items: List[ReportItem]) -> None:
            table.setRowCount(len(items))
            for row, item in enumerate(items):
                values = [item.model, item.identifier, item.code, item.old_value, item.new_value]
                for col, value in enumerate(values):
                    text = str(value)
                    cell = QTableWidgetItem(text)
                    cell.setToolTip(text)
                    table.setItem(row, col, cell)
            table.resizeRowsToContents()

        def set_report(self, report: ReportData, text: str | None = None):
            if report is None:
                return
            self._fill_table(self.tables["Обновлено"], report.updated)
            self._fill_table(self.tables["Пропущено"], report.skipped_same)
            self._fill_table(self.tables["Не найдено"], report.not_found)
            self._fill_table(self.tables["Ошибка обновления"], report.failed_update)
            counts = [
                len(report.updated),
                len(report.skipped_same),
                len(report.not_found),
                len(report.failed_update),
            ]
            for idx, title in enumerate(self.tab_order):
                self.tabs.setTabText(idx, f"{title} ({counts[idx]})")
            self.tabs.setCurrentIndex(0)

        def apply_theme(self) -> None:
            dark = _is_dark_theme(QtWidgets.QApplication.instance())
            if dark:
                table_style = (
                    "QTableView {"
                    "background-color: #1b1b1b;"
                    "color: #f0f0f0;"
                    "alternate-background-color: #2a2a2a;"
                    "gridline-color: #3a3a3a;"
                    "}"
                    "QHeaderView::section {"
                    "background-color: #1f1f1f;"
                    "color: #f0f0f0;"
                    "border: 1px solid #3a3a3a;"
                    "}"
                )
                tab_style = (
                    "QTabWidget::pane { border: none; }"
                    "QTabBar::tab {"
                    "background: #1b1b1b;"
                    "color: #f0f0f0;"
                    "border: 1px solid #3a3a3a;"
                    "border-radius: 12px;"
                    "padding: 6px 12px;"
                    "margin: 0 4px;"
                    "}"
                    "QTabBar::tab:hover {"
                    "background: rgba(247, 146, 30, 0.15);"
                    "border: 1px solid #F7921E;"
                    "}"
                    "QTabBar::tab:selected {"
                    "background: rgba(247, 146, 30, 0.25);"
                    "border: 1px solid #F7921E;"
                    "}"
                )
            else:
                table_style = ""
                tab_style = ""
            for table in self.tables.values():
                table.setStyleSheet(table_style)
            self.tabs.setStyleSheet(tab_style)

    class MainWindow(QMainWindow):
        def __init__(self):
            super().__init__()
            self.setWindowTitle("XML Uploader")
            self.setMinimumWidth(780)
            try:
                if os.path.exists(LOGO_ICO_PATH):
                    self.setWindowIcon(QIcon(LOGO_ICO_PATH))
            except Exception:
                pass

            central = QWidget()
            layout = QVBoxLayout(central)

            header = QWidget()
            header.setObjectName("header")
            hl = QHBoxLayout(header)
            hl.setContentsMargins(8, 8, 8, 8)
            logo_label = QLabel()
            if set_header_logo is not None:
                try:
                    set_header_logo(logo_label, ICON_DIR, height=56)
                except Exception:
                    pass
            if logo_label.pixmap() is None:
                logo_path = ""
                for fname in ("Manager-scaled.png", "logo.png", "logo_white.png"):
                    candidate = os.path.join(ICON_DIR, fname)
                    if os.path.exists(candidate):
                        logo_path = candidate
                        break
                if logo_path:
                    pm = QPixmap(logo_path)
                    if not pm.isNull():
                        logo_label.setPixmap(pm.scaledToHeight(56, QtCore.Qt.SmoothTransformation))
            logo_label.setAlignment(QtCore.Qt.AlignCenter)
            hl.addWidget(logo_label, 1)
            if ThemeSwitch is not None:
                try:
                    self.theme_switch = ThemeSwitch(icon_dir=ICON_DIR)
                    if set_header_logo is not None:
                        self.theme_switch.toggledTheme.connect(lambda _=None: set_header_logo(logo_label, ICON_DIR, height=56))
                    hl.addWidget(self.theme_switch)
                except Exception:
                    self.theme_switch = None
            else:
                self.theme_switch = None
            layout.addWidget(header)

            server_group = QGroupBox("Сервер")
            server_layout = QHBoxLayout(server_group)
            server_layout.addWidget(QLabel("localhost:"))
            self.base_url_edit = QLineEdit()
            self.base_url_edit.setPlaceholderText("5000")
            self.base_url_edit.setText("5000")
            self.base_url_edit.editingFinished.connect(self._on_server_edit_finished)
            server_layout.addWidget(self.base_url_edit, stretch=1)
            self.connect_btn = QPushButton("Подключить")
            self._connect_icon_path = os.path.join(ICON_DIR, "free-icon-login-2623062.png")
            if os.path.exists(self._connect_icon_path):
                self.connect_btn.setIcon(QIcon(self._connect_icon_path))
            self.connect_btn.clicked.connect(self._on_connect_clicked)
            server_layout.addWidget(self.connect_btn)
            self._apply_connect_icon()
            self.conn_status_icon = QLabel()
            self.conn_status_icon.setFixedSize(16, 16)
            self.conn_status_icon.setVisible(False)
            server_layout.addWidget(self.conn_status_icon)
            layout.addWidget(server_group)

            export_btn_layout = QHBoxLayout()
            export_btn_layout.addStretch(1)
            self.export_params_btn = QPushButton("Экспорт параметров")
            self.export_params_btn.clicked.connect(self._on_export_params_clicked)
            export_btn_layout.addWidget(self.export_params_btn)
            export_btn_layout.addStretch(1)
            layout.addLayout(export_btn_layout)

            file_group = QGroupBox("Файл")
            file_layout = QHBoxLayout(file_group)
            self.file_edit = QLineEdit()
            self.file_edit.setPlaceholderText("Путь к файлу (XML или Excel)")
            file_layout.addWidget(self.file_edit, stretch=1)
            self.file_btn = QPushButton("Выбрать файл")
            self.file_btn.clicked.connect(self.pick_file)
            file_layout.addWidget(self.file_btn)
            self.template_btn = QPushButton("Шаблон")
            self.template_btn.clicked.connect(self.download_template)
            file_layout.addWidget(self.template_btn)
            layout.addWidget(file_group)

            local_group = QGroupBox("Локальные контейнеры")
            local_layout = QVBoxLayout(local_group)

            proj_row = QHBoxLayout()
            proj_row.addWidget(QLabel("Проект:"))
            self.project_combo = QComboBox()
            self.project_combo.currentIndexChanged.connect(self.load_local_containers)
            proj_row.addWidget(self.project_combo, stretch=1)
            self.load_local_btn = QPushButton("Загрузить локальные")
            self.load_local_btn.clicked.connect(self.load_local_projects)
            proj_row.addWidget(self.load_local_btn)
            local_layout.addLayout(proj_row)

            self.container_scroll = QScrollArea()
            self.container_scroll.setWidgetResizable(True)
            try:
                self.container_scroll.setFrameShape(QFrame.NoFrame)
            except Exception:
                pass
            self.container_host = QWidget()
            self.container_layout = QVBoxLayout(self.container_host)
            self.container_layout.addStretch(1)
            self.container_scroll.setWidget(self.container_host)
            local_layout.addWidget(self.container_scroll)

            self.checked_containers = {}
            self.container_titles = {}

            layout.addWidget(local_group)

            self.run_btn = QPushButton("Загрузить изменения")
            self.run_btn.clicked.connect(self.run_upload_gui)
            layout.addWidget(self.run_btn)

            self.output_window = OutputWindow(self)

            self.setCentralWidget(central)
            QtCore.QTimer.singleShot(
                0,
                lambda: _set_dark_title_bar(
                    self,
                    _is_dark_theme(QtWidgets.QApplication.instance()),
                ),
            )
            QtCore.QTimer.singleShot(
                0,
                lambda: self._try_connect(show_success=False, show_error=True),
            )
            if self.theme_switch is not None:
                self.theme_switch.toggledTheme.connect(
                    lambda _=None: _set_dark_title_bar(
                        self,
                        _is_dark_theme(QtWidgets.QApplication.instance()),
                    )
                )
                self.theme_switch.toggledTheme.connect(
                    lambda _=None: _set_dark_title_bar(
                        self.output_window,
                        _is_dark_theme(QtWidgets.QApplication.instance()),
                    )
                )
                self.theme_switch.toggledTheme.connect(
                    lambda _=None: self.output_window.apply_theme()
                )
                self.theme_switch.toggledTheme.connect(
                    lambda _=None: self._apply_connect_icon()
                )

        def _apply_connect_icon(self):
            if not self._connect_icon_path:
                return
            if not os.path.exists(self._connect_icon_path):
                return
            pm = QPixmap(self._connect_icon_path)
            if pm.isNull():
                return
            if _is_dark_theme(QtWidgets.QApplication.instance()):
                tinted = QPixmap(pm.size())
                tinted.fill(QtCore.Qt.transparent)
                painter = QPainter(tinted)
                painter.setCompositionMode(QPainter.CompositionMode_Source)
                painter.drawPixmap(0, 0, pm)
                painter.setCompositionMode(QPainter.CompositionMode_SourceIn)
                painter.fillRect(tinted.rect(), QColor(255, 255, 255))
                painter.end()
                pm = tinted
            self.connect_btn.setIcon(QIcon(pm))

        def _apply_base_url(self):
            _set_internal_base_url(self.base_url_edit.text())

        def _set_connection_status(self, ok: bool):
            if not ok:
                self.conn_status_icon.setPixmap(QPixmap())
                self.conn_status_icon.setVisible(False)
                return
            icon_path = os.path.join(ICON_DIR, "ok.png")
            pm = QPixmap(icon_path) if os.path.exists(icon_path) else QPixmap()
            if not pm.isNull():
                pm = pm.scaled(16, 16, QtCore.Qt.KeepAspectRatio, QtCore.Qt.SmoothTransformation)
            self.conn_status_icon.setPixmap(pm)
            self.conn_status_icon.setVisible(not pm.isNull())

        def _try_connect(self, *, show_success: bool, show_error: bool):
            self._apply_base_url()
            projects = load_local_projects()
            if not projects:
                self._set_connection_status(False)
                if show_error:
                    QMessageBox.warning(self, "Ошибка", f"Не удалось подключиться к {INTERNAL_BASE_URL}")
                return
            self._set_connection_status(True)
            if show_success:
                QMessageBox.information(self, "Подключено", f"Успешное подключение: {INTERNAL_BASE_URL}")
            self.project_combo.clear()
            
            def get_project_created_date(project):
                for key in ("created", "createdAt", "creationDate", "createDate"):
                    if key in project:
                        return project.get(key)
                return None
            
            sorted_projects = sorted(projects, key=lambda p: get_project_created_date(p) or "", reverse=False)
            
            for p in sorted_projects:
                self.project_combo.addItem(p.get("title", "Без названия"), p.get("id"))
            if sorted_projects:
                self.load_local_containers()

        def _on_server_edit_finished(self):
            self._try_connect(show_success=False, show_error=True)

        def _on_connect_clicked(self):
            self._try_connect(show_success=True, show_error=True)

        def pick_file(self):
            path, _ = QFileDialog.getOpenFileName(self, "Выбрать файл", "", "Файлы (*.xml *.xlsx *.xls)")
            if path:
                self.file_edit.setText(path)
        
        def download_template(self):
            """Создает шаблон Excel для импорта параметров"""
            try:
                from ExcelTemplate import create_excel_template
            except ImportError:
                QMessageBox.critical(self, "Ошибка", "Модуль Excel не найден")
                return
            
            # Открываем диалог сохранения с названием по умолчанию
            save_path, _ = QFileDialog.getSaveFileName(
                self,
                "Сохранить шаблон",
                "Шаблон Excel.xlsx",
                "Файлы Excel (*.xlsx)"
            )
            
            if not save_path:
                return
            
            # Создаем шаблон через функцию из Excel.py
            try:
                wb = create_excel_template()
                wb.save(save_path)
                QMessageBox.information(self, "Успех", f"Шаблон сохранен:\n{save_path}")
            except Exception as exc:
                QMessageBox.critical(self, "Ошибка", f"Не удалось создать шаблон:\n{exc}")

        def _on_export_params_clicked(self):
            """Открывает диалог экспорта параметров"""
            try:
                from ExportParameter import ExportParamsDialog
            except ImportError:
                QMessageBox.critical(self, "Ошибка", "Модуль ExportParameter не найден")
                return
            
            base_url = f"http://localhost:{self.base_url_edit.text().strip()}"
            dialog = ExportParamsDialog(base_url, self)
            dialog.exec_()

        def load_local_projects(self):
            self._try_connect(show_success=False, show_error=True)


        def load_local_containers(self):
            for i in reversed(range(self.container_layout.count() - 1)):
                item = self.container_layout.itemAt(i)
                widget = item.widget()
                if widget:
                    widget.setParent(None)
            self.checked_containers.clear()
            self.container_titles.clear()

            proj_id = self.project_combo.currentData()
            if proj_id is None:
                return

            containers = load_local_containers(proj_id)
            for c in containers:
                cid = c.get("id")
                title = c.get("title", "Без названия")
                if cid is None:
                    continue
                cb = QCheckBox(title)
                self.container_layout.insertWidget(self.container_layout.count() - 1, cb)
                self.checked_containers[cid] = cb
                self.container_titles[cid] = title

        def run_upload_gui(self):
            self._apply_base_url()
            file_path = self.file_edit.text().strip()
            if not file_path or not os.path.exists(file_path):
                QMessageBox.critical(self, "Ошибка", "Укажите существующий файл (XML или Excel).")
                return
            selected = [(cid, self.container_titles.get(cid, "")) for cid, cb in self.checked_containers.items() if cb.isChecked()]
            if not selected:
                QMessageBox.critical(self, "Ошибка", "Выберите хотя бы один контейнер.")
                return

            code, report_text, report_data = run_upload(file_path, selected)
            self.output_window.set_report(report_data, report_text)
            self.output_window.show()
            self.output_window.raise_()
            self.output_window.activateWindow()
            if code != 0:
                QMessageBox.warning(self, "Не удалось загрузить изменения", report_text.splitlines()[0])


def main() -> int:
    parser = argparse.ArgumentParser(description="Upload XML/Excel parameter changes to local containers.")
    parser.add_argument("--gui", dest="gui", action="store_true", help="Run GUI")
    parser.add_argument("--file", dest="file_path", default="", help="Path to change list file (XML or Excel)")
    args = parser.parse_args()

    if args.gui or (PYSIDE and len(sys.argv) == 1):
        if not PYSIDE:
            print("PySide6 is not available. Install it or run without --gui.")
            return 1
        app = QApplication(sys.argv)
        try:
            if os.path.exists(LOGO_ICO_PATH):
                app.setWindowIcon(QIcon(LOGO_ICO_PATH))
        except Exception:
            pass
        if apply_app_style is not None:
            try:
                apply_app_style(app, theme="light", icon_dir=ICON_DIR)
            except Exception:
                pass
        w = MainWindow()
        w.show()
        return app.exec()

    file_path = args.file_path.strip()
    if not file_path:
        file_path = input("Файл (XML или Excel): ").strip()
    if not file_path or not os.path.exists(file_path):
        print("Файл не найден.")
        return 1

    try:
        file_format = detect_file_format(file_path)
    except ValueError as e:
        print(e)
        return 1

    try:
        if file_format == "excel":
            changes, param_defs = read_excel_payload(file_path)
        else:
            changes = read_xml_changes(file_path)
            param_defs = []
        if param_defs:
            print(f"Loaded {len(param_defs)} parameter definitions from Excel.")
        else:
            print(f"Loaded {len(changes)} changes from file.")
    except Exception as e:
        print(f"Failed to read file: {e}")
        return 1

    projects = load_local_projects()
    if not projects:
        print("Проекты не найдены.")
        return 1

    print("Проекты:")
    project_id, project_name = choose_item(projects, "title", "id")
    print(f"Выбран проект: {project_name} ({project_id})")

    containers = load_local_containers(project_id)
    if not containers:
        print("Контейнеры не найдены.")
        return 1

    print("Контейнеры:")
    selected_ids = choose_items(containers, "title", "id")
    selected_containers = [(c.get("id"), c.get("title", "")) for c in containers if c.get("id") in selected_ids]

    code, report_text, _ = run_upload(file_path, selected_containers)
    print(report_text)

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
