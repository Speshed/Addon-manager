# -*- coding: utf-8 -*-
"""Read Larix adapter Excel sheets without depending on a UI module."""
from __future__ import annotations

from dataclasses import dataclass
from typing import Iterable

import openpyxl


@dataclass(frozen=True)
class AdapterParameter:
    code: str
    name: str
    group: str
    is_numeric: bool
    source_key: str = ""


def list_adapter_sheets(path: str) -> list[str]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def _text(value) -> str:
    return "" if value is None else str(value).strip()


def _norm(value) -> str:
    return _text(value).lower().replace("ё", "е")


def _find_header(ws):
    wanted = {
        "group": {"группа параметров"},
        "name": {"наименование параметра"},
        "type": {"тип параметра"},
        "source_key": {"параметры"},
        "source_list": {"список параметров"},
    }
    max_row = min(int(ws.max_row or 0), 40)
    for row_idx in range(1, max_row + 1):
        found: dict[str, int] = {}
        for col_idx in range(1, int(ws.max_column or 0) + 1):
            value = _norm(ws.cell(row=row_idx, column=col_idx).value)
            for key, variants in wanted.items():
                if value in variants:
                    found[key] = col_idx
        if "name" in found and ("group" in found or "source_key" in found or "source_list" in found):
            return row_idx, found
    return None, {}


def _legacy_group_name(ws, header_row: int) -> str:
    for row_idx in range(1, min(header_row, 40) + 1):
        for col_idx in range(1, int(ws.max_column or 0) + 1):
            if _norm(ws.cell(row=row_idx, column=col_idx).value) == "укажите группу параметров:":
                return _text(ws.cell(row=row_idx, column=col_idx + 1).value)
    return ""


def read_adapter_parameters(path: str, sheet_name: str) -> list[AdapterParameter]:
    """Return target parameters created by an adapter sheet.

    Supports the current common template (Group/Name/Type/...) and the legacy
    layout where the group can be written above the table.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Лист '{sheet_name}' не найден")
        ws = wb[sheet_name]
        header_row, cols = _find_header(ws)
        if not header_row:
            raise ValueError("Не найдены заголовки адаптера: 'Группа параметров' / 'Наименование параметра'.")

        legacy_group = _legacy_group_name(ws, header_row)
        result: list[AdapterParameter] = []
        seen: set[str] = set()

        for row_idx in range(header_row + 1, int(ws.max_row or 0) + 1):
            name = _text(ws.cell(row=row_idx, column=cols["name"]).value) if cols.get("name") else ""
            group = _text(ws.cell(row=row_idx, column=cols["group"]).value) if cols.get("group") else ""
            param_type = _text(ws.cell(row=row_idx, column=cols["type"]).value) if cols.get("type") else ""
            source_key = _text(ws.cell(row=row_idx, column=cols["source_key"]).value) if cols.get("source_key") else ""

            if not name:
                continue
            if not group:
                group = legacy_group

            code = f"{group}.{name}" if group else name
            code = code.strip(".").strip()
            if not code or code.casefold() in seen:
                continue

            seen.add(code.casefold())
            result.append(
                AdapterParameter(
                    code=code,
                    name=name,
                    group=group,
                    is_numeric="чис" in param_type.casefold(),
                    source_key=source_key,
                )
            )
        return result
    finally:
        wb.close()


def read_adapter_mapping(path: str, sheet_name: str) -> dict[str, dict]:
    """Return legacy Excel-column -> Larix parameter mapping used by Validator."""
    mapping: dict[str, dict] = {}
    for item in read_adapter_parameters(path, sheet_name):
        key = item.source_key or item.name
        key = key.strip()
        if not key:
            continue
        mapping[key] = {"code": item.code, "isNumeric": bool(item.is_numeric)}
    return mapping




def find_adapter_mapping_duplicates(path: str, sheet_name: str) -> dict[str, list[str]]:
    """Return duplicate LOI mapping keys and target parameter codes.

    The same rule is used as in ``read_adapter_mapping``: the key comes from
    the ``Параметры`` column, with ``Наименование параметра`` as a fallback.
    """
    grouped: dict[str, list[str]] = {}
    for item in read_adapter_parameters(path, sheet_name):
        key = (item.source_key or item.name).strip()
        if not key:
            continue
        grouped.setdefault(key, []).append(item.code)

    duplicates: dict[str, list[str]] = {}
    for key, codes in grouped.items():
        unique_codes: list[str] = []
        seen_codes: set[str] = set()
        for code in codes:
            marker = str(code).casefold()
            if marker in seen_codes:
                continue
            seen_codes.add(marker)
            unique_codes.append(str(code))
        if len(unique_codes) > 1:
            duplicates[key] = unique_codes
    return duplicates

def prefer_adapter_sheet(sheet_names: Iterable[str]) -> str:
    names = [str(name) for name in sheet_names]
    for name in names:
        if name.strip().casefold() == "адаптер":
            return name
    return names[0] if names else ""
