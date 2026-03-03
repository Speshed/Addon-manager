# -*- coding: utf-8 -*-
from __future__ import annotations

import json
import os
import re
from typing import Dict, Iterable, List, Tuple, Any

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except Exception as exc:
    raise RuntimeError("openpyxl is required to export Excel.") from exc

COMMON_JSON = r'''
{
"Наборы для матриц":[
  {"Имя набора": "АР_Кладка", "Описание": "ОС.СП.1.1.1, ОС.СП.1.1.2, ОС.СП.3.1.1, ОС.СП.3.1.2, ОС.СП.3.1.3"},
  {"Имя набора": "АР_Перегородки", "Описание": "ОС.СП.3.2"},
  {"Имя набора": "АР_Фасад", "Описание": "ОС.ФС.1, ОС.ФС.2, ОС.ФС.3"},
  {"Имя набора": "АР_Витражи", "Описание": "ОС.ЗП.3.1, ОС.ЗП.3.2, ОС.ЗП.3.3"},
  {"Имя набора": "АР_Кровля", "Описание": "ОС.КР.1, ОС.КР.2.1, ОС.КР.2.2, ОС.КР.2.3, ОС.КР.2.4, ОС.КР.2.5, ОС.КР.2.6"},
  {"Имя набора": "КР_Фундаменты", "Описание": "ОС.КР.1, ОС.КР.2.1, ОС.КР.2.2, ОС.КР.2.3, ОС.КР.2.4, ОС.КР.2.5, ОС.КР.2.6"},
  {"Имя набора": "КР_Стены", "Описание": "ОС.ФС.1, ОС.ФС.2, ОС.ФС.3"}
],
"Матрица":{
  "группы": ["АР", "КР"],
  "наборы": {
    "АР": [
      {"имя": "АР_Кладка", "код": "1.01"},
      {"имя": "АР_Перегородки", "код": "1.02"},
      {"имя": "АР_Фасад", "код": "1.03"}
    ],
    "КР": [
      {"имя": "КР_Фундаменты", "код": "2.01"},
      {"имя": "КР_Стены", "код": "2.02"}
    ]
  },
  "данные": [
    {"код": "1.01", "имя": "АР_Кладка", "значения": {"АР_Кладка": "D/B", "АР_Перегородки": "B", "АР_Фасад": "A", "КР_Фундаменты": "B", "КР_Стены": "C"}},
    {"код": "1.02", "имя": "АР_Перегородки", "значения": {"АР_Кладка": null, "АР_Перегородки": "D/C", "АР_Фасад": "B", "КР_Фундаменты": "A", "КР_Стены": "B"}},
    {"код": "1.03", "имя": "АР_Фасад", "значения": {"АР_Кладка": null, "АР_Перегородки": null, "АР_Фасад": "D/B", "КР_Фундаменты": "C", "КР_Стены": "B"}},
    {"код": "2.01", "имя": "КР_Фундаменты", "значения": {"АР_Кладка": null, "АР_Перегородки": null, "АР_Фасад": null, "КР_Фундаменты": "D/C", "КР_Стены": "B"}},
    {"код": "2.02", "имя": "КР_Стены", "значения": {"АР_Кладка": null, "АР_Перегородки": null, "АР_Фасад": null, "КР_Фундаменты": null, "КР_Стены": "D/B"}}
  ]
},
"Параметры":{
  "loi_заголовки": ["Корпус", "Секция", "Комплект документации", "Этаж", "Наименование элемента"],
  "данные": [
    {"тип": "подсекция", "название": "Раздел КР - Стадия Р"},
    {"тип": "подсекция", "название": "03_Конструкции строений"},
    {"тип": "подсекция", "название": "03.01_Фундамент ж/б"},
    {"тип": "данные", "Элементы модели": "03.01.01.01_Фундаментная плита монолитная", "Категория Revit": "Фундамент несущей конструкции", "Пример Класса IFC": "IfcFooting", "Код по классификатору": "03.01.01.01", "Описание по классификатору": "Фундаментная плита железобетонная монолитная", "LOI": ["+", "+", "+", "+", "+"]},
    {"тип": "подсекция", "название": "03.03_Вертикальные ж/б конструкции"},
    {"тип": "данные", "Элементы модели": "03.03.01.01_Стены ж/б монолитные", "Категория Revit": "Стены", "Пример Класса IFC": "IfcWall", "Код по классификатору": "03.03.01.01", "Описание по классификатору": "Стена железобетонная монолитная", "LOI": ["+", null, "+", "+", null]},
    {"тип": "данные", "Элементы модели": "03.03.01.03_Колонны ж/б монолитные", "Категория Revit": "Несущие колонны", "Пример Класса IFC": "IfcColumn", "Код по классификатору": "03.03.01.03", "Описание по классификатору": "Колонна железобетонная монолитная", "LOI": ["+", "+", "+", "+", "+"]},
    {"тип": "подсекция", "название": "04_Инженерные сети"},
    {"тип": "подсекция", "название": "04.02_Системы вентиляции"},
    {"тип": "данные", "Элементы модели": "04.02.01_Воздуховоды", "Категория Revit": "Воздуховоды", "Пример Класса IFC": "IfcDuctSegment", "Код по классификатору": "04.02.01.01, 04.02.01.02, 04.02.01.07", "Описание по классификатору": "Воздуховоды систем вентиляции", "LOI": ["+", null, "+", null, null]},
    {"тип": "данные", "Элементы модели": "04.02.03_Материалы изоляции воздуховодов", "Категория Revit": "Изоляция воздуховодов", "Пример Класса IFC": "IfcCovering", "Код по классификатору": "04.02.03.01", "Описание по классификатору": "Изоляционные материалы для воздуховодов", "LOI": ["+", "+", "+", "+", "+"]},
    {"тип": "данные", "Элементы модели": "04.02.04_Воздухораспределительные устройства", "Категория Revit": "Воздухораспределители", "Пример Класса IFC": "IfcAirTerminal", "Код по классификатору": "04.02.04.03, 04.02.04.04", "Описание по классификатору": "Воздухораспределительные устройства", "LOI": ["+", "+", null, "+", null]},
    {"тип": "данные", "Элементы модели": "04.02.05_Арматура воздуховодов", "Категория Revit": "Арматура воздуховодов", "Пример Класса IFC": "IfcDuctFitting", "Код по классификатору": "04.02.05.01, 04.02.05.03, 04.02.05.05", "Описание по классификатору": "Арматура воздуховодов", "LOI": ["+", null, null, "+", "+"]}
  ]
},
"Адаптер":[
  {"Укажите группу параметров:": "Наименование параметра", "IBIM": "Тип параметра", "Список параметров": "Список параметров", "Замена с": "Замена с", "Замена на": "Замена на", "Параметры": "Параметры"},
  {"Укажите группу параметров:": "Корпус", "IBIM": "Текст", "Список параметров": "\"ADSK_Номер корпуса\", \"MARKS_Номер корпуса\"", "Замена с": "\"Корпус 1\", \"Корпус 2\"", "Замена на": "\"1\", \"2\"", "Параметры": "Корпус"},
  {"Укажите группу параметров:": "Секция", "IBIM": "Текст", "Список параметров": "\"ADSK_Номер секции\", \"ADSK_Номер секции\"", "Замена с": null, "Замена на": null, "Параметры": "Секция"},
  {"Укажите группу параметров:": "Комплект_документации", "IBIM": "Текст", "Список параметров": "\"ADSK_Комплект чертежей\"", "Замена с": null, "Замена на": null, "Параметры": "Комплект документации"},
  {"Укажите группу параметров:": "Этаж", "IBIM": "Текст", "Список параметров": "\"ADSK_Этаж\"", "Замена с": null, "Замена на": null, "Параметры": "Этаж"},
  {"Укажите группу параметров:": "Этаж_12", "IBIM": "Текст", "Список параметров": "\"ADSK_Этаж\"", "Замена с": null, "Замена на": null, "Параметры": "Этаж"}
]
}
'''


def export_common_excel(json_path: str, output_path: str) -> None:
    data = _load_json(json_path)

    thin = Side(style="thin", color="000000")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)
    base_font = Font(name="Tahoma", size=11)
    bold_font = Font(name="Tahoma", size=11, bold=False)
    white_bold_font = Font(name="Tahoma", size=11, bold=False, color="FFFFFF")

    fill_blue_header = PatternFill("solid", fgColor="8EA9DB")
    fill_gray = PatternFill("solid", fgColor="D9D9D9")
    fill_light_gray = PatternFill("solid", fgColor="F2F2F2")
    fill_orange = PatternFill("solid", fgColor="F4B183")
    fill_yellow = PatternFill("solid", fgColor="FFD966")
    fill_light_blue = PatternFill("solid", fgColor="DDEBF7")
    fill_a = PatternFill("solid", fgColor="2F75B5")
    fill_b = PatternFill("solid", fgColor="5B9BD5")
    fill_c = PatternFill("solid", fgColor="9DC3E6")

    align_center = Alignment(horizontal="center", vertical="center")
    align_left_top = Alignment(horizontal="left", vertical="top", wrap_text=True)
    align_left_center = Alignment(horizontal="left", vertical="center", wrap_text=True)

    def _apply_base_font(ws, min_row, max_row, min_col, max_col):
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                cell = ws.cell(r, c)
                f = cell.font
                if f is None:
                    cell.font = base_font
                    continue
                new_font = f.copy()
                new_font.name = "Tahoma"
                new_font.sz = 11
                new_font.bold = False
                cell.font = new_font

    def _apply_wrap(ws, min_row, max_row, min_col, max_col):
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                cell = ws.cell(r, c)
                a = cell.alignment
                if a is None:
                    cell.alignment = Alignment(wrap_text=True)
                    continue
                new_align = a.copy()
                new_align.wrap_text = True
                cell.alignment = new_align

    def _apply_border(ws, min_row, max_row, min_col, max_col):
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                ws.cell(r, c).border = border

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    try:
        normal_style = None
        for ns in wb._named_styles:
            if getattr(ns, "name", "") == "Normal":
                normal_style = ns
                break
        if normal_style is not None:
            normal_style.font = base_font
        else:
            from openpyxl.styles import NamedStyle
            wb.add_named_style(NamedStyle(name="Normal", font=base_font))
    except Exception:
        pass

    rows = data.get("Наборы для матриц", [])
    if rows:
        ws = wb.create_sheet("Наборы для матриц")
        headers = ["Имя набора", "Описание"]
        for col, name in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=name)
            cell.font = bold_font
            cell.fill = fill_blue_header
            cell.alignment = align_center
        for r_idx, row in enumerate(rows, start=2):
            ws.cell(row=r_idx, column=1, value=row.get("Имя набора", "")).alignment = align_left_top
            ws.cell(row=r_idx, column=2, value=row.get("Описание", "")).alignment = align_left_top
        ws.column_dimensions["A"].width = 32
        ws.column_dimensions["B"].width = 90
        ws.row_dimensions[1].height = 20
        _apply_border(ws, 1, 1 + len(rows), 1, 2)
        _apply_base_font(ws, 1, ws.max_row, 1, ws.max_column)
        _apply_wrap(ws, 1, ws.max_row, 1, ws.max_column)

    matrix_data = data.get("Матрица", {})
    if matrix_data:
        ws = wb.create_sheet("Матрица")
        группы = matrix_data.get("группы", [])
        наборы_by_group = matrix_data.get("наборы", {})
        данные = matrix_data.get("данные", [])

        all_nabory = []
        for g in группы:
            all_nabory.extend(наборы_by_group.get(g, []))

        ws.merge_cells(start_row=1, start_column=1, end_row=3, end_column=2)
        cell = ws.cell(1, 1, "Наборы элементов")
        cell.font = bold_font
        cell.fill = fill_light_gray
        cell.alignment = align_center

        col_idx = 3
        for g in группы:
            nabory = наборы_by_group.get(g, [])
            if not nabory:
                continue
            start_col = col_idx
            for _ in nabory:
                col_idx += 1
            end_col = col_idx - 1
            ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
            cell = ws.cell(1, start_col, value=g)
            cell.font = bold_font
            cell.fill = fill_gray
            cell.alignment = align_center

        col_idx = 3
        for g in группы:
            for n in наборы_by_group.get(g, []):
                cell = ws.cell(2, col_idx, value=n.get("имя", ""))
                cell.font = bold_font
                cell.fill = fill_light_gray
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True, text_rotation=90)
                cell = ws.cell(3, col_idx, value=n.get("код", ""))
                cell.font = bold_font
                cell.fill = fill_gray
                cell.alignment = align_center
                col_idx += 1

        for r_idx, row_data in enumerate(данные, start=4):
            ws.cell(r_idx, 1, row_data.get("код", "")).alignment = align_center
            ws.cell(r_idx, 2, row_data.get("имя", "")).alignment = align_left_center
            значения = row_data.get("значения", {})
            col_idx = 3
            for g in группы:
                for n in наборы_by_group.get(g, []):
                    val = значения.get(n.get("имя", ""), "")
                    cell = ws.cell(r_idx, col_idx, val if val is not None else "")
                    cell.alignment = align_center
                    if isinstance(val, str) and val.strip():
                        tag = val.strip()[0].upper()
                        if tag == "A":
                            cell.fill = fill_a
                            cell.font = white_bold_font
                        elif tag == "B":
                            cell.fill = fill_b
                            cell.font = white_bold_font
                        elif tag == "C":
                            cell.fill = fill_c
                            cell.font = bold_font
                    col_idx += 1

        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 36
        col_idx = 3
        for g in группы:
            for n in наборы_by_group.get(g, []):
                col_letter = get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = max(4, len(n.get("код", "")) + 1)
                col_idx += 1
        ws.row_dimensions[1].height = 20
        ws.row_dimensions[2].height = 90.75
        ws.row_dimensions[3].height = 28.5
        total_cols = 2 + len(all_nabory)
        _apply_border(ws, 1, 3 + len(данные), 1, total_cols)
        _apply_base_font(ws, 1, ws.max_row, 1, ws.max_column)
        _apply_wrap(ws, 1, ws.max_row, 1, ws.max_column)

    params_data = data.get("Параметры", {})
    if params_data:
        ws = wb.create_sheet("Параметры")
        loi_headers = params_data.get("loi_заголовки", [])
        rows_data = params_data.get("данные", [])
        total_cols = 5 + len(loi_headers)

        main_headers = ["Элементы модели", "Категория Revit", "Пример Класса IFC", "Код по классификатору", "Описание по классификатору"]
        for col, name in enumerate(main_headers, start=1):
            cell = ws.cell(1, col, value=name)
            cell.font = bold_font
            cell.fill = fill_gray
            cell.alignment = align_center

        if loi_headers:
            ws.merge_cells(start_row=1, start_column=6, end_row=1, end_column=total_cols)
            cell = ws.cell(1, 6, "LOI")
            cell.font = bold_font
            cell.fill = fill_orange
            cell.alignment = align_center

            for idx, h in enumerate(loi_headers, start=6):
                cell = ws.cell(2, idx, value=h)
                cell.font = bold_font
                cell.fill = fill_orange
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False, text_rotation=90)
            ws.row_dimensions[2].height = 180

        out_row = 3
        for row_data in rows_data:
            row_type = row_data.get("тип", "данные")
            if row_type == "подсекция":
                loi_vals = row_data.get("LOI", [])
                if loi_vals:
                    ws.merge_cells(start_row=out_row, start_column=1, end_row=out_row, end_column=5)
                cell = ws.cell(out_row, 1, row_data.get("название", ""))
                cell.font = bold_font
                cell.fill = fill_light_gray
                cell.alignment = align_left_center
                for idx, val in enumerate(loi_vals):
                    col = 6 + idx
                    cell = ws.cell(out_row, col, val if val is not None else "")
                    cell.alignment = align_center
                    if isinstance(val, str) and val.strip() == "+":
                        cell.fill = fill_yellow
            else:
                ws.cell(out_row, 1, row_data.get("Элементы модели", "")).alignment = align_left_top
                ws.cell(out_row, 2, row_data.get("Категория Revit", "")).alignment = align_left_top
                ws.cell(out_row, 3, row_data.get("Пример Класса IFC", "")).alignment = align_left_top
                ws.cell(out_row, 4, row_data.get("Код по классификатору", "")).alignment = align_left_top
                ws.cell(out_row, 5, row_data.get("Описание по классификатору", "")).alignment = align_left_top
                loi_vals = row_data.get("LOI", [])
                for idx, val in enumerate(loi_vals):
                    col = 6 + idx
                    cell = ws.cell(out_row, col, val if val is not None else "")
                    cell.alignment = align_center
                    if isinstance(val, str) and val.strip() == "+":
                        cell.fill = fill_yellow
            out_row += 1

        ws.column_dimensions["A"].width = 40
        ws.column_dimensions["B"].width = 35
        ws.column_dimensions["C"].width = 30
        ws.column_dimensions["D"].width = 35
        ws.column_dimensions["E"].width = 90
        for idx in range(len(loi_headers)):
            ws.column_dimensions[get_column_letter(6 + idx)].width = 4
        _apply_border(ws, 1, out_row - 1, 1, total_cols)
        _apply_base_font(ws, 1, ws.max_row, 1, ws.max_column)
        _apply_wrap(ws, 1, ws.max_row, 1, ws.max_column)

    adapter_rows = data.get("Адаптер", [])
    if adapter_rows:
        ws = wb.create_sheet("Адаптер")
        header_row = adapter_rows[0] if adapter_rows else {}
        header_keys = list(header_row.keys())
        if header_keys:
            cell = ws.cell(1, 1, header_keys[0])
            cell.font = bold_font
            cell = ws.cell(1, 2, "IBIM")
            cell.font = bold_font

            for col, key in enumerate(header_keys, start=1):
                cell = ws.cell(2, col, header_row.get(key, ""))
                cell.font = white_bold_font
                cell.fill = fill_blue_header
                cell.alignment = align_center

            out_row = 3
            for idx, row in enumerate(adapter_rows[1:], start=0):
                for col, key in enumerate(header_keys, start=1):
                    cell = ws.cell(out_row, col, row.get(key, ""))
                    cell.alignment = align_left_top
                    if idx % 2 == 0:
                        cell.fill = fill_light_blue
                out_row += 1

            ws.column_dimensions["A"].width = 32
            ws.column_dimensions["B"].width = 14
            ws.column_dimensions["C"].width = 50
            ws.column_dimensions["D"].width = 24
            ws.column_dimensions["E"].width = 20
            ws.column_dimensions["F"].width = 25
            _apply_border(ws, 2, out_row - 1, 1, max(1, len(header_keys)))
            _apply_base_font(ws, 1, ws.max_row, 1, ws.max_column)
            _apply_wrap(ws, 1, ws.max_row, 1, ws.max_column)

    wb.save(output_path)


def _load_json(path: str) -> Dict[str, Any]:
    data = json.loads(COMMON_JSON)
    if not isinstance(data, dict):
        raise RuntimeError("Invalid JSON structure.")
    return data


if __name__ == "__main__":
    base_dir = os.path.dirname(os.path.abspath(__file__))
    json_path = os.path.join(base_dir, "Общий файл.json")
    output_path = os.path.join(base_dir, "Шаблон Excel.xlsx")
    export_common_excel(json_path, output_path)
